import http.client
import json
import uuid

from django.shortcuts import render,redirect, get_object_or_404
from django.http import HttpResponse,JsonResponse
from django.contrib import messages
from django.db.models import Avg, Q, Sum, Count
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from .models import Usuario,Comentario,MensajeQueja,Reserva,Servicio,Empresa,EmpresaServicio, ReservaServicio, Plan, SuscripcionUsuario, HistorialPagosSuscripcion, PlanEmpresarial, SuscripcionEmpresarial, SolicitudServicioEmpresa, HistorialPagosSuscripcionEmpresarial, SolicitudContactoPlan, PlanServicio, PeriodoLiquidacion, DetalleLiquidacion, ConfiguracionPagos
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from .forms import ComentarioForm, ReservaForm, UsuariosForm,ComentarioClienteForm,QuejaForm,ServicioForm,EmpresaForm,ProfileUserForm,EmpresaRegistroForm, EmpresaPerfilForm, AdminProfileForm, SolicitudContactoPlanForm
from datetime import datetime,timedelta
import csv
from django.utils import timezone
from functools import wraps
from django.core.mail import send_mail
from django.conf import settings
from django.urls import reverse
from django.contrib import messages


# Decorador personalizado para verificar autenticación de empresa
def empresa_required(function):
    @wraps(function)
    def wrap(request, *args, **kwargs):
        print(f"🔍 @empresa_required verificando acceso a {function.__name__}")
        
        # Verificar si la empresa está autenticada en la sesión
        if not request.session.get('es_empresa', False):
                print(f"❌ Empresa no autenticada")
                # Si la petición es AJAX, devolver JSON en lugar de redirigir
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Empresa no autenticada'}, status=403)
                messages.error(request, 'Debes iniciar sesión como empresa para acceder a esta sección.')
                return redirect('logincrud')
        
        empresa_id = request.session.get('empresa_id')
        if not empresa_id:
            print(f"❌ No se encontró ID de empresa en sesión")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Sesión de empresa inválida'}, status=403)
            messages.error(request, 'Error: Sesión de empresa inválida.')
            return redirect('logincrud')
        
        try:
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            print(f"✅ Empresa autenticada: {empresa.nombre_empresa}")
            return function(request, *args, **kwargs)
        except Empresa.DoesNotExist:
            print(f"❌ Empresa no encontrada con ID: {empresa_id}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Empresa no encontrada'}, status=404)
            messages.error(request, 'Error: Empresa no encontrada.')
            return redirect('logincrud')
    return wrap


# Decorador personalizado para verificar autenticación de usuario regular
def usuario_required(function):
    @wraps(function)
    def wrap(request, *args, **kwargs):
        print(f"🔍 @usuario_required verificando acceso a {function.__name__}")
        
        # Verificar si el usuario está autenticado
        if not request.user.is_authenticated:
            print(f"❌ Usuario no autenticado")
            messages.error(request, 'Debes iniciar sesión para acceder a esta sección.')
            return redirect('logincrud')
        
        # Verificar que no sea una empresa autenticada
        if request.session.get('es_empresa', False):
            print(f"❌ Intento de acceso de empresa a sección de usuario")
            messages.error(request, 'Esta sección es solo para usuarios. Por favor, cambia a tu cuenta de usuario.')
            return redirect('logincrud')
        
        print(f"✅ Usuario autenticado: {request.user.username}")
        return function(request, *args, **kwargs)
    return wrap


# Decorador simple para vistas de administrador (evita errores si no existía)
def admin_required(function):
    @wraps(function)
    def wrap(request, *args, **kwargs):
        # Intentar verificar sesión o usuario staff
        if request.session.get('es_admin') or getattr(request.user, 'is_staff', False):
            return function(request, *args, **kwargs)
        messages.error(request, 'Debes ser administrador para acceder a esta sección.')
        return redirect('logincrud')
    return wrap


# Función auxiliar para enviar correo de confirmación de reserva profesional
def enviar_correo_confirmacion_reserva(usuario, empresa, servicios, fecha, hora, precio_total, numero_reserva=None):
    import logging
    from django.conf import settings  # Importar al inicio para evitar UnboundLocalError
    logger = logging.getLogger('lavado_auto.views')
    
    try:
        import uuid
        from datetime import datetime
        
        logger.info(f"🔄 Iniciando envío de correo de confirmación de reserva")
        
        nombre_usuario = getattr(usuario, 'nombre_completo', getattr(usuario, 'nombre_usuario', 'Cliente'))
        correo_destino = getattr(usuario, 'correo', getattr(usuario, 'email', None))
        
        logger.info(f"📧 Usuario: {nombre_usuario}, Email: {correo_destino}")
        
        if not correo_destino:
            logger.error(f"❌ No se encontró email para el usuario {nombre_usuario}")
            print(f"❌ No se encontró email para el usuario {nombre_usuario}")
            return False

        # Si no se recibió un número de reserva (pasado por la vista), generar uno temporal
        if not numero_reserva:
            numero_reserva = f"ANW-{str(uuid.uuid4())[:8].upper()}"
        fecha_actual = datetime.now().strftime("%d de %B de %Y")
        
        # Obtener información de la empresa
        # Preferir el campo `nombre_empresa` (modelo Empresa usa ese nombre); fallback a `nombre` por compatibilidad
        nombre_empresa = getattr(empresa, 'nombre_empresa', None) or getattr(empresa, 'nombre', 'AutoNew')
        # Direccion y telefono: usar los campos disponibles en el modelo con fallback a cadena vacía
        direccion_empresa = getattr(empresa, 'direccion', '') or getattr(empresa, 'direccion_empresa', '')
        telefono_empresa = getattr(empresa, 'telefono', '') or getattr(empresa, 'telefono_empresa', '')

        # Construir lista de servicios profesional
        servicios_html = ""
        servicios_text = ""
        for i, servicio in enumerate(servicios, 1):
            nombre_servicio = servicio.nombre_servicio if hasattr(servicio, 'nombre_servicio') else getattr(servicio, 'nombre', 'Servicio')
            precio_servicio = getattr(servicio, 'precio', '0')
            
            servicios_html += f"""
            <tr style="border-bottom: 1px solid #e5e7eb;">
                <td style="padding: 12px 0; color: #374151; font-weight: 500;">{i}. {nombre_servicio}</td>
                <td style="padding: 12px 0; text-align: right; color: #374151; font-weight: 600;">${precio_servicio}</td>
            </tr>"""
            
            servicios_text += f"{i}. {nombre_servicio} - ${precio_servicio}\n"

        # HTML profesional con diseño responsive
        html_message = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Confirmación de Reserva - AutoNew</title>
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f8fafc; line-height: 1.6;">
            <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);">
                
                <!-- Header -->
                <div style="background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%); padding: 30px 20px; text-align: center;">
                    <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 700;">
                        🚗 AutoNew
                    </h1>
                    <p style="color: #bfdbfe; margin: 8px 0 0 0; font-size: 16px;">
                        Cuidamos tu vehículo con profesionalismo
                    </p>
                </div>

                <!-- Confirmación -->
                <div style="padding: 30px 20px; text-align: center; background-color: #f0fdf4; border-bottom: 3px solid #22c55e;">
                    <div style="display: inline-block; background-color: #22c55e; color: white; padding: 12px 24px; border-radius: 50px; font-weight: 600; margin-bottom: 15px;">
                        ✓ RESERVA CONFIRMADA
                    </div>
                    <h2 style="color: #15803d; margin: 10px 0 5px 0; font-size: 24px;">
                        ¡Hola {nombre_usuario}!
                    </h2>
                    <p style="color: #166534; margin: 0; font-size: 16px;">
                        Tu reserva ha sido procesada exitosamente
                    </p>
                </div>

                <!-- Nombre de la Empresa (destacado) -->
                <div style="text-align: center; padding: 18px 20px;">
                    <h3 style="margin: 0; font-size: 20px; color: #0c4a6e; font-weight: 700;">{nombre_empresa}</h3>
                </div>

                <!-- Información de la Reserva -->
                <div style="padding: 30px 20px;">
                    <div style="background-color: #f1f5f9; border-left: 4px solid #3b82f6; padding: 20px; border-radius: 0 8px 8px 0; margin-bottom: 25px;">
                        <h3 style="color: #1e40af; margin: 0 0 15px 0; font-size: 18px;">
                            📋 Detalles de tu Cita
                        </h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0; color: #64748b; font-weight: 500; width: 40%;">Número de Reserva:</td>
                                <td style="padding: 8px 0; color: #1e293b; font-weight: 700; font-size: 16px;">{numero_reserva}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b; font-weight: 500;">Fecha:</td>
                                <td style="padding: 8px 0; color: #1e293b; font-weight: 600;">{fecha}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b; font-weight: 500;">Hora:</td>
                                <td style="padding: 8px 0; color: #1e293b; font-weight: 600;">{hora}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0; color: #64748b; font-weight: 500;">Empresa:</td>
                                    <td style="padding: 8px 0; color: #1e293b; font-weight: 600;">{nombre_empresa}</td>
                            </tr>
                        </table>
                    </div>

                    <!-- Servicios Contratados -->
                    <div style="margin-bottom: 25px;">
                        <h3 style="color: #1e40af; margin: 0 0 15px 0; font-size: 18px;">
                            🛠️ Servicios Contratados
                        </h3>
                        <table style="width: 100%; border-collapse: collapse; background-color: #ffffff; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden;">
                            {servicios_html}
                            <tr style="background-color: #f8fafc; border-top: 2px solid #3b82f6;">
                                <td style="padding: 15px 0; color: #1e40af; font-weight: 700; font-size: 16px;">TOTAL A PAGAR:</td>
                                <td style="padding: 15px 0; text-align: right; color: #1e40af; font-weight: 700; font-size: 18px;">${precio_total}</td>
                            </tr>
                        </table>
                    </div>

                    <!-- Información Importante -->
                    <div style="background-color: #fef3c7; border: 1px solid #f59e0b; border-radius: 8px; padding: 20px; margin-bottom: 25px;">
                        <h3 style="color: #92400e; margin: 0 0 15px 0; font-size: 16px;">
                            ⚠️ Información Importante
                        </h3>
                        <ul style="color: #78350f; margin: 0; padding-left: 20px; font-size: 14px;">
                            <li style="margin-bottom: 8px;">Llega <strong>10 minutos antes</strong> de tu cita programada</li>
                            <li style="margin-bottom: 8px;">Trae tu vehículo <strong>preparado</strong> (retira objetos personales)</li>
                            <li style="margin-bottom: 8px;">Presenta este correo o el <strong>número de reserva</strong></li>
                            <li style="margin-bottom: 8px;">En caso de cancelación, hazlo con <strong>24 horas de anticipación</strong></li>
                        </ul>
                        {f'<div style="margin-top:12px;color: #075985; font-weight: 600;">📍 UBICACIÓN: {direccion_empresa}</div>' if direccion_empresa else ''}
                        {f'<div style="margin-top:6px;color: #075985; font-weight: 600;">📞 TELÉFONO: {telefono_empresa}</div>' if telefono_empresa else ''}
                    </div>

                    <!-- Ubicación -->
                    {f'''<div style="background-color: #f0f9ff; border: 1px solid #0ea5e9; border-radius: 8px; padding: 20px; margin-bottom: 25px;">
                        <h3 style="color: #0c4a6e; margin: 0 0 10px 0; font-size: 16px;">
                            📍 Ubicación
                        </h3>
                        <p style="color: #075985; margin: 0; font-weight: 500;">{direccion_empresa}</p>
                        {f'<p style="color: #075985; margin: 5px 0 0 0;">📞 {telefono_empresa}</p>' if telefono_empresa else ''}
                    </div>''' if direccion_empresa else ''}

                    <!-- Botones de Acción -->
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="#" style="display: inline-block; background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: 600; margin: 0 10px 10px 0; font-size: 16px;">
                            Ver Mis Reservas
                        </a>
                        <a href="#" style="display: inline-block; background-color: #f8fafc; color: #374151; border: 2px solid #d1d5db; padding: 13px 28px; text-decoration: none; border-radius: 8px; font-weight: 600; margin: 0 10px 10px 0; font-size: 16px;">
                            Contactar Soporte
                        </a>
                    </div>
                </div>

                <!-- Footer -->
                <div style="background-color: #1e293b; padding: 25px 20px; text-align: center;">
                    <p style="color: #94a3b8; margin: 0 0 10px 0; font-size: 14px;">
                        Gracias por confiar en AutoNew para el cuidado de tu vehículo
                    </p>
                    <p style="color: #64748b; margin: 0; font-size: 12px;">
                        © {datetime.now().year} AutoNew. Todos los derechos reservados.
                    </p>
                    <div style="margin-top: 15px;">
                        <span style="color: #64748b; font-size: 12px;">
                            Este correo fue enviado el {fecha_actual}
                        </span>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """

        # Versión texto plano profesional
        plain_message = f"""
========================================
    CONFIRMACIÓN DE RESERVA - AUTONEW
========================================

¡Hola {nombre_usuario}!

Tu reserva ha sido confirmada exitosamente.

========== {nombre_empresa} ==========

DETALLES DE LA CITA:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Número de Reserva: {numero_reserva}
• Fecha: {fecha}
• Hora: {hora}
• Empresa: {nombre_empresa}

SERVICIOS CONTRATADOS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{servicios_text}
TOTAL A PAGAR: ${precio_total}

INFORMACIÓN IMPORTANTE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Llega 10 minutos antes de tu cita
• Trae tu vehículo preparado (sin objetos personales)
• Presenta este correo o el número de reserva
• Para cancelaciones, hazlo con 24 horas de anticipación

{f'UBICACIÓN: {direccion_empresa}' if direccion_empresa else ''}
{f'TELÉFONO: {telefono_empresa}' if telefono_empresa else ''}

Gracias por confiar en AutoNew para el cuidado de tu vehículo.

© {datetime.now().year} AutoNew - Enviado el {fecha_actual}
        """

        # Registrar intento de envío
        subject = f'✅ Reserva Confirmada #{numero_reserva} - AutoNew'
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@autonew.com')
        
        logger.info(f"📤 Enviando correo:")
        logger.info(f"   - Asunto: {subject}")
        logger.info(f"   - Desde: {from_email}")
        logger.info(f"   - Para: {correo_destino}")
        logger.info(f"   - Número de reserva: {numero_reserva}")
        
        print(f"📤 Enviando correo de confirmación:")
        print(f"   - Usuario: {nombre_usuario}")
        print(f"   - Email: {correo_destino}")
        print(f"   - Asunto: {subject}")
        print(f"   - Desde: {from_email}")
        
        # Verificar configuración de correo
        if not hasattr(settings, 'EMAIL_HOST_USER') or not settings.EMAIL_HOST_USER:
            logger.error("❌ EMAIL_HOST_USER no está configurado en settings.py")
            print("❌ ERROR: EMAIL_HOST_USER no está configurado en settings.py")
            print("   Por favor configura tu email y contraseña de aplicación de Gmail")
            return False
        
        if settings.EMAIL_HOST_USER == 'tu-email@gmail.com':
            logger.error("❌ EMAIL_HOST_USER tiene el valor por defecto. Debe configurarse con un email real")
            print("❌ ERROR: Debes cambiar 'tu-email@gmail.com' por tu email real en settings.py")
            return False

        send_mail(
            subject=subject,
            message=plain_message,
            from_email=from_email,
            recipient_list=[correo_destino],
            html_message=html_message,
            fail_silently=False,  # Cambiado a False para ver errores
        )
        
        logger.info(f"✅ Correo enviado exitosamente a {correo_destino}")
        print(f"✅ Correo enviado exitosamente a {correo_destino}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Error enviando correo: {str(e)}")
        print(f"❌ Error enviar_correo_confirmacion_reserva: {e}")
        print(f"   Tipo de error: {type(e).__name__}")
        
        # Información adicional para debugging
        if 'authentication' in str(e).lower():
            print("💡 Posible problema de autenticación Gmail:")
            print("   1. Verifica que el email y contraseña sean correctos")
            print("   2. Usa una 'Contraseña de Aplicación' de Gmail, no tu contraseña normal")
            print("   3. Asegúrate de tener habilitada la autenticación en 2 pasos")
        elif 'connection' in str(e).lower():
            print("💡 Posible problema de conexión:")
            print("   1. Verifica tu conexión a internet")
            print("   2. Comprueba la configuración del servidor SMTP")
        
        return False

# Create your views here.



def home(request):
    comentarios = Comentario.objects.all().order_by('-fecha')  # Recupera todos los comentarios
    
    # Paginación para servicios - 8 por página
    servicios_list = Servicio.objects.all().order_by('nombre_servicio')
    servicios_paginator = Paginator(servicios_list, 8)
    servicios_page = request.GET.get('servicios_page', 1)
    
    try:
        servicios = servicios_paginator.page(servicios_page)
    except PageNotAnInteger:
        servicios = servicios_paginator.page(1)
    except EmptyPage:
        servicios = servicios_paginator.page(servicios_paginator.num_pages)
    
    # Paginación para empresas - 4 por página
    empresas_list = Empresa.objects.filter(verificada=True).order_by('nombre_empresa')
    empresas_paginator = Paginator(empresas_list, 4)
    empresas_page = request.GET.get('empresas_page')
    
    try:
        empresas = empresas_paginator.page(empresas_page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        empresas = empresas_paginator.page(1)
    except EmptyPage:
        # Si page está fuera del rango, mostrar la última página
        empresas = empresas_paginator.page(empresas_paginator.num_pages)
    
    return render(request, 'home.html', {
        'comentarios': comentarios, 
        'servicios': servicios,
        'empresas': empresas
    })


@admin_required
@require_POST
def marcar_reservas_empresa(request, empresa_id):
    """Marcar como pagadas las reservas seleccionadas para una empresa.
    Espera un POST con 'reservas' conteniendo IDs separados por comas o múltiples valores.
    Devuelve redirect al detalle de la empresa con un mensaje.
    """
    # Leer lista de ids desde POST (puede venir como reservas[]=1&reservas[]=2 o como texto csv)
    reservas_ids = request.POST.getlist('reservas')
    if not reservas_ids:
        reservas_csv = request.POST.get('reservas_csv', '')
        if reservas_csv:
            reservas_ids = [s.strip() for s in reservas_csv.split(',') if s.strip()]

    if not reservas_ids:
        messages.warning(request, 'No se seleccionaron reservas para marcar como pagadas.')
        return redirect('detalle_pagos_empresa', empresa_id=empresa_id)

    # Filtrar reservas que pertenezcan a la empresa y estén en estado completado y no pagadas a empresa
    qs = Reserva.objects.filter(pk__in=reservas_ids, empresa__id_empresa=empresa_id).filter(Q(estado='completado')).filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True))
    updated = qs.update(pagado_empresa=True)
    messages.success(request, f'Se marcaron {updated} reserva(s) como pagadas a la empresa.')
    return redirect('detalle_pagos_empresa', empresa_id=empresa_id)

def faq(request):
    """Vista para la página de Preguntas Frecuentes (FAQ)"""
    return render(request, 'pages_informativas/faq.html')

def blog(request):
    """Vista para la página de Blog/Noticias"""
    return render(request, 'pages_informativas/blog.html')

def servicios_ajax(request):
    """Vista AJAX para cargar servicios paginados en el home"""
    servicios_list = Servicio.objects.all().order_by('nombre_servicio')
    
    # Paginación para servicios - 8 por página en el home
    paginator = Paginator(servicios_list, 8)
    page = request.GET.get('page', 1)
    
    try:
        servicios = paginator.page(page)
    except PageNotAnInteger:
        servicios = paginator.page(1)
    except EmptyPage:
        servicios = paginator.page(paginator.num_pages)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Respuesta AJAX - renderizar servicios paginados
        return render(request, 'servicios_partial.html', {'servicios': servicios})
    else:
        # Fallback - redireccionar a home si no es AJAX
        return redirect('home')

def empresas_ajax(request):
    """Vista AJAX para cargar empresas paginadas sin recargar la página"""
    empresas_list = Empresa.objects.filter(verificada=True).order_by('nombre_empresa')
    
    # Paginación para empresas - 4 por página
    paginator = Paginator(empresas_list, 4)
    page = request.GET.get('page')
    
    try:
        empresas = paginator.page(page)
    except PageNotAnInteger:
        empresas = paginator.page(1)
    except EmptyPage:
        empresas = paginator.page(paginator.num_pages)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Respuesta AJAX - solo renderizar la sección de empresas
        return render(request, 'empresas_partial.html', {'empresas': empresas})
    else:
        # Fallback - redireccionar a home si no es AJAX
        return redirect('home')

def empresas(request):
    if request.method == 'POST':
        form = EmpresaRegistroForm(request.POST)
        if form.is_valid():
            # Verificar que el email no esté ya registrado
            email = form.cleaned_data['email']
            if Empresa.objects.filter(email=email).exists():
                messages.error(request, 'Ya existe una empresa registrada con este correo electrónico.')
                return render(request, 'empresas.html', {'form': form})
            
            # Crear la empresa
            empresa = form.save(commit=False)
            # Encriptar la contraseña
            empresa.contrasena = make_password(form.cleaned_data['contrasena'])
            # Establecer explícitamente verificada como False (aunque es el default)
            empresa.verificada = False
            empresa.save()
            
            messages.success(request, f'¡Empresa "{empresa.nombre_empresa}" registrada exitosamente! Tu cuenta está pendiente de verificación por nuestro equipo. Bienvenido a AutoNew.')
            # Crear un nuevo formulario vacío después del registro exitoso
            form = EmpresaRegistroForm()
        else:
            messages.error(request, 'Hubo errores en el formulario. Por favor, verifica los datos ingresados.')
    else:
        form = EmpresaRegistroForm()
    
    return render(request, 'empresas/empresas.html', {'form': form})

def login(request):
    # Si ya está logueado, redirigir al home
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        # REGISTRO DE USUARIO
        if 'nombre_completo' in request.POST:
            nombre_completo = request.POST['nombre_completo']
            nombre_usuario = request.POST['nombre_usuario']
            correo = request.POST['correo']
            telefono = request.POST.get('telefono')
            direccion = request.POST.get('direccion')
            contrasena1 = request.POST['contrasena1']
            contrasena2 = request.POST['contrasena2']
            
            # Crear diccionario con los datos del formulario para preservarlos
            form_data = {
                'nombre_completo': nombre_completo,
                'nombre_usuario': nombre_usuario,
                'correo': correo,
                'telefono': telefono,
                'direccion': direccion,
            }
            
            # Validar que las contraseñas coincidan
            if contrasena1 != contrasena2:
                messages.error(request, "Las contraseñas no coinciden, intentalo de nuevo")
                return render(request, 'login.html', {
                    'show_register': True,
                    'form_data': form_data
                })
            
            # Validar que el nombre de usuario no exista
            if Usuario.objects.filter(nombre_usuario=nombre_usuario).exists():
                messages.error(request, "El nombre de usuario ya esta registrado, intenta con otro.")
                return render(request, 'login.html', {
                    'show_register': True,
                    'form_data': form_data
                })
            
            # Validar que el correo no exista
            if Usuario.objects.filter(correo=correo).exists():
                messages.error(request, "El correo ya esta registrado, intenta con otro.")
                return render(request, 'login.html', {
                    'show_register': True,
                    'form_data': form_data
                })
            
            # Concatenar el prefijo "57" al teléfono
            telefono_completo = "57" + telefono
            
            # Crear el nuevo usuario
            nuevo_usuario = Usuario.objects.create_user(
                nombre_completo=nombre_completo,
                nombre_usuario=nombre_usuario,
                correo=correo,
                telefono=telefono_completo,
                direccion=direccion,
                password=contrasena1  # ✅ Usar 'password' no 'contrasena'
            )
            
            messages.success(request, "El usuario ha sido creado exitosamente. Ahora puedes iniciar sesión.")
            return redirect('login')
        
        # INICIO DE SESIÓN
        else:
            nombre_usuario = request.POST['nombre_usuario']
            contrasena = request.POST['contrasena']
            
            # Verificar si el usuario existe
            try:
                usuario_check = Usuario.objects.get(nombre_usuario=nombre_usuario)
                
                # Verificar si el usuario está activo
                if not usuario_check.is_active:
                    messages.error(request, 'Tu cuenta ha sido desactivada. Contacta al administrador para más información.')
                    return redirect('login')
                
                # Verificar si puede intentar hacer login (no está bloqueado)
                if not usuario_check.can_attempt_login():
                    if usuario_check.lockout_time:
                        time_since_lockout = timezone.now() - usuario_check.lockout_time
                        remaining_minutes = max(0, 15 - int(time_since_lockout.total_seconds() / 60))
                        messages.error(request, f'Tu cuenta está temporalmente bloqueada por seguridad debido a múltiples intentos fallidos. Intenta nuevamente en {remaining_minutes} minutos.')
                    else:
                        messages.error(request, 'Tu cuenta está temporalmente bloqueada por seguridad. Contacta al administrador.')
                    return redirect('login')
                
            except Usuario.DoesNotExist:
                # Si el usuario no existe, mostrar mensaje genérico
                messages.error(request, 'Usuario o contraseña incorrectos.')
                return redirect('login')
            
            # Autenticar usuario
            usuario = authenticate(request, username=nombre_usuario, password=contrasena)
            
            if usuario is not None:
                # Login exitoso - resetear intentos fallidos
                usuario.reset_failed_attempts()
                
                # Verificar nuevamente que el usuario esté activo (por seguridad)
                if usuario.is_active:
                    auth_login(request, usuario)
                    messages.success(request, f'Bienvenido de nuevo, {usuario.nombre_usuario}!')
                    # Redirigir a la página solicitada o al home por defecto
                    next_url = request.GET.get('next', 'home')
                    return redirect(next_url)
                else:
                    messages.error(request, 'Tu cuenta ha sido desactivada. Contacta al administrador para más información.')
                    return redirect('login')
            else:
                # Login fallido - incrementar intentos fallidos
                try:
                    usuario_fallido = Usuario.objects.get(nombre_usuario=nombre_usuario)
                    usuario_fallido.increment_failed_attempts()
                    
                    # Mostrar mensaje específico según los intentos restantes
                    remaining_attempts = usuario_fallido.get_remaining_attempts()
                    
                    if usuario_fallido.lockout_time:  # Si tiene tiempo de bloqueo activo
                        messages.error(request, 'Has excedido el número máximo de intentos de login. Tu cuenta ha sido bloqueada temporalmente por 15 minutos por seguridad.')
                    elif remaining_attempts <= 3 and remaining_attempts > 0:
                        messages.warning(request, f'Usuario o contraseña incorrectos. Te quedan {remaining_attempts} intentos antes de que tu cuenta sea bloqueada temporalmente.')
                    else:
                        messages.error(request, 'Usuario o contraseña incorrectos.')
                        
                except Usuario.DoesNotExist:
                    # Si el usuario no existe, mostrar mensaje genérico
                    messages.error(request, 'Usuario o contraseña incorrectos.')
                
                return redirect('login')
    
    # GET: Mostrar la página de login/registro
    else:
        # Verificar si se debe mostrar el formulario de registro
        show_register = request.GET.get('register') == 'true'
        return render(request, 'auth/login.html', {'show_register': show_register})
    

@login_required
def perfil_usuario(request):
    usuario = request.user  # Obtén el usuario actualmente autenticado
    form = ProfileUserForm(instance=usuario)  # Crea un formulario con los datos del usuario

    if request.method == 'POST':
        form = ProfileUserForm(request.POST, instance=usuario)  # Ya no incluye request.FILES

        if form.is_valid():
            # Manejar la actualización de la contraseña antes de guardar
            contrasena11 = request.POST.get('contrasena11', '').strip()
            contrasena22 = request.POST.get('contrasena22', '').strip()

            # Validar contraseñas si se proporcionan
            if contrasena11 or contrasena22:
                if not contrasena11 or not contrasena22:
                    messages.error(request, 'Debes completar ambos campos de contraseña.')
                    return render(request, 'usuarios/perfil_usuario.html', {'usuario': usuario, 'form': form})
                
                if contrasena11 != contrasena22:
                    messages.error(request, 'Las contraseñas no coinciden.')
                    return render(request, 'usuarios/perfil_usuario.html', {'usuario': usuario, 'form': form})
                
                if len(contrasena11) < 6:
                    messages.error(request, 'La contraseña debe tener al menos 6 caracteres.')
                    return render(request, 'usuarios/perfil_usuario.html', {'usuario': usuario, 'form': form})

            # Guardar los datos del usuario
            usuario_actualizado = form.save()

            # Cambiar la contraseña si se proporcionó
            if contrasena11:
                usuario_actualizado.set_password(contrasena11)
                usuario_actualizado.save()
                messages.success(request, 'Perfil y contraseña actualizados correctamente.')
            else:
                messages.success(request, 'Perfil actualizado correctamente.')
            
            # Redirigir al perfil para mostrar la notificación
            return redirect('perfil')
        else:
            # Si el formulario no es válido, mostrar errores
            messages.error(request, 'Por favor corrige los errores en el formulario.')

    return render(request, 'usuarios/perfil_usuario.html', {'usuario': usuario, 'form': form})

@login_required
def logout(request):
    auth_logout(request)
    return redirect('home')



def nosotros(request):
    return render(request, 'pages_informativas/nosotros.html')



def servicios(request):
    comentarios = Comentario.objects.all().order_by('-fecha')  # Recupera todos los comentarios
    servicios_list = Servicio.objects.all().order_by('nombre_servicio')  # Obtener todos los servicios disponibles
    
    # Para el carrusel, mostrar todos los servicios sin paginación
    servicios = servicios_list
    
    return render(request, 'servicios/servicios.html', {'comentarios': comentarios, 'servicios': servicios})

def servicios_page_ajax(request):
    """Vista AJAX para cargar todos los servicios en el carrusel"""
    servicios = Servicio.objects.all().order_by('nombre_servicio')
    
    # Renderizar el template parcial con todos los servicios
    return render(request, 'servicios/servicios_partial.html', {'servicios': servicios})

@usuario_required
def reservas(request):
    ahora = timezone.now()
    hoy = ahora.date()
    servicios = Servicio.objects.all().order_by('nombre_servicio')  # Obtener todos los servicios disponibles
    empresas = Empresa.objects.filter(verificada=True).order_by('nombre_empresa')    # Obtener solo las empresas verificadas
    empresaservicio = EmpresaServicio.objects.all()
    
    # Crear un diccionario con el conteo de empresas por servicio
    conteo_empresas_por_servicio = {}
    for servicio in servicios:
        conteo = EmpresaServicio.objects.filter(
            servicio=servicio,
            empresa__verificada=True
        ).count()
        conteo_empresas_por_servicio[servicio.id_servicio] = conteo
    
    # Debug: Verificar que las empresas existen
    print(f"🔍 Empresas encontradas: {empresas.count()}")
    print(f"🔍 Total servicios: {servicios.count()}")
    print(f"🔍 Total relaciones empresa-servicio: {empresaservicio.count()}")
    print(f"🔍 Conteo de empresas por servicio: {conteo_empresas_por_servicio}")
    
    # Obtener el usuario actual (si está autenticado)
    usuario = request.user if request.user.is_authenticated else None
    
    fecha_seleccionada = None
    servicios_filtrados = []
    
    # Inicializar la variable 'ocupadas' antes de cualquier otra cosa.
    reservas_existentes = Reserva.objects.all()
    
    ocupadas = {
        "fechas": {reserva.fecha for reserva in reservas_existentes},
        "horas": {}
    }
    
    for reserva in reservas_existentes:
        if reserva.fecha not in ocupadas["horas"]:
            ocupadas["horas"][reserva.fecha] = set()
        ocupadas["horas"][reserva.fecha].add(reserva.hora.strftime('%H:%M'))
    
    if request.method == "POST":
        print(f"🔍 POST request recibido para reservas")
        print(f"📝 POST data: {dict(request.POST)}")
        
        # Obtener datos del formulario paso a paso
        servicios_ids_str = request.POST.get('servicios_ids', '')  # IDs separados por coma
        empresa_id = request.POST.get('empresa_id')
        fecha_seleccionada = request.POST.get('fecha')
        hora_seleccionada = request.POST.get('hora')  # Ya viene en formato 24h desde el frontend
        
        print(f"🏢 Empresa ID: {empresa_id}")
        print(f"📅 Fecha: {fecha_seleccionada}")
        print(f"🕐 Hora: {hora_seleccionada}")
        print(f"🛠️ Servicios IDs string: {servicios_ids_str}")
        
        # Convertir string de IDs a lista
        try:
            servicios_ids = [int(id.strip()) for id in servicios_ids_str.split(',') if id.strip()]
        except (ValueError, AttributeError):
            servicios_ids = []
        
        print(f"🛠️ Servicios IDs list: {servicios_ids}")
        
        # Verificar si es una petición AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Validar que se hayan seleccionado servicios
        if not servicios_ids:
            error_msg = "Debes seleccionar al menos un servicio."
            if is_ajax:
                response = JsonResponse({'success': False, 'message': error_msg})
                response['Content-Type'] = 'application/json'
                return response
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Validar que se haya seleccionado empresa
        if not empresa_id:
            error_msg = "Debes seleccionar una empresa."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Validar que se haya seleccionado fecha y hora
        if not fecha_seleccionada or not hora_seleccionada:
            error_msg = "Debes seleccionar fecha y hora para la reserva."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Verificar que todos los servicios existan
        try:
            servicios_seleccionados = Servicio.objects.filter(id_servicio__in=servicios_ids)
            if len(servicios_seleccionados) != len(servicios_ids):
                raise Servicio.DoesNotExist
        except (Servicio.DoesNotExist, ValueError):
            error_msg = "Uno o más servicios seleccionados no existen."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Verificar que la empresa exista
        try:
            empresa = Empresa.objects.get(id_empresa=empresa_id, verificada=True)
        except Empresa.DoesNotExist:
            error_msg = "La empresa seleccionada no existe o no está verificada."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Verificar que la empresa ofrezca todos los servicios seleccionados
        servicios_empresa = Servicio.objects.filter(
            id_servicio__in=EmpresaServicio.objects.filter(empresa=empresa).values('servicio')
        )
        
        servicios_no_disponibles = []
        for servicio in servicios_seleccionados:
            if servicio not in servicios_empresa:
                servicios_no_disponibles.append(servicio.nombre_servicio)
        
        if servicios_no_disponibles:
            error_msg = f"Los siguientes servicios no están disponibles en {empresa.nombre_empresa}: {', '.join(servicios_no_disponibles)}"
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Validar formato de fecha
        try:
            fecha_obj = datetime.strptime(fecha_seleccionada, '%Y-%m-%d').date()
        except ValueError:
            error_msg = "Formato de fecha inválido."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Validar que la fecha no sea en el pasado
        if fecha_obj < hoy:
            error_msg = "No puedes reservar una fecha en el pasado."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Validar formato de hora
        try:
            hora_obj = datetime.strptime(hora_seleccionada, '%H:%M').time()
        except ValueError:
            error_msg = "Formato de hora inválido."
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
        
        # Verificar si la fecha y hora están ocupadas para esta empresa específica
        reserva_existente = Reserva.objects.filter(
            empresa=empresa,
            fecha=fecha_obj,
            hora=hora_obj
        ).exists()
        
        print(f"🔍 Verificando disponibilidad:")
        print(f"   - Empresa: {empresa.nombre_empresa}")
        print(f"   - Fecha: {fecha_obj}")
        print(f"   - Hora: {hora_obj}")
        print(f"   - ¿Existe reserva?: {reserva_existente}")
        
        if reserva_existente:
            error_msg = f"Lo siento, la fecha {fecha_seleccionada} a las {hora_seleccionada} ya está ocupada en {empresa.nombre_empresa}."
            print(f"❌ {error_msg}")
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg})
            messages.error(request, error_msg)
            return redirect('reservas')
            
        # Verificar si el usuario tiene una suscripción activa
        suscripcion_activa = None
        tiene_suscripcion = False
        usar_suscripcion = False
        
        print(f"🔍 Verificando suscripción para usuario: {usuario.nombre_usuario if usuario else 'Anónimo'}")
        
        try:
            from .models import SuscripcionUsuario
            suscripcion_activa = SuscripcionUsuario.objects.filter(
                usuario=usuario,
                estado='activa'
            ).first()
            
            print(f"🔍 Suscripción encontrada: {suscripcion_activa}")
            
            if suscripcion_activa:
                # Usar la misma lógica que mi_suscripcion
                suscripcion_activa.reiniciar_contador_mensual()
                
                print(f"📋 Plan: {suscripcion_activa.plan.nombre}")
                print(f"📊 Servicios utilizados: {suscripcion_activa.servicios_utilizados_mes}")
                print(f"📊 Servicios totales permitidos: {suscripcion_activa.plan.cantidad_servicios_mes}")
                print(f"⏰ Fecha fin: {suscripcion_activa.fecha_fin}")
                print(f"✅ ¿Puede usar servicio?: {suscripcion_activa.puede_usar_servicio()}")
                
                tiene_suscripcion = True
                print(f"✅ Usuario tiene suscripción activa (estado='activa')")
                
                # Verificar si puede usar servicios de la suscripción
                if suscripcion_activa.puede_usar_servicio():
                    usar_suscripcion = True
                    print(f"✅ Usuario puede usar suscripción. Servicios seleccionados: {len(servicios_seleccionados)}")
                else:
                    print(f"⚠️ Usuario {usuario.nombre_usuario} ha agotado sus servicios del mes. Servicios restantes: {suscripcion_activa.servicios_restantes()}")
                    # No bloquear, solo usar pago individual
                    usar_suscripcion = False
            else:
                print(f"❌ Usuario no tiene suscripción activa o está vencida")
                
        except Exception as e:
            print(f"❌ Error verificando suscripción: {e}")
            import traceback
            traceback.print_exc()
        
        print(f"🎯 Resultado: usar_suscripcion = {usar_suscripcion}, tiene_suscripcion = {tiene_suscripcion}")
        
        # Crear la reserva
        reserva = Reserva(
            empresa=empresa,
            fecha=fecha_obj,
            hora=hora_obj,
            usuario=usuario,
            suscripcion_utilizada=suscripcion_activa if usar_suscripcion else None,
            es_pago_individual=not usar_suscripcion
        )
        reserva.save()
        
        # Separar servicios incluidos en el plan de los servicios adicionales
        servicios_incluidos_plan = []
        servicios_adicionales = []
        precio_total = 0
        servicios_nombres = []
        
        # Diccionario para almacenar descuentos de PlanServicio
        descuentos_plan = {}
        
        if usar_suscripcion and suscripcion_activa:
            # Obtener los servicios incluidos en el plan del usuario
            from .models import PlanServicio
            servicios_del_plan = suscripcion_activa.plan.servicios_incluidos.all()
            servicios_del_plan_ids = list(servicios_del_plan.values_list('id_servicio', flat=True))
            
            # Obtener los descuentos de cada servicio desde PlanServicio
            plan_servicios = PlanServicio.objects.filter(
                plan=suscripcion_activa.plan,
                servicio__in=servicios_seleccionados
            ).select_related('servicio')
            
            for plan_servicio in plan_servicios:
                descuentos_plan[plan_servicio.servicio.id_servicio] = plan_servicio.porcentaje_descuento
                print(f"🔍 Descuento para {plan_servicio.servicio.nombre_servicio}: {plan_servicio.porcentaje_descuento}%")
            
            print(f"🔍 Servicios incluidos en el plan {suscripcion_activa.plan.nombre}: {servicios_del_plan_ids}")
            
            # Separar servicios seleccionados entre incluidos y adicionales
            for servicio in servicios_seleccionados:
                if servicio.id_servicio in servicios_del_plan_ids:
                    servicios_incluidos_plan.append(servicio)
                    print(f"✅ Servicio incluido en plan: {servicio.nombre_servicio}")
                else:
                    servicios_adicionales.append(servicio)
                    precio_total += servicio.precio
                    print(f"💰 Servicio adicional (se cobrará): {servicio.nombre_servicio} - ${servicio.precio}")
            
            # VALIDACIÓN: Verificar que no se excedan los servicios restantes del plan
            servicios_restantes = suscripcion_activa.servicios_restantes()
            if len(servicios_incluidos_plan) > servicios_restantes:
                error_msg = f"Has seleccionado {len(servicios_incluidos_plan)} servicios del plan, pero solo tienes {servicios_restantes} servicios disponibles. Por favor, reduce tu selección."
                print(f"❌ {error_msg}")
                if is_ajax:
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
                return redirect('reservas')
        else:
            # Si no usa suscripción, todos los servicios son adicionales
            servicios_adicionales = list(servicios_seleccionados)
        
        # Crear las relaciones entre reserva y servicios (múltiples)
        for servicio in servicios_seleccionados:
            # Determinar si es servicio del plan y obtener descuento
            es_servicio_plan = servicio in servicios_incluidos_plan and usar_suscripcion
            descuento_plan = descuentos_plan.get(servicio.id_servicio, 0) if es_servicio_plan else 0
            precio_original = servicio.precio
            
            # Calcular precio con descuento si aplica
            if es_servicio_plan and descuento_plan > 0:
                precio_con_descuento = float(precio_original) * (1 - float(descuento_plan) / 100)
                precio_aplicado = precio_con_descuento
                print(f"💵 {servicio.nombre_servicio}: Precio original ${precio_original} -> Con descuento {descuento_plan}% -> ${precio_aplicado:.2f}")
            elif es_servicio_plan:
                # Servicio incluido sin descuento especial (precio completo cubierto por plan)
                precio_aplicado = 0
                print(f"✅ {servicio.nombre_servicio}: Incluido en plan (sin costo)")
            else:
                # Servicio adicional (sin descuento)
                precio_aplicado = precio_original
                print(f"💰 {servicio.nombre_servicio}: Servicio adicional ${precio_aplicado}")
            
            ReservaServicio.objects.create(
                reserva=reserva,
                servicio=servicio,
                es_servicio_plan=es_servicio_plan,
                descuento_plan_individual=descuento_plan,
                precio_original=precio_original,
                precio_aplicado=precio_aplicado
            )
            servicios_nombres.append(servicio.nombre_servicio)
        
        # Calcular precio total usando el método del modelo que suma todos los precio_aplicado
        precio_total = reserva.calcular_total_reserva()
        
        # Si usa suscripción, incrementar el contador solo con los servicios incluidos en el plan
        if usar_suscripcion and suscripcion_activa and len(servicios_incluidos_plan) > 0:
            servicios_del_plan_utilizados = len(servicios_incluidos_plan)
            suscripcion_activa.servicios_utilizados_mes += servicios_del_plan_utilizados
            suscripcion_activa.save()
            print(f"✅ Servicios del plan utilizados: {servicios_del_plan_utilizados}")
            print(f"✅ Servicios adicionales cobrados: {len(servicios_adicionales)} (${precio_total})")
            print(f"✅ Contador actualizado: {suscripcion_activa.servicios_utilizados_mes}/{suscripcion_activa.plan.cantidad_servicios_mes if suscripcion_activa.plan.cantidad_servicios_mes > 0 else 'Ilimitado'}")
        else:
            print(f"💰 Reserva creada con pago individual. Precio total: ${precio_total}")
        
        # Enviar correo de confirmación
        numero_reserva = None
        try:
            correo_enviado = enviar_correo_confirmacion_reserva(
                usuario=usuario,
                empresa=reserva.empresa,
                servicios=servicios_seleccionados,
                fecha=fecha_seleccionada,
                hora=hora_seleccionada,
                precio_total=precio_total,
                numero_reserva=reserva.numero_reserva
            )
            if correo_enviado:
                # El número de reserva se genera dentro de la función de correo
                import uuid
                numero_reserva = f"ANW-{str(uuid.uuid4())[:8].upper()}"
                print(f"✅ Correo de confirmación enviado a {usuario.correo}")
            else:
                print(f"❌ Error enviando correo a {usuario.correo}")
        except Exception as e:
            print(f"❌ Excepción enviando correo: {e}")
        
        # Crear mensaje de éxito personalizado más detallado
        if len(servicios_nombres) == 1:
            base_msg = f"Tu cita para {servicios_nombres[0]} ha sido reservada para el {fecha_seleccionada} a las {hora_seleccionada}."
        else:
            base_msg = f"Tu cita para {len(servicios_nombres)} servicios ha sido reservada para el {fecha_seleccionada} a las {hora_seleccionada}."
        
        # Agregar información detallada sobre el tipo de pago
        if usar_suscripcion and len(servicios_incluidos_plan) > 0:
            servicios_restantes = suscripcion_activa.servicios_restantes()
            
            if len(servicios_adicionales) > 0:
                # Hay servicios incluidos y adicionales
                servicios_incluidos_nombres = [s.nombre_servicio for s in servicios_incluidos_plan]
                servicios_adicionales_nombres = [s.nombre_servicio for s in servicios_adicionales]
                
                success_msg = f"{base_msg}\n"
                success_msg += f"✅ Servicios incluidos en tu plan: {', '.join(servicios_incluidos_nombres)}\n"
                success_msg += f"💰 Servicios adicionales (${precio_total}): {', '.join(servicios_adicionales_nombres)}\n"
                success_msg += f"Te quedan {servicios_restantes} servicios en tu plan este mes."
            else:
                # Solo servicios incluidos
                success_msg = f"{base_msg} ✅ Todos los servicios están incluidos en tu plan. Te quedan {servicios_restantes} servicios este mes."
        elif usar_suscripcion and len(servicios_incluidos_plan) == 0:
            # Tiene suscripción pero ningún servicio está incluido
            success_msg = f"{base_msg} 💰 Ninguno de estos servicios está incluido en tu plan actual, por lo que se cobrará individualmente (${precio_total})."
        else:
            if tiene_suscripcion:
                success_msg = f"{base_msg} 💰 Has agotado tus servicios del mes, por lo que esta reserva se cobrará individualmente (${precio_total})."
            else:
                success_msg = f"{base_msg} 💰 Total a pagar: ${precio_total}."
        
        if is_ajax:
            # Obtener detalles de servicios desde el modelo
            detalle_servicios = reserva.obtener_detalle_servicios()
            
            response = JsonResponse({
                'success': True,
                'message': success_msg,
                'reserva': {
                    'id': reserva.id_reserva,
                    # Priorizar el número almacenado en el modelo (generado en save())
                    'numero_reserva': reserva.numero_reserva if getattr(reserva, 'numero_reserva', None) else (numero_reserva or f"ANW-{reserva.id_reserva:08d}"),
                    'empresa': empresa.nombre_empresa,
                    'servicios': servicios_nombres,
                    'precio_total': str(precio_total),
                    'fecha': fecha_seleccionada,
                    'hora': hora_seleccionada,
                    'servicios_incluidos_plan': [
                        {
                            'nombre': item['nombre'],
                            'precio_original': str(item['precio_original']),
                            'precio_aplicado': str(item['precio_aplicado']),
                            'descuento': str(item['descuento']),
                            'ahorro': str(item['ahorro']),
                            'es_gratis': item['precio_aplicado'] == 0
                        } for item in detalle_servicios['servicios_plan']
                    ],
                    'servicios_adicionales': [
                        {
                            'nombre': item['nombre'],
                            'precio_original': str(item['precio_original']),
                            'precio_aplicado': str(item['precio_aplicado']),
                            'es_gratis': False
                        } for item in detalle_servicios['servicios_adicionales']
                    ],
                    'servicios_detalle': [
                        {
                            'nombre': servicio.nombre_servicio,
                            'precio_original': str(servicio.precio),
                            'precio_aplicado': str(descuentos_plan.get(servicio.id_servicio, 0) if servicio in servicios_incluidos_plan else servicio.precio),
                            'descuento': str(descuentos_plan.get(servicio.id_servicio, 0)) if servicio in servicios_incluidos_plan else '0',
                            'es_plan': servicio in servicios_incluidos_plan
                        } for servicio in servicios_seleccionados
                    ],
                    'ahorro_total': str(detalle_servicios['ahorro_total'])
                }
            })
            # Asegurar que el Content-Type sea correcto
            response['Content-Type'] = 'application/json'
            return response
        
        messages.success(request, success_msg)
        return redirect('citas')  # Redirigir a la página de citas del usuario
    
    # Generar horas disponibles
    horas_disponibles = {}
    for i in range(15):
        fecha = hoy + timedelta(days=i)
        horas_disponibles[fecha] = []
        
        for h in range(8, 21):  # De 8:00 AM a 8:00 PM
            hora_formateada_24h = f"{h:02}:00"
            
            if fecha == hoy and h < ahora.hour:
                continue
            
            # Verifica si la hora está ocupada para esta fecha
            if hora_formateada_24h not in ocupadas["horas"].get(fecha, set()):
                # Convertir a formato 12h con AM/PM para mostrar al usuario
                hora_12h = convertir_hora_12h(hora_formateada_24h)
                horas_disponibles[fecha].append(hora_12h)
    
    fechas_disponibles = [hoy + timedelta(days=i) for i in range(15)]
    
    # Obtener información de la suscripción del usuario si está autenticado
    suscripcion_info = None
    if usuario and usuario.is_authenticated:
        print(f"🔍 Obteniendo información de suscripción para: {usuario.nombre_usuario}")
        try:
            from .models import SuscripcionUsuario
            suscripcion_activa = SuscripcionUsuario.objects.filter(
                usuario=usuario,
                estado='activa'
            ).first()
            
            print(f"🔍 Suscripción encontrada en GET: {suscripcion_activa}")
            
            if suscripcion_activa:
                # Usar la misma lógica que mi_suscripcion - solo verificar estado='activa'
                # Reiniciar contador mensual como en mi_suscripcion
                suscripcion_activa.reiniciar_contador_mensual()
                
                servicios_restantes = suscripcion_activa.servicios_restantes()
                suscripcion_info = {
                    'tiene_suscripcion': True,
                    'plan_nombre': suscripcion_activa.plan.nombre,
                    'plan_id': suscripcion_activa.plan.id_plan,
                    'servicios_restantes': servicios_restantes,
                    'servicios_utilizados': suscripcion_activa.servicios_utilizados_mes,
                    'servicios_totales': suscripcion_activa.plan.cantidad_servicios_mes,
                    'fecha_fin': suscripcion_activa.fecha_fin,
                    'puede_usar_servicio': suscripcion_activa.puede_usar_servicio(),
                    'esta_activa': True,  # Si está en BD con estado='activa', considerarlo activo
                    'suscripcion_id': suscripcion_activa.id_suscripcion
                }
                print(f"✅ Info suscripción: Plan={suscripcion_activa.plan.nombre}, Restantes={servicios_restantes}, Utilizados={suscripcion_activa.servicios_utilizados_mes}")
            else:
                suscripcion_info = {'tiene_suscripcion': False}
                print(f"❌ No hay suscripción con estado='activa'")
        except Exception as e:
            print(f"❌ Error obteniendo información de suscripción: {e}")
            import traceback
            traceback.print_exc()
            suscripcion_info = {'tiene_suscripcion': False}
    else:
        suscripcion_info = {'tiene_suscripcion': False}
        print(f"🔍 Usuario no autenticado, sin suscripción")
    
    return render(request, 'reservas/reservas.html', {
        'ocupadas': ocupadas,
        'horas_disponibles': horas_disponibles,
        'fechas_disponibles': fechas_disponibles,
        'hoy': hoy,
        'servicios_filtrados': servicios_filtrados,
        'empresas': empresas,
        'servicios': servicios,
        'empresaservicio': empresaservicio,
        'conteo_empresas_por_servicio': conteo_empresas_por_servicio,
        'suscripcion_info': suscripcion_info
    })

# Vista para obtener servicios por empresa (AJAX)
@usuario_required
def obtener_servicios(request):
    empresa_id = request.GET.get('empresa_id')
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            # Forma más directa: usar la relación many-to-many
            servicios = empresa.servicios.all()
            # (Si se quisiera mantener la otra forma correcta sería usando values_list('servicio', flat=True))
            servicios_data = [
                {
                    'id_servicio': s.id_servicio,
                    'nombre_servicio': s.nombre_servicio,
                    'descripcion': s.descripcion,
                    'precio': s.precio
                } for s in servicios
            ]
            empresa_data = {
                'nombre_empresa': empresa.nombre_empresa,
                'direccion': empresa.direccion,
                'telefono': empresa.telefono
            }
            return JsonResponse({'servicios': servicios_data, 'empresa': empresa_data})
        except Empresa.DoesNotExist:
            return JsonResponse({'servicios': [], 'empresa': None})
    return JsonResponse({'servicios': [], 'empresa': None})

# Nueva vista para obtener empresas filtradas por servicios (AJAX)
@usuario_required
def obtener_servicios_plan(request):
    """
    Obtiene los servicios incluidos en el plan del usuario y los servicios adicionales
    """
    try:
        # Obtener la suscripción activa del usuario
        suscripcion = SuscripcionUsuario.objects.filter(
            usuario=request.user,
            estado='activa'
        ).first()
        
        if not suscripcion:
            return JsonResponse({
                'success': False,
                'message': 'No tienes una suscripción activa'
            })
        
        # Obtener servicios incluidos en el plan
        servicios_plan = suscripcion.plan.servicios_incluidos.all()
        servicios_plan_data = [
            {
                'id_servicio': servicio.id_servicio,
                'nombre_servicio': servicio.nombre_servicio,
                'descripcion': servicio.descripcion,
                'precio': float(servicio.precio)
            } for servicio in servicios_plan
        ]
        
        # Obtener todos los servicios disponibles (para los adicionales)
        todos_los_servicios = Servicio.objects.all()
        servicios_ids_del_plan = set(servicios_plan.values_list('id_servicio', flat=True))
        
        servicios_adicionales = todos_los_servicios.exclude(id_servicio__in=servicios_ids_del_plan)
        servicios_adicionales_data = [
            {
                'id_servicio': servicio.id_servicio,
                'nombre_servicio': servicio.nombre_servicio,
                'descripcion': servicio.descripcion,
                'precio': float(servicio.precio)
            } for servicio in servicios_adicionales
        ]
        
        return JsonResponse({
            'success': True,
            'servicios_plan': servicios_plan_data,
            'servicios_adicionales': servicios_adicionales_data,
            'puede_usar_servicio': suscripcion.puede_usar_servicio(),
            'servicios_restantes': suscripcion.servicios_restantes()
        })
        
    except Exception as e:
        print(f"Error obteniendo servicios del plan: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor'
        })

@usuario_required
def obtener_empresas_por_servicios(request):
    """
    Obtiene las empresas que ofrecen todos los servicios seleccionados
    """
    servicios_ids = request.GET.getlist('servicios[]')  # Lista de IDs de servicios
    
    print(f"🔍 Buscando empresas para servicios: {servicios_ids}")
    
    if not servicios_ids:
        return JsonResponse({'empresas': [], 'message': 'No se han seleccionado servicios'})
    
    try:
        # Convertir a enteros
        servicios_ids = [int(sid) for sid in servicios_ids]
        
        # Verificar que todos los servicios existan
        servicios_validos = Servicio.objects.filter(id_servicio__in=servicios_ids)
        if len(servicios_validos) != len(servicios_ids):
            return JsonResponse({'empresas': [], 'error': 'Algunos servicios no existen'})
        
        # Obtener empresas que tienen TODOS los servicios seleccionados
        empresas = Empresa.objects.filter(verificada=True)
        
        # Filtrar empresas que ofrecen todos los servicios
        for servicio_id in servicios_ids:
            empresas = empresas.filter(
                id_empresa__in=EmpresaServicio.objects.filter(
                    servicio_id=servicio_id
                ).values('empresa_id')
            )
        
        print(f"✅ Empresas encontradas que ofrecen todos los servicios: {empresas.count()}")
        
    except (ValueError, TypeError) as e:
        print(f"❌ Error procesando servicios_ids: {e}")
        return JsonResponse({'empresas': [], 'error': 'IDs de servicios inválidos'})
    
    # Preparar datos de respuesta con información completa
    empresas_data = []
    for empresa in empresas:
        empresas_data.append({
            'id': empresa.id_empresa,
            'nombre': empresa.nombre_empresa,
            'direccion': empresa.direccion,
            'telefono': empresa.telefono,
            'email': empresa.email,
            'verificada': empresa.verificada
        })
    
    return JsonResponse({
        'empresas': empresas_data,
        'total': len(empresas_data),
        'servicios_solicitados': servicios_ids
    })

# Vista para obtener información de la empresa (AJAX)
@usuario_required
def obtener_info_empresa(request):
    empresa_id = request.GET.get('empresa_id')
    if empresa_id:
        try:
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            return JsonResponse({
                'nombre': empresa.nombre_empresa,
                'direccion': empresa.direccion,
                'telefono': empresa.telefono
            })
        except Empresa.DoesNotExist:
            return JsonResponse({
                'nombre': 'No disponible',
                'direccion': 'No disponible',
                'telefono': 'No disponible'
            })
    return JsonResponse({
        'nombre': 'No disponible',
        'direccion': 'No disponible',
        'telefono': 'No disponible'
    })

# Vista para obtener información del servicio (AJAX)
def obtener_info_servicio(request):
    servicio_id = request.GET.get('servicio_id')
    if servicio_id:
        try:
            servicio = Servicio.objects.get(id_servicio=servicio_id)
            return JsonResponse({
                'nombre_servicio': servicio.nombre_servicio,
                'descripcion': servicio.descripcion,
                'precio': servicio.precio
            })
        except Servicio.DoesNotExist:
            return JsonResponse({
                'nombre_servicio': 'Servicio no encontrado',
                'descripcion': 'No disponible',
                'precio': 0
            })
    return JsonResponse({
        'nombre_servicio': 'Selecciona un servicio',
        'descripcion': '',
        'precio': 0
    })

# Función helper para convertir hora de 24h a 12h con AM/PM
def convertir_hora_12h(hora_24h):
    """Convierte hora en formato 24h (HH:MM) a formato 12h con AM/PM"""
    try:
        hora_obj = datetime.strptime(hora_24h, '%H:%M').time()
        # Formatear a 12 horas con AM/PM
        return hora_obj.strftime('%I:%M %p')
    except:
        return hora_24h  # Retorna la hora original si hay error

# Función helper para convertir hora de 12h a 24h
def convertir_hora_24h(hora_12h):
    """Convierte hora en formato 12h con AM/PM a formato 24h (HH:MM)"""
    try:
        hora_obj = datetime.strptime(hora_12h, '%I:%M %p').time()
        return hora_obj.strftime('%H:%M')
    except:
        return hora_12h  # Retorna la hora original si hay error

# Vista para obtener horas disponibles (AJAX)
@usuario_required
def get_horas(request):
    empresa_id = request.GET.get('empresa_id')
    fecha_str = request.GET.get('fecha')
    
    print(f"🕐 get_horas llamado con empresa_id={empresa_id}, fecha={fecha_str}")
    
    if not empresa_id or not fecha_str:
        print(f"❌ Parámetros faltantes: empresa_id={empresa_id}, fecha={fecha_str}")
        return JsonResponse({'success': False, 'horas_disponibles': [], 'error': 'Parámetros faltantes'})
    
    try:
        # Verificar que la empresa existe
        empresa = Empresa.objects.get(id_empresa=empresa_id, verificada=True)
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        print(f"📅 Fecha objeto: {fecha_obj}")
        
        # Verificar que la fecha no esté en el pasado
        if fecha_obj < timezone.now().date():
            print(f"❌ Fecha en el pasado")
            return JsonResponse({'success': False, 'horas_disponibles': [], 'message': 'No se pueden reservar fechas pasadas'})

        # Obtener TODAS las reservas para esta empresa y fecha
        reservas_existentes = Reserva.objects.filter(
            empresa=empresa,
            fecha=fecha_obj
        )
        
        print(f"🔍 Reservas existentes para empresa {empresa.nombre_empresa} en fecha {fecha_obj}: {reservas_existentes.count()}")
        
        # Obtener todas las horas ocupadas
        horas_ocupadas = set()
        for reserva in reservas_existentes:
            hora_str = reserva.hora.strftime('%H:%M')
            horas_ocupadas.add(hora_str)
            print(f"⏰ Hora ocupada: {hora_str}")

        print(f"🚫 Total horas ocupadas: {horas_ocupadas}")

        horas_disponibles = []
        ahora = timezone.now()
        
        # Generar horas de 8:00 AM a 6:00 PM (horario comercial)
        for h in range(8, 19):  # 8:00 a 18:00 (6:00 PM)
            hora_formateada_24h = f"{h:02}:00"
            
            # Si es hoy, no mostrar horas que ya pasaron
            if fecha_obj == ahora.date() and h <= ahora.hour:
                print(f"⏳ Hora {hora_formateada_24h} ya pasó")
                continue
            
            # Si la hora no está ocupada, agregarla
            if hora_formateada_24h not in horas_ocupadas:
                # Crear formato para el frontend con display y value
                horas_disponibles.append({
                    'value': hora_formateada_24h,
                    'display': hora_formateada_24h
                })
                print(f"✅ Hora disponible: {hora_formateada_24h}")
            else:
                print(f"❌ Hora ocupada: {hora_formateada_24h}")

        print(f"📋 Horas disponibles finales: {[h['value'] for h in horas_disponibles]}")
        return JsonResponse({
            'success': True,
            'horas_disponibles': horas_disponibles,
            'fecha': fecha_str,
            'empresa': empresa.nombre_empresa
        })

    except Empresa.DoesNotExist:
        print(f"❌ Empresa no encontrada: {empresa_id}")
        return JsonResponse({'success': False, 'horas_disponibles': [], 'error': 'Empresa no encontrada'})
    except ValueError as e:
        print(f"❌ Error de formato de fecha: {e}")
        return JsonResponse({'success': False, 'horas_disponibles': [], 'error': 'Formato de fecha inválido'})


def get_horas_edicion(request):
    """
    Obtiene horas disponibles para edición de citas.
    Verifica que haya al menos 12 horas entre ahora y la hora original de la cita.
    Si se cumple, muestra todas las horas disponibles del día en esa empresa.
    """
    empresa_id = request.GET.get('empresa_id')
    fecha_str = request.GET.get('fecha')
    hora_original_str = request.GET.get('hora_original')  # Formato HH:MM:SS o HH:MM
    
    print(f"🕐 get_horas_edicion llamado con empresa_id={empresa_id}, fecha={fecha_str}, hora_original={hora_original_str}")
    
    if not empresa_id or not fecha_str or not hora_original_str:
        print(f"❌ Parámetros faltantes")
        return JsonResponse({
            'success': False, 
            'horas': [], 
            'error': 'Parámetros faltantes'
        })
    
    try:
        # Verificar que la empresa existe
        empresa = Empresa.objects.get(id_empresa=empresa_id, verificada=True)
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        # Parsear la hora original (puede venir en formato HH:MM:SS o HH:MM)
        if len(hora_original_str.split(':')) == 3:
            hora_original = datetime.strptime(hora_original_str, '%H:%M:%S').time()
        else:
            hora_original = datetime.strptime(hora_original_str, '%H:%M').time()
        
        print(f"📅 Fecha: {fecha_obj}, Hora original: {hora_original}")
        
        # Crear datetime de la cita original
        from django.utils import timezone
        fecha_hora_original = timezone.make_aware(
            datetime.combine(fecha_obj, hora_original)
        )
        
        # Obtener la hora actual
        ahora = timezone.now()
        
        # Calcular cuántas horas faltan para la cita original
        tiempo_hasta_cita = fecha_hora_original - ahora
        horas_hasta_cita = tiempo_hasta_cita.total_seconds() / 3600
        
        print(f"⏰ Ahora: {ahora}")
        print(f"📍 Cita original: {fecha_hora_original}")
        print(f"⏳ Horas hasta la cita: {horas_hasta_cita:.2f} horas")
        
        # Verificar que haya al menos 12 horas de anticipación
        if horas_hasta_cita < 12:
            print(f"❌ No hay 12 horas de anticipación (solo {horas_hasta_cita:.2f} horas)")
            return JsonResponse({
                'success': False,
                'horas': [],
                'error': f'Solo se puede editar con al menos 12 horas de anticipación. Faltan {horas_hasta_cita:.2f} horas para tu cita.',
                'mensaje': 'No puedes editar esta cita porque no hay 12 horas de anticipación.',
                'horas_faltantes': round(horas_hasta_cita, 2)
            })
        
        print(f"✅ Hay suficiente tiempo para editar ({horas_hasta_cita:.2f} horas >= 12 horas)")
        
        # Obtener todas las reservas existentes para esta empresa y fecha
        reservas_existentes = Reserva.objects.filter(
            empresa=empresa,
            fecha=fecha_obj
        )
        
        print(f"🔍 Reservas existentes: {reservas_existentes.count()}")
        
        # Obtener todas las horas ocupadas (excluyendo la reserva actual si existe)
        reserva_id_actual = request.GET.get('reserva_id')
        horas_ocupadas = set()
        for reserva in reservas_existentes:
            # No incluir la hora de la reserva que se está editando
            if reserva_id_actual and str(reserva.id_reserva) == str(reserva_id_actual):
                continue
            hora_str = reserva.hora.strftime('%H:%M')
            horas_ocupadas.add(hora_str)
        
        print(f"🚫 Horas ocupadas: {horas_ocupadas}")
        
        horas_disponibles = []
        
        # Generar horas de 8:00 AM a 6:00 PM (horario comercial)
        for h in range(8, 19):  # 8:00 a 18:00 (6:00 PM)
            hora_formateada_24h = f"{h:02d}:00"
            
            # Crear datetime para esta hora
            fecha_hora_candidata = timezone.make_aware(
                datetime.combine(fecha_obj, datetime.strptime(hora_formateada_24h, '%H:%M').time())
            )
            
            # VALIDACIÓN 1: Si la hora candidata ya pasó, no mostrarla
            if fecha_hora_candidata <= ahora:
                print(f"⏳ Hora {hora_formateada_24h} ya pasó")
                continue
            
            # VALIDACIÓN 2: Verificar que haya al menos 12 horas hasta esta nueva hora
            tiempo_hasta_nueva_hora = fecha_hora_candidata - ahora
            horas_hasta_nueva_hora = tiempo_hasta_nueva_hora.total_seconds() / 3600
            
            if horas_hasta_nueva_hora < 12:
                print(f"⏳ Hora {hora_formateada_24h} no tiene 12h de anticipación ({horas_hasta_nueva_hora:.2f}h)")
                continue
            
            # VALIDACIÓN 3: La hora no debe estar ocupada
            if hora_formateada_24h in horas_ocupadas:
                print(f"❌ Hora {hora_formateada_24h} ocupada")
                continue
            
            # Si pasa todas las validaciones, agregar
            horas_disponibles.append(hora_formateada_24h)
            print(f"✅ Hora disponible: {hora_formateada_24h} ({horas_hasta_nueva_hora:.2f}h de anticipación)")
        
        print(f"📋 Total horas disponibles: {len(horas_disponibles)}")
        
        if not horas_disponibles:
            return JsonResponse({
                'success': False,
                'horas': [],
                'mensaje': 'No hay horas disponibles que cumplan con el requisito de 12 horas de anticipación.',
                'info': 'Todas las horas disponibles deben tener al menos 12 horas de anticipación desde ahora.'
            })
        
        return JsonResponse({
            'success': True,
            'horas': horas_disponibles,
            'fecha': fecha_str,
            'empresa': empresa.nombre_empresa,
            'hora_original': hora_original_str,
            'horas_hasta_cita_original': round(horas_hasta_cita, 2)
        })

    except Empresa.DoesNotExist:
        print(f"❌ Empresa no encontrada: {empresa_id}")
        return JsonResponse({
            'success': False, 
            'horas': [], 
            'error': 'Empresa no encontrada'
        })
    except ValueError as e:
        print(f"❌ Error de formato: {e}")
        return JsonResponse({
            'success': False, 
            'horas': [], 
            'error': f'Formato de datos inválido: {str(e)}'
        })
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        return JsonResponse({'success': False, 'horas_disponibles': [], 'error': 'Error interno del servidor'})




def planes(request):
    return render(request, 'planes/planes.html')


def planes_empresariales(request):
    """Vista para mostrar los planes empresariales disponibles"""
    # Obtener todos los planes empresariales activos, ordenados por precio
    planes_empresariales = PlanEmpresarial.objects.filter(activo=True).prefetch_related('servicios_incluidos').order_by('precio_mensual_por_vehiculo')
    
    context = {
        'planes_empresariales': planes_empresariales,
        'titulo': 'Planes Empresariales',
        'descripcion': 'Soluciones especializadas para flotas de vehículos y empresas de transporte'
    }
    
    return render(request, 'planes/planes_empresariales.html', context)


def solicitar_contacto_plan(request):
    """Vista AJAX para procesar solicitudes de contacto de planes empresariales"""
    if request.method == 'POST':
        try:
            plan_id = request.POST.get('plan_id')
            
            # Verificar que el plan existe
            plan = get_object_or_404(PlanEmpresarial, id_plan=plan_id, activo=True)
            
            # Debug: Imprimir datos recibidos
            print("Datos POST recibidos:", request.POST)
            
            # Crear el formulario con los datos recibidos
            form = SolicitudContactoPlanForm(request.POST)
            
            if form.is_valid():
                # Crear la solicitud
                solicitud = form.save(commit=False)
                solicitud.plan = plan
                solicitud.estado = 'pendiente'
                
                # Capturar IP del solicitante
                def get_client_ip(request):
                    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                    if x_forwarded_for:
                        ip = x_forwarded_for.split(',')[0]
                    else:
                        ip = request.META.get('REMOTE_ADDR')
                    return ip
                
                solicitud.ip_solicitante = get_client_ip(request)
                
                # Capturar User Agent
                solicitud.user_agent = request.META.get('HTTP_USER_AGENT', '')
                
                # Si quieres que fecha_contacto se llene automáticamente (aunque conceptualmente debería llenarse cuando se contacte al cliente)
                # solicitud.fecha_contacto = timezone.now()
                
                solicitud.save()
                
                # Debug: Verificar datos guardados
                print(f"Solicitud guardada - ID: {solicitud.id_solicitud}")
                print(f"Cargo guardado: '{solicitud.cargo}'")
                print(f"IP guardada: '{solicitud.ip_solicitante}'")
                print(f"User Agent guardado: '{solicitud.user_agent[:100]}...'")
                
                # Intentar enviar email de notificación
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    asunto = f"Nueva solicitud de contacto - Plan {plan.nombre}"
                    mensaje = f"""
                    Nueva solicitud de contacto para el plan empresarial: {plan.nombre}
                    
                    Datos del solicitante:
                    - Nombre: {solicitud.nombre_completo}
                    - Email: {solicitud.email}
                    - Teléfono: {solicitud.telefono}
                    - Empresa: {solicitud.empresa}
                    - Cargo: {solicitud.cargo}
                    - Cantidad de vehículos: {solicitud.cantidad_vehiculos}
                    
                    Mensaje adicional:
                    {solicitud.mensaje_adicional}
                    
                    Fecha de solicitud: {solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M')}
                    """
                    
                    send_mail(
                        asunto,
                        mensaje,
                        settings.DEFAULT_FROM_EMAIL,
                        [settings.DEFAULT_FROM_EMAIL],
                        fail_silently=True,
                    )
                except Exception as e:
                    # Log del error pero no fallar la respuesta
                    print(f"Error enviando email: {e}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Su solicitud de contacto ha sido enviada exitosamente. Nos pondremos en contacto con usted a la brevedad.',
                    'solicitud_id': solicitud.id_solicitud
                })
            else:
                # Devolver errores del formulario
                return JsonResponse({
                    'success': False,
                    'message': 'Por favor, corrija los errores en el formulario.',
                    'errors': form.errors
                })
                
        except PlanEmpresarial.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'El plan seleccionado no existe o no está disponible.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error interno del servidor: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido.'
    })


@login_required
def reservas_usuario(request):
    ahora = datetime.now()
    hoy = ahora.date()
    
    # Obtener reservas del usuario actual with prefetch_related para optimizar
    reservas_pendientes = Reserva.objects.filter(
        estado='pendiente', 
        usuario=request.user
    ).prefetch_related('servicios', 'empresa', 'reservaservicio_set')
    
    reservas_completadas = Reserva.objects.filter(
        estado='completado', 
        usuario=request.user
    ).prefetch_related('servicios', 'empresa', 'reservaservicio_set')
    
    # Obtener todas las reservas del usuario para el total
    total_reservas = Reserva.objects.filter(usuario=request.user)
    
    # Debug: Imprimir información sobre las reservas
    print(f"Usuario: {request.user.nombre_usuario}")
    print(f"Reservas pendientes: {reservas_pendientes.count()}")
    print(f"Reservas completadas: {reservas_completadas.count()}")
    print(f"Total reservas: {total_reservas.count()}")
    for reserva in reservas_completadas:
        print(f"  - ID: {reserva.id_reserva}, Empresa: {reserva.empresa.nombre_empresa}, Usuario: {reserva.usuario.nombre_usuario}")

    # Asegurarse de que cada reserva tenga un atributo legible numero_reserva_display
    def asegurar_numero_display(qs):
        for r in qs:
            nr = getattr(r, 'numero_reserva', None)
            if nr:
                r.numero_reserva_display = nr
            else:
                try:
                    r.numero_reserva_display = f"ANW-{r.id_reserva:08d}"
                except Exception:
                    r.numero_reserva_display = str(getattr(r, 'id_reserva', ''))

    asegurar_numero_display(reservas_pendientes)
    asegurar_numero_display(reservas_completadas)
    
    servicios = Servicio.objects.all() 
    empresas = Empresa.objects.all()

    horas_disponibles = {}
    ocupadas = {
        "fechas": {reserva.fecha for reserva in reservas_pendientes},
        "horas": {}
    }

    # Obtener horas ocupadas de reservas no completadas
    for reserva in reservas_pendientes:
        if reserva.fecha not in ocupadas["horas"]:
            ocupadas["horas"][reserva.fecha] = set()
        ocupadas["horas"][reserva.fecha].add(reserva.hora.strftime('%H:%M'))

    # Filtrar fechas y horas disponibles desde hoy
    for i in range(15):  # Desde hoy hasta 15 días adelante
        fecha = hoy + timedelta(days=i)
        horas_disponibles[fecha] = []

        for h in range(8, 16):  # De 08:00 a 15:00
            hora_formateada_24h = f"{h:02d}:00"

            # Si la fecha es hoy, verifica que la hora no haya pasado
            if fecha == hoy and h < ahora.hour:
                continue  # Ignora horas pasadas para hoy

            # Verifica si la hora está ocupada en la fecha seleccionada
            if hora_formateada_24h not in ocupadas["horas"].get(fecha, set()):
                # Convertir a formato 12h con AM/PM para mostrar al usuario
                hora_12h = convertir_hora_12h(hora_formateada_24h)
                horas_disponibles[fecha].append(hora_12h)

    if request.method == "POST":
        # Manejar la eliminación de reservas
        if 'eliminar' in request.POST:
            reserva_id = request.POST.get('eliminar')
            try:
                reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)  # Filtrar por usuario
                reserva_info = f"{reserva.empresa.nombre_empresa} - {reserva.fecha} {reserva.hora}"
                
                # Si la reserva utiliza suscripción, devolver los servicios utilizados
                if reserva.suscripcion_utilizada and not reserva.es_pago_individual:
                    suscripcion = reserva.suscripcion_utilizada
                    servicios_count = reserva.servicios.count()
                    if servicios_count == 0:
                        servicios_count = 1  # Al menos un servicio por defecto
                    
                    # Restar los servicios del contador (devolver servicios)
                    suscripcion.servicios_utilizados_mes = max(0, suscripcion.servicios_utilizados_mes - servicios_count)
                    suscripcion.save()
                    print(f"✅ Servicios devueltos a la suscripción. Servicios utilizados: {suscripcion.servicios_utilizados_mes}/{suscripcion.plan.cantidad_servicios_mes if suscripcion.plan.cantidad_servicios_mes > 0 else 'Ilimitado'}")
                
                reserva.delete()
                messages.success(request, f'Reserva eliminada con éxito: {reserva_info}')
                
            except Exception as e:
                print(f"Error al eliminar reserva {reserva_id}: {str(e)}")
                messages.error(request, 'Error al eliminar la reserva. Por favor, inténtelo de nuevo.')
            return redirect('citas')
        
        # Manejar la creación o edición de reservas
        if 'id_reserva' in request.POST:  # Para editar
            reserva_id = request.POST.get('id_reserva')
            reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)  # Filtrar por usuario

            # Actualiza la reserva
            reserva.empresa_id = request.POST.get('empresa')  # Actualiza el lugar
            reserva.fecha = request.POST.get('fecha')  # Actualiza la fecha
            # Convertir hora de 12h a 24h antes de guardar
            hora_12h = request.POST.get('hora')
            reserva.hora = convertir_hora_24h(hora_12h)  # Actualiza la hora
            reserva.servicio_id = request.POST.get('servicio')  # Actualiza el servicio
            reserva.save()
            messages.success(request, 'Reserva actualizada con éxito.')
            return redirect('citas')
        else:  # Para crear
            form = ReservaForm(request.POST)
            if form.is_valid():
                # Verificar si el usuario tiene una suscripción activa
                suscripcion_activa = None
                tiene_suscripcion = False
                
                try:
                    from .models import SuscripcionUsuario
                    suscripcion_activa = SuscripcionUsuario.objects.filter(
                        usuario=request.user,
                        estado='activa'
                    ).first()
                    
                    if suscripcion_activa and suscripcion_activa.esta_activa():
                        tiene_suscripcion = True
                        # Verificar si puede usar más servicios este mes
                        if not suscripcion_activa.puede_usar_servicio():
                            messages.error(request, f"Has agotado tus servicios del mes. Servicios restantes: {suscripcion_activa.servicios_restantes()}")
                            return redirect('citas')
                except Exception as e:
                    print(f"Error verificando suscripción: {e}")
                
                reserva = form.save(commit=False)
                reserva.usuario = request.user  # Asignar el usuario actual
                reserva.suscripcion_utilizada = suscripcion_activa if tiene_suscripcion else None
                reserva.es_pago_individual = not tiene_suscripcion
                reserva.save()
                
                # Si tiene suscripción activa, incrementar el contador de servicios utilizados
                if tiene_suscripcion and suscripcion_activa:
                    # Contar los servicios asociados a esta reserva
                    servicios_count = reserva.servicios.count()
                    if servicios_count == 0:
                        servicios_count = 1  # Al menos un servicio por defecto
                    
                    suscripcion_activa.servicios_utilizados_mes += servicios_count
                    suscripcion_activa.save()
                    print(f"✅ Servicios utilizados actualizados: {suscripcion_activa.servicios_utilizados_mes}/{suscripcion_activa.plan.cantidad_servicios_mes if suscripcion_activa.plan.cantidad_servicios_mes > 0 else 'Ilimitado'}")
                
                messages.success(request, 'Reserva creada con éxito.')
                return redirect('citas')
    else:
        form = ReservaForm()

    # Generar lista de fechas disponibles
    fechas_disponibles = [hoy + timedelta(days=i) for i in range(15)] 

    # Renderiza la vista con los datos necesarios
    return render(request, 'reservas/reservas_usuario.html', {
        'reservas_pendientes': reservas_pendientes,
        'reservas_completadas': reservas_completadas,
        'total_reservas': total_reservas,
        'servicios': servicios,
        'empresas': empresas,
        'horas_disponibles': horas_disponibles,
        'fechas_disponibles': fechas_disponibles,
        'hoy': hoy,
    })

# =============================================================================
# NUEVAS VISTAS PARA LA GESTIÓN AVANZADA DE CITAS
# =============================================================================

@login_required
def obtener_detalles_reserva(request, reserva_id):
    """Vista AJAX para obtener detalles completos de una reserva"""
    try:
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)
        
        # Obtener los servicios asociados con su precio aplicado
        servicios_data = []
        total_precio = 0
        total_servicios_adicionales = 0  # Solo servicios que NO son del plan
        
        # Obtener las relaciones ReservaServicio para acceder al precio_aplicado
        reserva_servicios = ReservaServicio.objects.filter(reserva=reserva).select_related('servicio')
        
        for rs in reserva_servicios:
            # Manejar precio_aplicado que puede ser None en la BD
            precio_aplicado_val = rs.precio_aplicado if rs.precio_aplicado is not None else 0
            try:
                precio_aplicado_float = float(precio_aplicado_val)
            except Exception:
                precio_aplicado_float = 0.0

            precio_original_float = float(rs.servicio.precio) if getattr(rs.servicio, 'precio', None) is not None else 0.0

            # Un servicio es gratis si es del plan O si el precio aplicado es 0
            es_gratis = bool(rs.es_servicio_plan) or precio_aplicado_float == 0.0

            servicios_data.append({
                'nombre': rs.servicio.nombre_servicio,
                'descripcion': rs.servicio.descripcion,
                'precio': precio_original_float,
                'precio_aplicado': precio_aplicado_float,
                'es_gratis': es_gratis,
                'es_servicio_plan': rs.es_servicio_plan  # Agregamos este campo también
            })

            total_precio += precio_aplicado_float

            # Solo sumar si NO es servicio del plan
            if not rs.es_servicio_plan:
                total_servicios_adicionales += precio_aplicado_float
        
        # Verificar si todos los servicios son del plan
        todos_servicios_son_plan = all(rs.es_servicio_plan for rs in reserva_servicios)
        
        # Formatear fecha en español
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
        fecha_obj = reserva.fecha
        fecha_formateada = f"{dias_semana[fecha_obj.weekday()]}, {fecha_obj.day} de {meses[fecha_obj.month - 1]} de {fecha_obj.year}"
        
        # Formatear hora a 12h
        hora_formateada = reserva.hora.strftime('%I:%M %p')
        
        reserva_data = {
            'id': reserva.id_reserva,
            'empresa': {
                'id': reserva.empresa.id_empresa,  # Agregado: ID de la empresa
                'nombre': reserva.empresa.nombre_empresa,
                'direccion': reserva.empresa.direccion,
                'telefono': reserva.empresa.telefono,
                'email': reserva.empresa.email
            },
            'fecha': fecha_formateada,
            'fecha_original': str(reserva.fecha),
            'hora': hora_formateada,
            'hora_original': str(reserva.hora),
            'estado': reserva.estado,
            'servicios': servicios_data,
            'total': total_precio,
            'total_servicios_adicionales': total_servicios_adicionales,  # Solo servicios adicionales
            'todos_servicios_son_plan': todos_servicios_son_plan,  # Flag para saber si todos son del plan
            'es_pago_individual': reserva.es_pago_individual,
            'tiene_plan': reserva.suscripcion_utilizada is not None,
            'suscripcion_info': {
                'utilizada': reserva.suscripcion_utilizada is not None,
                'plan_nombre': reserva.suscripcion_utilizada.plan.nombre if reserva.suscripcion_utilizada else None
            } if reserva.suscripcion_utilizada else None
        }
        
        return JsonResponse({
            'success': True,
            'reserva': reserva_data
        })
        
    except Reserva.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Reserva no encontrada'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Error en obtener_detalles_reserva: {e}")
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener detalles: {str(e)}'
        })

@login_required
def cancelar_reserva(request, reserva_id):
    """Vista AJAX para cancelar una reserva"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'})
    
    try:
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)
        
        # Verificar que la reserva se pueda cancelar (no esté completada)
        if reserva.estado == 'completado':
            return JsonResponse({
                'success': False,
                'message': 'No se puede cancelar una reserva que ya está completada'
            })
        
        if reserva.estado == 'cancelada':
            return JsonResponse({
                'success': False,
                'message': 'Esta reserva ya está cancelada'
            })
        
        # Verificar que la cancelación sea con al menos 24 horas de anticipación
        from django.utils import timezone
        ahora = timezone.now()
        fecha_hora_reserva = timezone.make_aware(
            datetime.combine(reserva.fecha, reserva.hora)
        )
        
        if fecha_hora_reserva <= ahora + timedelta(hours=24):
            return JsonResponse({
                'success': False,
                'message': 'Las reservas deben cancelarse con al menos 24 horas de anticipación'
            })
        
        # Si la reserva utiliza suscripción, devolver los servicios
        if reserva.suscripcion_utilizada and not reserva.es_pago_individual:
            suscripcion = reserva.suscripcion_utilizada
            servicios_count = reserva.servicios.count()
            if servicios_count == 0:
                servicios_count = 1
            
            # Restar los servicios del contador (devolver servicios)
            suscripcion.servicios_utilizados_mes = max(0, suscripcion.servicios_utilizados_mes - servicios_count)
            suscripcion.save()
        
        # Cambiar estado a cancelada
        reserva.estado = 'cancelada'
        reserva.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Reserva cancelada exitosamente'
        })
        
    except Reserva.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Reserva no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al cancelar reserva: {str(e)}'
        })

@login_required
def editar_reserva(request, reserva_id):
    """Vista para editar una reserva existente - Solo permite cambio de hora con 12h de anticipación"""
    try:
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)
        
        # Verificar que la reserva se pueda editar
        if reserva.estado != 'pendiente':
            messages.error(request, 'Solo se pueden editar reservas pendientes')
            return redirect('citas')
        
        # Verificar que la edición sea con al menos 12 horas de anticipación
        from django.utils import timezone
        ahora = timezone.now()
        fecha_hora_reserva = timezone.make_aware(
            datetime.combine(reserva.fecha, reserva.hora)
        )
        
        # Calcular cuántas horas faltan para la cita original
        tiempo_hasta_cita = fecha_hora_reserva - ahora
        horas_hasta_cita = tiempo_hasta_cita.total_seconds() / 3600
        
        # Verificar que haya al menos 12 horas hasta la cita original
        if horas_hasta_cita < 12:
            messages.error(
                request, 
                f'Las reservas deben editarse con al menos 12 horas de anticipación. '
                f'Tu cita es en {horas_hasta_cita:.1f} horas. '
                f'Si deseas cambiar la fecha, lugar o servicios, debes cancelar esta cita y crear una nueva.'
            )
            return redirect('citas')
        
        if request.method == 'POST':
            # Solo se permite cambiar la hora, no la fecha ni la empresa
            nueva_hora_str = request.POST.get('hora')
            
            # Validaciones
            if not nueva_hora_str:
                messages.error(request, 'Debes seleccionar una nueva hora')
                return redirect('citas')
            
            try:
                # Parsear la nueva hora (puede venir en formato HH:MM:SS o HH:MM)
                if len(nueva_hora_str.split(':')) == 3:
                    nueva_hora = datetime.strptime(nueva_hora_str, '%H:%M:%S').time()
                else:
                    nueva_hora = datetime.strptime(nueva_hora_str, '%H:%M').time()
                
                # Verificar que la nueva hora sea diferente a la actual
                if nueva_hora == reserva.hora:
                    messages.warning(request, 'La hora seleccionada es la misma que la actual')
                    return redirect('citas')
                
                # Crear datetime para la nueva hora (mismo día y empresa)
                nueva_fecha_hora = timezone.make_aware(
                    datetime.combine(reserva.fecha, nueva_hora)
                )
                
                # Verificar que la nueva hora no haya pasado ya
                if nueva_fecha_hora <= ahora:
                    messages.error(request, 'No puedes seleccionar una hora que ya pasó')
                    return redirect('citas')
                
                # Verificar que haya al menos 12 horas hasta la nueva hora
                tiempo_hasta_nueva_hora = nueva_fecha_hora - ahora
                horas_hasta_nueva_hora = tiempo_hasta_nueva_hora.total_seconds() / 3600
                
                if horas_hasta_nueva_hora < 12:
                    messages.error(
                        request,
                        f'La nueva hora debe tener al menos 12 horas de anticipación. '
                        f'La hora seleccionada solo tiene {horas_hasta_nueva_hora:.1f} horas de anticipación.'
                    )
                    return redirect('citas')
                
                # Verificar disponibilidad de la nueva hora (misma fecha y empresa)
                reserva_existente = Reserva.objects.filter(
                    empresa=reserva.empresa,
                    fecha=reserva.fecha,
                    hora=nueva_hora
                ).exclude(id_reserva=reserva_id).exists()
                
                if reserva_existente:
                    messages.error(request, 'La hora seleccionada ya está ocupada')
                    return redirect('citas')
                
                # Actualizar solo la hora de la reserva
                hora_anterior = reserva.hora.strftime("%H:%M")
                reserva.hora = nueva_hora
                reserva.save()
                
                messages.success(
                    request, 
                    f'¡Cita actualizada exitosamente! Cambiaste de {hora_anterior} a {nueva_hora.strftime("%H:%M")}'
                )
                return redirect('citas')
                
            except ValueError as e:
                messages.error(request, f'Hora inválida proporcionada: {str(e)}')
                return redirect('citas')
        
        # GET request: mostrar formulario de edición
        context = {
            'reserva': reserva,
        }
        return render(request, 'reservas/editar_reserva.html', context)
        
    except Reserva.DoesNotExist:
        messages.error(request, 'Reserva no encontrada')
        return redirect('citas')

@login_required
def repetir_reserva(request, reserva_id):
    """Vista para repetir una reserva completada"""
    try:
        reserva_original = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)
        
        # Solo permitir repetir reservas completadas
        if reserva_original.estado != 'completado':
            messages.error(request, 'Solo se pueden repetir reservas completadas')
            return redirect('citas')
        
        # Redirigir a la página de reservas con los datos pre-cargados
        servicios_ids = list(reserva_original.servicios.values_list('id_servicio', flat=True))
        params = {
            'empresa': reserva_original.empresa.id_empresa,
            'servicios': ','.join(map(str, servicios_ids)),
            'repetir': 'true'
        }
        
        from urllib.parse import urlencode
        url = f"/reservas/?{urlencode(params, doseq=True)}"
        messages.info(request, 'Selecciona la nueva fecha y hora para tu cita')
        return redirect(url)
        
    except Reserva.DoesNotExist:
        messages.error(request, 'Reserva no encontrada')
        return redirect('citas')

@login_required
def calificar_servicio(request, reserva_id):
    """Vista para calificar un servicio completado"""
    try:
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, usuario=request.user)
        
        # Solo permitir calificar reservas completadas
        if reserva.estado != 'completado':
            messages.error(request, 'Solo se pueden calificar reservas completadas')
            return redirect('citas')
        
        if request.method == 'POST':
            calificacion = request.POST.get('calificacion')
            comentario_texto = request.POST.get('comentario', '')
            
            if not calificacion:
                messages.error(request, 'Debes seleccionar una calificación')
                return redirect('calificar_servicio', reserva_id=reserva_id)
            
            try:
                # Crear comentario con calificación
                comentario = Comentario.objects.create(
                    comentario=f"Calificación: {calificacion}/5 estrellas. {comentario_texto}".strip(),
                    usuario=request.user
                )
                
                messages.success(request, 'Gracias por tu calificación!')
                return redirect('citas')
                
            except Exception as e:
                messages.error(request, 'Error al guardar la calificación')
                return redirect('calificar_servicio', reserva_id=reserva_id)
        
        # GET request: mostrar formulario de calificación
        context = {
            'reserva': reserva,
        }
        return render(request, 'servicios/calificar_servicio.html', context)
        
    except Reserva.DoesNotExist:
        messages.error(request, 'Reserva no encontrada')
        return redirect('citas')

@login_required
def get_empresas_verificadas(request):
    """Vista AJAX para obtener empresas verificadas"""
    try:
        empresas = Empresa.objects.filter(verificada=True).values(
            'id_empresa', 'nombre_empresa', 'direccion'
        )
        
        empresas_data = []
        for empresa in empresas:
            empresas_data.append({
                'id': empresa['id_empresa'],
                'nombre': empresa['nombre_empresa'],
                'direccion': empresa['direccion']
            })
        
        return JsonResponse({
            'success': True,
            'empresas': empresas_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al obtener empresas: {str(e)}'
        })



@login_required
def contacto(request):
    # Obtener todos los servicios para el dropdown
    servicios = Servicio.objects.all().order_by('nombre_servicio')
    
    if request.method == 'POST':
        # Obtener todos los datos del formulario PQRS
        tipo_pqrs = request.POST.get('tipo_pqrs')
        urgencia = request.POST.get('urgencia')
        nombre_contacto = request.POST.get('nombre_contacto')
        email_contacto = request.POST.get('email_contacto')
        servicio_relacionado = request.POST.get('servicio_relacionado')
        contenido = request.POST.get('contenido')
        acepto_terminos = request.POST.get('acepto_terminos') == 'on'

        # Validar campos requeridos
        if not all([tipo_pqrs, contenido, acepto_terminos]):
            messages.error(request, "Por favor, completa todos los campos requeridos y acepta los términos.")
            return redirect('contacto')

        # Crear el mensaje PQRS
        mensaje = MensajeQueja(
            tipo_pqrs=tipo_pqrs,
            urgencia=urgencia or 'media',
            nombre_contacto=nombre_contacto,
            email_contacto=email_contacto,
            servicio_relacionado=servicio_relacionado,
            contenido=contenido,
            acepto_terminos=acepto_terminos,
            usuario=request.user
        )
        
        # Si se seleccionó un servicio de la base de datos, asociarlo
        if servicio_relacionado and servicio_relacionado.startswith('servicio_'):
            try:
                servicio_id = servicio_relacionado.replace('servicio_', '')
                servicio = Servicio.objects.get(id_servicio=servicio_id)
                mensaje.servicio_bd = servicio
            except Servicio.DoesNotExist:
                pass
        
        mensaje.save()
        
        # Mensaje de éxito personalizado según el tipo de PQRS
        tipo_display = dict(MensajeQueja.TIPOS_PQRS).get(tipo_pqrs, 'solicitud')
        messages.success(
            request, 
            f"Tu {tipo_display.lower()} ha sido enviada exitosamente. "
            f"Número de radicado: {mensaje.numero_radicado}. "
            f"Será respondida en el menor tiempo posible."
        )
        return redirect('contacto')

    context = {
        'servicios': servicios,
    }
    return render(request, 'pages_informativas/contacto.html', context) 

def resetCorreo(request):
    return render(request, 'auth/reset_correo.html')
def resetContrasena(request):
    return render(request, 'reset_contrasena.html')


@login_required
def comentarios(request):
    if request.method == 'POST':
        form = ComentarioClienteForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)  
            comentario.usuario = request.user  
            comentario.save() 
            
            # Si es una solicitud AJAX, devolver JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Tu comentario ha sido publicado exitosamente.',
                    'comentario_id': comentario.id_comentario,
                    'usuario': comentario.usuario.username,
                    'texto': comentario.comentario,
                    'fecha': comentario.fecha.strftime('%d/%m/%Y %H:%M')
                })
            else:
                messages.success(request, "Tu comentario ha sido publicado exitosamente.")
                return redirect('comentarios')
        else:
            # Si es AJAX y hay errores
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'message': 'Hubo un error al procesar tu comentario.'
                })

    else:
        form = ComentarioClienteForm()

    return render(request, 'comentarios/comentarios.html', {'form': form})






############################################################ comienzo de crud


def login_crud(request):
    # Si ya está logueado como admin, redirigir al home de admin
    if request.user.is_authenticated and hasattr(request.user, 'rol') and request.user.rol == 'admin':
        return redirect('homecrud')
    
    if request.method == 'POST':
        nombre_usuario = request.POST['nombre_usuario']
        contrasena = request.POST['contrasena']
        tipo_usuario = request.POST.get('tipo_usuario', 'admin')  # Por defecto admin
        
        print(f"🔍 Intento de login - Tipo: {tipo_usuario}, Usuario: {nombre_usuario}")
        
        if tipo_usuario == 'admin':
            # Autenticación del administrador (sin sistema de bloqueo por intentos)
            usuario = authenticate(request, username=nombre_usuario, password=contrasena)
            
            if usuario is not None:
                # Verificar si el usuario tiene rol de admin y está activo
                if hasattr(usuario, 'rol') and usuario.rol == 'admin' and usuario.is_active:
                    auth_login(request, usuario)  # Iniciar sesión con el usuario autenticado
                    messages.success(request, f'Bienvenido administrador, {usuario.nombre_usuario}!')
                    return redirect('homecrud')
                elif not usuario.is_active:
                    messages.error(request, 'Tu cuenta de administrador ha sido desactivada. Contacta al superadministrador.')
                    return redirect('logincrud')
                else:
                    messages.error(request, 'Acceso denegado. Solo los administradores pueden acceder.')
                    return redirect('logincrud')
            else:
                # Login fallido - solo mostrar mensaje genérico (sin contar intentos)
                messages.error(request, 'Usuario o contraseña de administrador incorrectos.')
                return redirect('logincrud')
                
        elif tipo_usuario == 'empresa':
            # Autenticación de empresa
            try:
                # Buscar la empresa por email (usando nombre_usuario como email)
                empresa = Empresa.objects.get(email=nombre_usuario)
                
                # Verificar si la cuenta está activa
                if not empresa.is_active:
                    messages.error(request, 'Su cuenta empresarial ha sido desactivada. Contacte al administrador para más información.')
                    return redirect('logincrud')
                
                # Verificar si puede intentar hacer login (no está bloqueada)
                if not empresa.can_attempt_login():
                    if empresa.lockout_time:
                        time_since_lockout = timezone.now() - empresa.lockout_time
                        remaining_minutes = max(0, 15 - int(time_since_lockout.total_seconds() / 60))
                        messages.error(request, f'Su cuenta empresarial está temporalmente bloqueada por seguridad debido a múltiples intentos fallidos. Intente nuevamente en {remaining_minutes} minutos.')
                    else:
                        messages.error(request, 'Su cuenta empresarial está temporalmente bloqueada por seguridad. Contacte al administrador.')
                    return redirect('logincrud')
                
                # Verificar la contraseña
                if check_password(contrasena, empresa.contrasena):
                    # Resetear intentos fallidos en login exitoso
                    empresa.reset_failed_attempts()
                    
                    # Verificar si la empresa está verificada
                    if empresa.verificada:
                        # Guardar información de la empresa en la sesión
                        request.session['empresa_id'] = empresa.id_empresa
                        request.session['empresa_nombre'] = empresa.nombre_empresa
                        request.session['empresa_email'] = empresa.email
                        request.session['es_empresa'] = True
                        
                        messages.success(request, f'Bienvenida empresa {empresa.nombre_empresa}!')
                        return redirect('home_empresas')  # Redirigir al home de empresas
                    else:
                        mensaje_verificacion = (
                            f'¡Hola {empresa.nombre_empresa}! Su cuenta empresarial fue creada exitosamente el {empresa.fecha_registro.strftime("%d/%m/%Y a las %H:%M")}. '
                            'Sin embargo, aún no ha sido verificada por nuestro equipo administrativo. '
                            'Para garantizar la seguridad y calidad de nuestros servicios, todas las empresas deben pasar por un proceso de verificación antes de poder acceder al sistema. '
                            'Nuestros administradores revisarán su información y aprobarán su cuenta en un plazo máximo de 24-48 horas hábiles. '
                            'Una vez verificada, podrá acceder completamente a todas las funcionalidades del sistema. '
                            f'Si tiene alguna pregunta, puede contactarnos. Email de contacto: {empresa.email}'
                        )
                        messages.warning(request, mensaje_verificacion)
                        return redirect('logincrud')
                else:
                    # Login fallido - incrementar intentos fallidos
                    empresa.increment_failed_attempts()
                    
                    # Mostrar mensaje específico según los intentos restantes
                    remaining_attempts = empresa.get_remaining_attempts()
                    
                    if empresa.lockout_time:  # Si tiene tiempo de bloqueo activo
                        messages.error(request, 'Ha excedido el número máximo de intentos de login empresarial. Su cuenta ha sido bloqueada temporalmente por 15 minutos por seguridad.')
                    elif remaining_attempts <= 3 and remaining_attempts > 0:
                        messages.warning(request, f'Contraseña de empresa incorrecta. Le quedan {remaining_attempts} intentos antes de que su cuenta sea bloqueada temporalmente.')
                    else:
                        messages.error(request, 'Contraseña de empresa incorrecta.')
                    
                    return redirect('logincrud')
                    
            except Empresa.DoesNotExist:
                messages.error(request, 'Email de empresa no encontrado. Verifique el correo electrónico.')
                return redirect('logincrud')
            except Exception as e:
                print(f"❌ Error en login de empresa: {str(e)}")
                messages.error(request, 'Error interno. Intente nuevamente.')
                return redirect('logincrud')
        else:
            messages.error(request, 'Tipo de usuario no válido.')
            return redirect('logincrud')
    
    return render(request, 'auth/login_crud.html')
    

def logout_view(request):
    auth_logout(request)  # Cerrar sesión
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('home')  # Redirigir a la página de home normal


# ==================== RECUPERACIÓN DE CONTRASEÑA PARA EMPRESAS ====================

def empresa_password_reset(request):
    """Vista para solicitar el restablecimiento de contraseña de empresa"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            empresa = Empresa.objects.get(email=email)
            
            # Generar token único para el reset
            token = str(uuid.uuid4())
            empresa.token_reset = token
            empresa.save()
            
            # Crear URL de reset
            reset_url = request.build_absolute_uri(f'/empresa/reset/{token}/')
            
            # Enviar correo con el enlace
            subject = 'Recuperación de contraseña - AUTONEW Empresa'
            html_message = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #3b82f6 0%, #1e3a8a 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
                    .content {{ background: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }}
                    .button {{ display: inline-block; padding: 14px 28px; background: linear-gradient(135deg, #3b82f6 0%, #1e3a8a 100%); color: white; text-decoration: none; border-radius: 8px; font-weight: 600; margin: 20px 0; }}
                    .footer {{ text-align: center; padding: 20px; color: #6b7280; font-size: 14px; }}
                    .warning {{ background: #fef3c7; padding: 15px; border-left: 4px solid #f59e0b; margin: 20px 0; border-radius: 4px; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>🔐 Recuperación de Contraseña</h1>
                        <p>AUTONEW - Panel Empresarial</p>
                    </div>
                    <div class="content">
                        <h2>Hola, {empresa.nombre_empresa}</h2>
                        <p>Hemos recibido una solicitud para restablecer la contraseña de tu cuenta empresarial.</p>
                        <p>Para crear una nueva contraseña, haz clic en el siguiente botón:</p>
                        <div style="text-align: center;">
                            <a href="{reset_url}" class="button">Restablecer Contraseña</a>
                        </div>
                        <p>O copia y pega este enlace en tu navegador:</p>
                        <p style="word-break: break-all; background: white; padding: 10px; border-radius: 4px; border: 1px solid #e5e7eb;">
                            {reset_url}
                        </p>
                        <div class="warning">
                            <strong>⚠️ Importante:</strong>
                            <ul>
                                <li>Este enlace expirará en 24 horas por seguridad</li>
                                <li>Si no solicitaste este cambio, ignora este correo</li>
                                <li>Tu contraseña actual permanecerá activa hasta que establezcas una nueva</li>
                            </ul>
                        </div>
                    </div>
                    <div class="footer">
                        <p>Este es un correo automático, por favor no respondas a este mensaje.</p>
                        <p>&copy; 2025 AUTONEW. Todos los derechos reservados.</p>
                    </div>
                </div>
            </body>
            </html>
            """
            
            plain_message = f"""
            Recuperación de Contraseña - AUTONEW Empresa
            
            Hola, {empresa.nombre_empresa}
            
            Hemos recibido una solicitud para restablecer la contraseña de tu cuenta empresarial.
            
            Para crear una nueva contraseña, visita el siguiente enlace:
            {reset_url}
            
            IMPORTANTE:
            - Este enlace expirará en 24 horas por seguridad
            - Si no solicitaste este cambio, ignora este correo
            - Tu contraseña actual permanecerá activa hasta que establezcas una nueva
            
            Este es un correo automático, por favor no respondas a este mensaje.
            
            © 2025 AUTONEW. Todos los derechos reservados.
            """
            
            try:
                send_mail(
                    subject,
                    plain_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [empresa.email],
                    html_message=html_message,
                    fail_silently=False,
                )
                messages.success(request, 'Se ha enviado un correo con las instrucciones para restablecer tu contraseña.')
            except Exception as e:
                print(f"❌ Error al enviar correo: {str(e)}")
                messages.error(request, 'Error al enviar el correo. Por favor, intenta nuevamente.')
                return redirect('empresa_password_reset')
            
            return redirect('empresa_password_reset_done')
            
        except Empresa.DoesNotExist:
            # Por seguridad, no revelamos si el email existe o no
            messages.success(request, 'Si el correo está registrado, recibirás las instrucciones para restablecer tu contraseña.')
            return redirect('empresa_password_reset_done')
    
    return render(request, 'auth/empresa/reset_correo.html')


def empresa_password_reset_done(request):
    """Vista de confirmación de que se envió el correo"""
    return render(request, 'auth/empresa/reset_correo_enviado.html')


def empresa_password_reset_confirm(request, token):
    """Vista para confirmar el token y establecer nueva contraseña"""
    try:
        empresa = Empresa.objects.get(token_reset=token)
        
        if request.method == 'POST':
            nueva_contrasena = request.POST.get('nueva_contrasena')
            confirmar_contrasena = request.POST.get('confirmar_contrasena')
            
            # Validaciones
            if not nueva_contrasena or not confirmar_contrasena:
                messages.error(request, 'Todos los campos son obligatorios.')
                return render(request, 'auth/empresa/reset_contrasena.html', {'token': token})
            
            if nueva_contrasena != confirmar_contrasena:
                messages.error(request, 'Las contraseñas no coinciden.')
                return render(request, 'auth/empresa/reset_contrasena.html', {'token': token})
            
            if len(nueva_contrasena) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
                return render(request, 'auth/empresa/reset_contrasena.html', {'token': token})
            
            # Actualizar contraseña
            empresa.contrasena = make_password(nueva_contrasena)
            empresa.token_reset = None  # Invalidar el token
            empresa.save()
            
            # No agregar mensaje aquí porque la página de éxito ya muestra la confirmación
            return redirect('empresa_password_reset_complete')
        
        return render(request, 'auth/empresa/reset_contrasena.html', {'token': token, 'empresa': empresa})
        
    except Empresa.DoesNotExist:
        messages.error(request, 'El enlace de recuperación es inválido o ha expirado.')
        return redirect('empresa_password_reset')


def empresa_password_reset_complete(request):
    """Vista de confirmación de que se restableció la contraseña"""
    return render(request, 'auth/empresa/reset_completo.html')


@admin_required
def comentarios_crud(request):
    comentarios = Comentario.objects.all()
    
    # Calcular estadísticas
    total_comentarios = comentarios.count()
    total_usuarios = comentarios.values('usuario').distinct().count()

    if request.method == "POST":
        # Manejar la eliminación de comenatrios
        if 'eliminar' in request.POST:
            comentario_id = request.POST.get('eliminar')
            comentario = get_object_or_404(Comentario, id_comentario=comentario_id)
            comentario.delete()
            messages.error(request, 'El comentario se a eliminado.')
            return redirect('comentarioscrud')
    else:
        form = ComentarioForm()
    
    context = {
        'comentarios': comentarios, 
        'form': form,
        'total_comentarios': total_comentarios,
        'total_usuarios': total_usuarios,
    }
    
    return render(request, 'comentarios/comentarios_crud.html', context)




@admin_required
def quejas_crud(request):
    quejas = MensajeQueja.objects.all().order_by('-fecha_envio')
    
    # Calcular estadísticas
    from django.db.models import Count
    total_count = quejas.count()
    resueltas_count = quejas.filter(estado__in=['resuelto', 'cerrado']).count()
    efectividad = round((resueltas_count / total_count * 100), 1) if total_count > 0 else 0
    
    estadisticas = {
        'total': total_count,
        'por_tipo': quejas.values('tipo_pqrs').annotate(count=Count('tipo_pqrs')),
        'por_estado': quejas.values('estado').annotate(count=Count('estado')),
        'por_urgencia': quejas.values('urgencia').annotate(count=Count('urgencia')),
        'pendientes': quejas.filter(estado__in=['recibido', 'en_proceso']).count(),
        'resueltas': resueltas_count,
        'efectividad': efectividad,
    }

    if request.method == "POST":
        # Manejar la eliminación de quejas/PQRS
        if 'eliminar' in request.POST:
            queja_id = request.POST.get('eliminar')
            queja = get_object_or_404(MensajeQueja, id_mensaje=queja_id)
            queja.delete()
            messages.error(request, 'La solicitud PQRS ha sido eliminada.')
            return redirect('quejascrud')

    if request.method == "POST":
        # Manejar la respuesta a la queja
        if 'id_reserva' in request.POST:
            respuesta = request.POST.get('respuesta')
            queja_id = request.POST.get('id_reserva')
            queja = get_object_or_404(MensajeQueja, id_mensaje=queja_id)

            # Obtener el usuario asociado a la queja
            if queja.usuario:
                usuario = queja.usuario
                
                # Guardar la respuesta
                queja.respuesta = respuesta
                queja.estado = 'resuelto'
                queja.fecha_respuesta = timezone.now()
                queja.save()

                messages.success(request, 'La respuesta se ha enviado.')

                # Enviar mensaje de WhatsApp (código existente)
                try:
                    conn = http.client.HTTPSConnection("kqqk31.api.infobip.com")
                    payload = json.dumps({
                        "messages": [
                            {
                                "from": "447860099299",
                                "to": usuario.telefono,
                                "messageId": "c2dbb13f-2a4a-48d7-97c2-085d5d3d6108",
                                "content": {
                                    "templateName": "message_test",
                                    "templateData": {
                                        "body": {
                                            "placeholders": [respuesta]
                                        }
                                    },
                                    "language": "es"
                                }
                            }
                        ]
                    })
                    headers = {
                        'Authorization': 'App e5eb011e80db665606dd35c15e897846-c8cf84ba-fd3f-401e-a267-7e95cc6bce48',
                        'Content-Type': 'application/json',
                        'Accept': 'application/json'
                    }
                    conn.request("POST", "/whatsapp/1/message/template", payload, headers)
                    res = conn.getresponse()
                    data = res.read()
                    print(data.decode("utf-8"))
                except Exception as e:
                    print("Error al enviar mensaje:", e)
                    messages.error(request, 'Error al enviar el mensaje de WhatsApp.')
            else:
                messages.error(request, 'No se puede enviar respuesta: solicitud sin usuario asociado.')

            return redirect('quejascrud')
    else:
        form = QuejaForm()
    
    context = {
        'quejas': quejas, 
        'form': form,
        'estadisticas': estadisticas
    }
    return render(request, 'comentarios/quejas_crud.html', context)



@admin_required
def usuarios_crud(request):
    from django.core.paginator import Paginator
    from django.http import JsonResponse
    
    # Obtener el tipo de usuarios a mostrar (activos, inactivos o bloqueados)
    tab = request.GET.get('tab', 'activos')
    
    if tab == 'inactivos':
        usuarios = Usuario.objects.filter(is_active=False)
    elif tab == 'bloqueados':
        usuarios = Usuario.objects.filter(lockout_time__isnull=False)  # Usuarios con bloqueo temporal activo
    else:
        usuarios = Usuario.objects.filter(is_active=True)
    
    form = UsuariosForm() 

    # Calcular estadísticas GENERALES (todos los usuarios)
    todos_usuarios = Usuario.objects.all()
    total_usuarios = todos_usuarios.filter(is_active=True).count()
    total_usuarios_inactivos = todos_usuarios.filter(is_active=False).count()
    total_admins = todos_usuarios.filter(rol='admin', is_active=True).count()
    total_clientes = todos_usuarios.filter(rol='cliente', is_active=True).count()
    total_usuarios_bloqueados = todos_usuarios.filter(lockout_time__isnull=False).count()  # Bloqueados temporalmente
    usuarios_con_intentos_fallidos = todos_usuarios.filter(failed_login_attempts__gt=0).count()

    # Filtrar según los parámetros de búsqueda
    nombre_usuario = request.GET.get('nombre_usuario', '')
    correo = request.GET.get('correo', '')
    telefono = request.GET.get('telefono', '')
    rol = request.GET.get('rol', '')

    if nombre_usuario:
        usuarios = usuarios.filter(nombre_usuario__icontains=nombre_usuario)
    if correo:
        usuarios = usuarios.filter(correo__icontains=correo)
    if telefono:
        usuarios = usuarios.filter(telefono__icontains=telefono)
    if rol:
        usuarios = usuarios.filter(rol=rol)

    # Implementar paginación
    page_number = request.GET.get('page', 1)
    paginator = Paginator(usuarios, 50)  # 50 usuarios por página
    usuarios_paginated = paginator.get_page(page_number)

    # Si es una petición AJAX para cargar más usuarios
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        usuarios_data = []
        for usuario in usuarios_paginated:
            usuarios_data.append({
                'id_usuario': usuario.id_usuario,
                'nombre_completo': usuario.nombre_completo,
                'nombre_usuario': usuario.nombre_usuario,
                'correo': usuario.correo,
                'telefono': usuario.telefono,
                'direccion': usuario.direccion,
                'rol': usuario.rol,
                'is_active': usuario.is_active,
            })
        
        return JsonResponse({
            'usuarios': usuarios_data,
            'has_next': usuarios_paginated.has_next(),
            'has_previous': usuarios_paginated.has_previous(),
            'total_pages': paginator.num_pages,
            'current_page': usuarios_paginated.number,
        })

    if request.method == "POST":
        # Manejar la inactivación de los usuarios
        if 'eliminar' in request.POST:
            usuario_id  = request.POST.get('eliminar')
            usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
            usuario.is_active = False
            usuario.save()
            messages.success(request, 'El usuario ha sido inactivado.')
            return redirect('usuarioscrud')
        
        # Manejar la reactivación de los usuarios
        if 'reactivar' in request.POST:
            usuario_id  = request.POST.get('reactivar')
            usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
            usuario.is_active = True
            usuario.save()
            messages.success(request, 'El usuario ha sido reactivado.')
            return redirect('/usuarioscrud/?tab=activos')
        
        # Manejar la creación de nuevos usuarios
        if 'nombre_completo' in request.POST and 'id_usuario' not in request.POST:
            contrasena1 = request.POST.get('contrasena1')
            contrasena2= request.POST.get('contrasena2')
            nombre_usuario = request.POST.get('nombre_usuario')
            correo = request.POST.get('correo')
            telefono = request.POST.get('telefono')

            # Verificar si el nombre de usuario ya está en uso
            if Usuario.objects.filter(nombre_usuario=nombre_usuario).exists():
                messages.error(request, 'El nombre de usuario ya está en uso.')
            # Verificar si el correo ya está en uso
            elif Usuario.objects.filter(correo=correo).exists():
                messages.error(request, 'El correo ya está en uso.')
            else:
                # Concatenar el prefijo "57" al teléfono ingresado
                telefono_completo = "57" + telefono
                if contrasena1 == contrasena2:
                    nuevo_usuario = Usuario(
                        nombre_completo=request.POST.get('nombre_completo'),
                        nombre_usuario=nombre_usuario,
                        correo=correo,
                        telefono=telefono_completo,  # Guardar el número completo
                        direccion=request.POST.get('direccion'),
                        rol=request.POST.get('rol'),
                        password=make_password(contrasena1)  # Hash de la contraseña
                    )
                    nuevo_usuario.save()
                    messages.success(request, 'El usuario ha sido creado.')
                    return redirect('usuarioscrud')
                else:
                    messages.error(request, 'Las contraseñas no coinciden.')
                    return redirect('usuarioscrud')

        # Manejar la actualización de usuarios
        if 'id_usuario' in request.POST:
            usuario_id = request.POST.get('id_usuario')
            usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
            usuario.nombre_completo = request.POST.get('nombre_completo')
            usuario.nombre_usuario = request.POST.get('nombre_usuario')
            usuario.correo = request.POST.get('correo')
            usuario.direccion = request.POST.get('direccion')
            usuario.rol = request.POST.get('rol')

            # Actualizar la contraseña solo si se proporciona una nueva
            contrasena11 = request.POST.get('contrasena11')
            contrasena22= request.POST.get('contrasena22')

            # Solo actualizar la contraseña si el campo no está vacío
            if contrasena11:
                if contrasena11 == contrasena22:
                    usuario.password = make_password(contrasena11)
                else:
                    messages.error(request, 'Las contraseñas no coinciden.')
                    return redirect('usuarioscrud')

            # Guardar los cambios en el usuario
            usuario.save()
            messages.success(request, 'El usuario ha sido actualizado.')
            return redirect('usuarioscrud')

    else:
        form = UsuariosForm()

    # Contexto con estadísticas calculadas
    context = {
        'usuarios': usuarios_paginated, 
        'form': form,
        'total_usuarios': total_usuarios,
        'total_usuarios_inactivos': total_usuarios_inactivos,
        'total_usuarios_bloqueados': total_usuarios_bloqueados,
        'usuarios_con_intentos_fallidos': usuarios_con_intentos_fallidos,
        'total_admins': total_admins,
        'total_clientes': total_clientes,
        'tab_actual': tab,
        'paginator': paginator,
        'page_obj': usuarios_paginated,
        'total_filtered': paginator.count,
    }

    return render(request, 'usuarios/usuarios_crud.html', context)

@admin_required
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
    
    if request.method == 'POST':
        form = UsuariosForm(request.POST, instance=usuario)
        if form.is_valid():
            # Manejar la actualización de contraseña si se proporciona
            contrasena1 = request.POST.get('contrasena1')
            contrasena2 = request.POST.get('contrasena2')
            
            # Solo actualizar la contraseña si se proporciona una nueva
            if contrasena1:
                if contrasena1 == contrasena2:
                    usuario.password = make_password(contrasena1)
                else:
                    messages.error(request, 'Las contraseñas no coinciden.')
                    return render(request, 'editar_usuario.html', {'form': form, 'usuario': usuario})
            
            # Guardar el formulario sin la contraseña primero
            usuario_actualizado = form.save(commit=False)
            
            # Concatenar el prefijo "57" al teléfono si no lo tiene ya
            telefono = usuario_actualizado.telefono
            if telefono and not telefono.startswith('57'):
                usuario_actualizado.telefono = '57' + telefono
            
            usuario_actualizado.save()
            messages.success(request, 'El usuario ha sido actualizado correctamente.')
            return redirect('usuarioscrud')
    else:
        # Remover el prefijo "57" del teléfono para mostrar solo el número local
        initial_data = {
            'nombre_completo': usuario.nombre_completo,
            'nombre_usuario': usuario.nombre_usuario,
            'correo': usuario.correo,
            'telefono': usuario.telefono[2:] if usuario.telefono and usuario.telefono.startswith('57') else usuario.telefono,
            'direccion': usuario.direccion,
            'rol': usuario.rol,
        }
        form = UsuariosForm(initial=initial_data)
    
    return render(request, 'usuarios/editar_usuario.html', {'form': form, 'usuario': usuario})

@admin_required
def desbloquear_usuario(request, usuario_id):
    """Vista para que un administrador pueda desbloquear a un usuario"""
    usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
    
    if request.method == 'POST':
        usuario.reset_failed_attempts()
        messages.success(request, f'El usuario {usuario.nombre_usuario} ha sido desbloqueado exitosamente.')
        return redirect('usuarioscrud')
    
    return redirect('usuarioscrud')

@admin_required
def citas_crud(request):
    # Auto-completar reservas vencidas (más de 4 horas después de la cita)
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Q
    
    ahora = timezone.now()
    hace_4_horas = ahora - timedelta(hours=4)
    
    # Buscar y actualizar reservas pendientes que ya pasaron más de 4 horas
    # Usando una consulta más eficiente con datetime combinado
    reservas_a_completar = Reserva.objects.filter(
        estado='pendiente'
    )
    
    # Filtrar manualmente las que ya pasaron 4 horas
    reservas_vencidas = []
    for reserva in reservas_a_completar:
        try:
            # Crear datetime combinando fecha y hora de la reserva
            fecha_hora_reserva = timezone.make_aware(
                datetime.combine(reserva.fecha, reserva.hora),
                timezone.get_current_timezone()
            )
            
            # Verificar si ya pasaron 4 horas
            if fecha_hora_reserva <= hace_4_horas:
                reservas_vencidas.append(reserva.id_reserva)
        except Exception:
            # Si hay error con alguna fecha, continuar con la siguiente
            continue
    
    # Actualizar todas las reservas vencidas de una vez
    reservas_actualizadas = 0
    if reservas_vencidas:
        reservas_actualizadas = Reserva.objects.filter(
            id_reserva__in=reservas_vencidas,
            estado='pendiente'
        ).update(estado='completado')
    
    # Mostrar mensaje si se completaron reservas automáticamente
    if reservas_actualizadas > 0:
        messages.info(
            request, 
            f"✅ {reservas_actualizadas} reserva{'s' if reservas_actualizadas != 1 else ''} fueron marcadas automáticamente como completadas por haber pasado más de 4 horas."
        )
        # Log para debugging (opcional)
        print(f"🔄 Auto-completado: {reservas_actualizadas} reservas marcadas como completadas automáticamente")
    
    hoy = ahora.date()
    
    # Inicializar el queryset de reservas (ordenadas por más recientes primero)
    reservas = Reserva.objects.select_related('empresa', 'usuario').prefetch_related('reservaservicio_set__servicio').order_by('-fecha', '-hora')
    servicios = Servicio.objects.all() 
    empresas = Empresa.objects.all()
    
    # Aplicar filtros basados en los parámetros GET
    filtros_aplicados = {}
    
    # Filtro por empresa
    empresa_id = request.GET.get('empresa')
    if empresa_id:
        reservas = reservas.filter(empresa_id=empresa_id)
        filtros_aplicados['empresa'] = empresa_id
    
    # Filtro por servicio
    servicio_id = request.GET.get('servicio')
    if servicio_id:
        reservas = reservas.filter(servicios__id_servicio=servicio_id)
        filtros_aplicados['servicio'] = servicio_id
    
    # Filtro por estado
    estado = request.GET.get('estado')
    if estado:
        if estado == 'no_completado':
            reservas = reservas.filter(estado='pendiente')
        elif estado == 'completado':
            reservas = reservas.filter(estado='completado')
        elif estado == 'cancelada':
            reservas = reservas.filter(estado='cancelada')
        filtros_aplicados['estado'] = estado
    # Nota: Removido el filtro por defecto para mostrar todas las citas
    
    # Filtro por cliente (nombre)
    cliente = request.GET.get('cliente')
    if cliente:
        reservas = reservas.filter(usuario__nombre_completo__icontains=cliente)
        filtros_aplicados['cliente'] = cliente
    
    # Filtro por fecha desde
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        reservas = reservas.filter(fecha__gte=fecha_desde)
        filtros_aplicados['fecha_desde'] = fecha_desde
    
    # Filtro por fecha hasta
    fecha_hasta = request.GET.get('fecha_hasta')
    if fecha_hasta:
        reservas = reservas.filter(fecha__lte=fecha_hasta)
        filtros_aplicados['fecha_hasta'] = fecha_hasta
    
    # Filtro por nombre de usuario
    usuario = request.GET.get('usuario')
    if usuario:
        reservas = reservas.filter(usuario__nombre_usuario__icontains=usuario)
        filtros_aplicados['usuario'] = usuario
    
    # Filtro por ID de cita
    id_cita = request.GET.get('id_cita')
    if id_cita:
        reservas = reservas.filter(id_reserva=id_cita)
        filtros_aplicados['id_cita'] = id_cita

    horas_disponibles = {}
    ocupadas = {
        "fechas": {reserva.fecha for reserva in reservas},
        "horas": {}
    }

    # Obtener horas ocupadas
    for reserva in reservas:
        if reserva.fecha not in ocupadas["horas"]:
            ocupadas["horas"][reserva.fecha] = set()
        ocupadas["horas"][reserva.fecha].add(reserva.hora.strftime('%H:%M'))

    # Filtrar fechas y horas disponibles desde hoy
    for i in range(15):  # Desde hoy hasta 15 días adelante
        fecha = hoy + timedelta(days=i)
        horas_disponibles[fecha] = []

        for h in range(8, 16):  # De 08:00 a 15:00
            hora_formateada_24h = f"{h:01}:00"

            # Si la fecha es hoy, verifica que la hora no haya pasado
            if fecha == hoy and h < ahora.hour:
                continue  # Ignora horas pasadas para hoy

            # Verifica si la hora está ocupada en la fecha seleccionada
            if hora_formateada_24h not in ocupadas["horas"].get(fecha, set()):
                # Convertir a formato 12h con AM/PM para mostrar al usuario
                hora_12h = convertir_hora_12h(hora_formateada_24h)
                horas_disponibles[fecha].append(hora_12h)

    if request.method == "POST":
        # Manejar la cancelación de reservas
        if 'eliminar' in request.POST:
            reserva_id = request.POST.get('eliminar')
            reserva = get_object_or_404(Reserva, id_reserva=reserva_id)
            
            # Si la reserva utiliza suscripción, devolver los servicios utilizados
            if reserva.suscripcion_utilizada and not reserva.es_pago_individual:
                suscripcion = reserva.suscripcion_utilizada
                servicios_count = reserva.servicios.count()
                if servicios_count == 0:
                    servicios_count = 1  # Al menos un servicio por defecto
                
                # Restar los servicios del contador (devolver servicios)
                suscripcion.servicios_utilizados_mes = max(0, suscripcion.servicios_utilizados_mes - servicios_count)
                suscripcion.save()
                print(f"✅ Servicios devueltos a la suscripción (Admin). Servicios utilizados: {suscripcion.servicios_utilizados_mes}/{suscripcion.plan.cantidad_servicios_mes if suscripcion.plan.cantidad_servicios_mes > 0 else 'Ilimitado'}")
            
            # Cambiar estado a cancelada en lugar de eliminar
            reserva.estado = 'cancelada'
            reserva.save()
            messages.success(request, 'Reserva cancelada con éxito.')
            return redirect('citascrud')

        # Manejar la creación o edición de reservas
        if 'id_reserva' in request.POST:  # Para editar
            reserva_id = request.POST.get('id_reserva')
            reserva = get_object_or_404(Reserva, id_reserva=reserva_id)

            # Actualizar los datos de la reserva
            empresa_id = request.POST.get('empresa')
            fecha = request.POST.get('fecha')
            hora_12h = request.POST.get('hora')
            servicio_id = request.POST.get('servicio')

            if empresa_id:
                empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
                reserva.empresa = empresa
            
            if fecha:
                reserva.fecha = fecha
            
            if hora_12h:
                # Convertir hora de 12h a 24h antes de guardar
                reserva.hora = convertir_hora_24h(hora_12h)
            
            reserva.save()

            # Actualizar el servicio asociado
            if servicio_id:
                servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
                # Eliminar la relación anterior
                ReservaServicio.objects.filter(reserva=reserva).delete()
                # Crear la nueva relación
                ReservaServicio.objects.create(reserva=reserva, servicio=servicio)

            messages.success(request, 'Reserva actualizada con éxito.')
            return redirect('citascrud')
        else:  # Para crear
            form = ReservaForm(request.POST)
            if form.is_valid():
                reserva = form.save(commit=False)
                # En el caso del admin, usar el usuario logueado como fallback
                if not reserva.usuario:
                    reserva.usuario = request.user
                
                # Verificar si el usuario de la reserva tiene una suscripción activa
                suscripcion_activa = None
                tiene_suscripcion = False
                
                try:
                    from .models import SuscripcionUsuario
                    suscripcion_activa = SuscripcionUsuario.objects.filter(
                        usuario=reserva.usuario,
                        estado='activa'
                    ).first()
                    
                    if suscripcion_activa and suscripcion_activa.esta_activa():
                        tiene_suscripcion = True
                        # Verificar si puede usar más servicios este mes
                        if not suscripcion_activa.puede_usar_servicio():
                            messages.error(request, f"El usuario {reserva.usuario.nombre_usuario} ha agotado sus servicios del mes. Servicios restantes: {suscripcion_activa.servicios_restantes()}")
                            return redirect('citascrud')
                except Exception as e:
                    print(f"Error verificando suscripción: {e}")
                
                reserva.suscripcion_utilizada = suscripcion_activa if tiene_suscripcion else None
                reserva.es_pago_individual = not tiene_suscripcion
                reserva.save()
                
                # Si tiene suscripción activa, incrementar el contador de servicios utilizados
                if tiene_suscripcion and suscripcion_activa:
                    # Contar los servicios asociados a esta reserva
                    servicios_count = reserva.servicios.count()
                    if servicios_count == 0:
                        servicios_count = 1  # Al menos un servicio por defecto
                    
                    suscripcion_activa.servicios_utilizados_mes += servicios_count
                    suscripcion_activa.save()
                    print(f"✅ Servicios utilizados actualizados (Admin): {suscripcion_activa.servicios_utilizados_mes}/{suscripcion_activa.plan.cantidad_servicios_mes if suscripcion_activa.plan.cantidad_servicios_mes > 0 else 'Ilimitado'}")
                
                messages.success(request, 'Reserva creada con éxito.')
                return redirect('citascrud')
    else:
        form = ReservaForm()

    # Generar lista de fechas disponibles
    fechas_disponibles = [hoy + timedelta(days=i) for i in range(15)] 

    # Para las estadísticas, necesitamos todas las reservas sin filtrar (ordenadas por más recientes primero)
    todas_las_reservas = Reserva.objects.select_related('empresa', 'usuario').prefetch_related('reservaservicio_set__servicio').order_by('-fecha', '-hora')

    # Implementar paginación
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.http import JsonResponse
    
    paginator = Paginator(reservas, 50)  # 50 reservas por página
    page_number = request.GET.get('page', 1)
    
    # Validar el número de página
    try:
        if page_number is None or page_number == '':
            page_number = 1
        else:
            page_number = int(page_number)
            if page_number < 1:
                page_number = 1
    except (ValueError, TypeError):
        page_number = 1
    
    page_obj = paginator.get_page(page_number)
    
    # Si es una petición AJAX para cargar más reservas
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        reservas_data = []
        for reserva in page_obj:
            # Formatear los datos de la reserva para JSON
            servicios_lista = []
            for rs in reserva.reservaservicio_set.all():
                servicios_lista.append({
                    'nombre': rs.servicio.nombre_servicio,
                    'precio': float(rs.servicio.precio),
                    'es_servicio_plan': rs.es_servicio_plan,
                    'precio_original': float(rs.precio_original) if rs.precio_original else float(rs.servicio.precio),
                    'precio_aplicado': float(rs.precio_aplicado) if rs.precio_aplicado else float(rs.servicio.precio),
                    'descuento_plan_individual': float(rs.descuento_plan_individual) if rs.descuento_plan_individual else 0,
                    'descuento_empresarial': float(rs.descuento_empresarial) if rs.descuento_empresarial else 0,
                })
            
            # Calcular el total con descuentos
            total_con_descuentos = sum(float(rs.precio_aplicado) if rs.precio_aplicado else float(rs.servicio.precio) for rs in reserva.reservaservicio_set.all())
            
            reservas_data.append({
                'id_reserva': reserva.id_reserva,
                'usuario_nombre': reserva.usuario.nombre_completo,
                'usuario_username': reserva.usuario.nombre_usuario,
                'usuario_telefono': reserva.usuario.telefono or '',
                'empresa_nombre': reserva.empresa.nombre_empresa if reserva.empresa else 'Sin empresa',
                'empresa_id': reserva.empresa.id_empresa if reserva.empresa else None,
                'fecha': reserva.fecha.strftime('%d/%m/%Y'),
                'hora': reserva.hora.strftime('%H:%M'),
                'estado': reserva.estado,
                'es_pago_individual': reserva.es_pago_individual,
                'es_reserva_empresarial': reserva.es_reserva_empresarial,
                'tipo_vehiculo': reserva.tipo_vehiculo or '',
                'placa_vehiculo': reserva.placa_vehiculo or '',
                'conductor_asignado': reserva.conductor_asignado or '',
                'servicios': servicios_lista,
                'total_precio': total_con_descuentos,
            })
        
        return JsonResponse({
            'reservas': reservas_data,
            'has_next': page_obj.has_next(),
            'page_number': page_obj.number,
            'total_pages': paginator.num_pages,
        })

    # Renderiza la vista con los datos necesarios
    return render(request, 'reservas/citas_crud.html', {
        'page_obj': page_obj,
        'reservas': page_obj,  # Mantener compatibilidad
        'todas_las_reservas': todas_las_reservas,  # Para las estadísticas
        'form': form,
        'servicios': servicios,
        'empresas': empresas,
        'horas_disponibles': horas_disponibles,
        'fechas_disponibles': fechas_disponibles,
        'hoy': hoy,
        'filtros_aplicados': filtros_aplicados,
    })


@admin_required
def crear_cita_admin(request):
    """Vista para que el administrador cree citas desde el CRUD"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    ahora = timezone.now()
    hoy = ahora.date()
    servicios = Servicio.objects.all()
    empresas = Empresa.objects.filter(verificada=True)
    usuarios = User.objects.filter(is_superuser=False)  # Obtener usuarios no admin
    
    # Obtener reservas existentes para calcular horas ocupadas
    reservas_existentes = Reserva.objects.all()
    ocupadas = {
        "fechas": {reserva.fecha for reserva in reservas_existentes},
        "horas": {}
    }
    
    for reserva in reservas_existentes:
        if reserva.fecha not in ocupadas["horas"]:
            ocupadas["horas"][reserva.fecha] = set()
        ocupadas["horas"][reserva.fecha].add(reserva.hora.strftime('%H:%M'))
    
    if request.method == "POST":
        usuario_id = request.POST.get('usuario')
        empresa_id = request.POST.get('empresa')
        fecha_seleccionada = request.POST.get('fecha')
        hora_12h = request.POST.get('hora')
        servicios_ids = request.POST.getlist('servicios')
        
        # Validaciones
        if not usuario_id:
            messages.error(request, "Debes seleccionar un usuario.")
            return redirect('crear_cita_admin')
            
        if not empresa_id:
            messages.error(request, "Debes seleccionar una empresa.")
            return redirect('crear_cita_admin')
            
        if not fecha_seleccionada:
            messages.error(request, "Debes seleccionar una fecha.")
            return redirect('crear_cita_admin')
            
        if not hora_12h:
            messages.error(request, "Debes seleccionar una hora.")
            return redirect('crear_cita_admin')
            
        if not servicios_ids:
            messages.error(request, "Debes seleccionar al menos un servicio.")
            return redirect('crear_cita_admin')
        
        try:
            usuario = User.objects.get(id=usuario_id)
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            servicios_seleccionados = Servicio.objects.filter(id_servicio__in=servicios_ids)
            
            if len(servicios_seleccionados) != len(servicios_ids):
                messages.error(request, "Uno o más servicios seleccionados no existen.")
                return redirect('crear_cita_admin')
            
            # Convertir hora a formato 24h
            hora = convertir_hora_24h(hora_12h)
            
            # Verificar si la fecha y hora están ocupadas
            fecha_obj = datetime.strptime(fecha_seleccionada, '%Y-%m-%d').date()
            if fecha_obj in ocupadas["fechas"] and hora in ocupadas["horas"].get(fecha_obj, set()):
                messages.error(request, f"La fecha {fecha_seleccionada} a las {hora_12h} ya está ocupada.")
                return redirect('crear_cita_admin')
            
            # Verificar suscripción del usuario
            suscripcion_activa = None
            usar_suscripcion = False
            
            try:
                from .models import SuscripcionUsuario
                suscripcion_activa = SuscripcionUsuario.objects.filter(
                    usuario=usuario,
                    estado='activa'
                ).first()
                
                if suscripcion_activa and suscripcion_activa.esta_activa():
                    if suscripcion_activa.puede_usar_servicio():
                        usar_suscripcion = True
            except Exception as e:
                print(f"Error verificando suscripción: {e}")
            
            # Crear la reserva
            reserva = Reserva(
                empresa=empresa,
                fecha=fecha_seleccionada,
                hora=hora,
                usuario=usuario,
                suscripcion_utilizada=suscripcion_activa if usar_suscripcion else None,
                es_pago_individual=not usar_suscripcion
            )
            reserva.save()
            
            # Crear las relaciones entre reserva y servicios
            precio_total = 0
            servicios_nombres = []
            for servicio in servicios_seleccionados:
                ReservaServicio.objects.create(reserva=reserva, servicio=servicio)
                precio_total += servicio.precio
                servicios_nombres.append(servicio.nombre_servicio)
            
            # Si usa suscripción, incrementar el contador
            if usar_suscripcion and suscripcion_activa:
                suscripcion_activa.servicios_utilizados_mes += len(servicios_seleccionados)
                suscripcion_activa.save()
            
            messages.success(request, f"Cita creada exitosamente para {usuario.nombre_usuario} el {fecha_seleccionada} a las {hora_12h}.")
            return redirect('citascrud')
            
        except User.DoesNotExist:
            messages.error(request, "El usuario seleccionado no existe.")
        except Empresa.DoesNotExist:
            messages.error(request, "La empresa seleccionada no existe.")
        except Exception as e:
            messages.error(request, f"Error al crear la cita: {str(e)}")
        
        return redirect('crear_cita_admin')
    
    # Generar fechas disponibles (próximos 15 días)
    fechas_disponibles = []
    for i in range(15):
        fecha = hoy + timedelta(days=i)
        fechas_disponibles.append(fecha)
    
    # Generar horas disponibles (8:00 AM a 3:00 PM)
    horas_disponibles = []
    for h in range(8, 16):
        hora_24h = f"{h:02d}:00"
        hora_12h = convertir_hora_12h(hora_24h)
        horas_disponibles.append({
            'valor': hora_12h,
            'texto': hora_12h
        })
    
    context = {
        'servicios': servicios,
        'empresas': empresas,
        'usuarios': usuarios,
        'fechas_disponibles': fechas_disponibles,
        'horas_disponibles': horas_disponibles,
        'ocupadas': ocupadas,
    }
    
    return render(request, 'reservas/crear_cita_admin.html', context)






@admin_required
def cambiar_estado_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, id_reserva=reserva_id)
    
    if reserva.estado == 'pendiente':
        reserva.estado = 'completado'
        reserva.save()
        messages.success(request, 'La reserva ha sido marcada como completada.')
    else:
        messages.info(request, 'La reserva ya está completada.')

    return redirect('citascrud') 


@admin_required
def servicios_crud(request):
    from django.db.models import Q
    
    # Filtros de búsqueda
    busqueda = request.GET.get('busqueda', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()
    
    # Base queryset con optimización y orden
    servicios_qs = Servicio.objects.all().prefetch_related('empresaservicio_set__empresa').order_by('-id_servicio')
    
    # Aplicar filtros
    if busqueda:
        servicios_qs = servicios_qs.filter(
            Q(nombre_servicio__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(precio__icontains=busqueda)
        )
    
    # Filtrar según los parámetros de búsqueda (para compatibilidad con código existente)
    nombre_servicio = request.GET.get('nombre_servicio', '')
    if nombre_servicio and not busqueda:  # Solo si no hay búsqueda general
        servicios_qs = servicios_qs.filter(nombre_servicio__icontains=nombre_servicio)
    
    # Calcular estadísticas de manera más eficiente
    total_servicios = servicios_qs.count()
    total_asignaciones = EmpresaServicio.objects.count()
    total_solicitudes_pendientes = SolicitudServicioEmpresa.objects.filter(estado='pendiente').count()
    
    # Paginación de servicios - 8 por página
    paginator = Paginator(servicios_qs, 8)
    page = request.GET.get('page', 1)
    
    try:
        servicios = paginator.page(page)
    except PageNotAnInteger:
        servicios = paginator.page(1)
    except EmptyPage:
        servicios = paginator.page(paginator.num_pages)
    
    # Obtener solicitudes pendientes con limit para no sobrecargar
    solicitudes_servicios = SolicitudServicioEmpresa.objects.filter(
        estado='pendiente'
    ).select_related('empresa', 'servicio_solicitado')[:10]  # Solo las primeras 10
    
    # Filtros activos para el template
    filtros_activos = {
        'busqueda': busqueda,
        'estado': estado_filtro,
    }
    
    form = ServicioForm()

    if request.method == "POST":
        # Manejar aprobación/rechazo de solicitudes
        if 'aprobar_solicitud' in request.POST:
            solicitud_id = request.POST.get('aprobar_solicitud')
            try:
                solicitud = get_object_or_404(SolicitudServicioEmpresa, id_solicitud=solicitud_id)
                resultado = solicitud.aprobar_solicitud()
                if resultado:
                    messages.success(request, f'Solicitud de {solicitud.empresa.nombre_empresa} para el servicio {solicitud.servicio_solicitado.nombre_servicio} aprobada exitosamente.')
                else:
                    messages.info(request, f'Solicitud de {solicitud.empresa.nombre_empresa} procesada. La empresa ya tenía acceso al servicio {solicitud.servicio_solicitado.nombre_servicio}.')
            except Exception as e:
                messages.error(request, f'Error al aprobar la solicitud: {str(e)}')
            return redirect('servicioscrud')

        if 'rechazar_solicitud' in request.POST:
            solicitud_id = request.POST.get('rechazar_solicitud')
            motivo = request.POST.get('motivo_rechazo', 'Solicitud rechazada por el administrador')
            try:
                solicitud = get_object_or_404(SolicitudServicioEmpresa, id_solicitud=solicitud_id)
                solicitud.rechazar_solicitud(motivo)
                messages.warning(request, f'Solicitud de {solicitud.empresa.nombre_empresa} para el servicio {solicitud.servicio_solicitado.nombre_servicio} rechazada.')
            except Exception as e:
                messages.error(request, f'Error al rechazar la solicitud: {str(e)}')
            return redirect('servicioscrud')

        # Manejar la eliminación de los servicios
        if 'eliminar' in request.POST:
            servicio_id = request.POST.get('eliminar')
            servicio_a_eliminar = get_object_or_404(Servicio, id_servicio=servicio_id)
            servicio_a_eliminar.delete()
            messages.error(request, 'El Servicio ha sido eliminado.')
            return redirect('servicioscrud')

        # Manejar la creación de nuevos servicios
        if 'nombre_servicio' in request.POST and 'id_servicio' not in request.POST:
            nombre_servicio = request.POST.get('nombre_servicio')
            descripcion = request.POST.get('descripcion')
            precio = request.POST.get('precio')

            nuevo_servicio = Servicio(
                nombre_servicio=nombre_servicio,
                descripcion=descripcion,
                precio=precio
            )
            nuevo_servicio.save()
            messages.success(request, 'El servicio ha sido creado.')
            return redirect('servicioscrud')

        # Manejar la actualización de servicios
        if 'id_servicio' in request.POST:
            servicio_id = request.POST.get('id_servicio')
            servicio = get_object_or_404(Servicio, id_servicio=servicio_id)

            # Obtener los nuevos valores desde el POST
            nuevo_nombre_servicio = request.POST.get('nombre_servicio')
            nueva_descripcion = request.POST.get('descripcion')
            nuevo_precio = request.POST.get('precio')

            # Solo actualizar si se proporciona un nuevo valor
            if nuevo_nombre_servicio:
                servicio.nombre_servicio = nuevo_nombre_servicio

            if nueva_descripcion:
                servicio.descripcion = nueva_descripcion

            if nuevo_precio:
                servicio.precio = nuevo_precio

            servicio.save()
            messages.success(request, 'El servicio ha sido actualizado.')
            return redirect('servicioscrud')

        # Manejar la asignación/desasignación de empresas
        if 'asignar_empresa' in request.POST:
            servicio_id = request.POST.get('servicio_id')
            servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
            empresa_ids = request.POST.getlist('empresas')  # Recoge las empresas seleccionadas

            # Obtener las empresas que ya están asociadas al servicio
            empresas_asociadas = EmpresaServicio.objects.filter(servicio=servicio)

            # Eliminar las relaciones de empresa-servicio que no están seleccionadas
            for empresa_servicio in empresas_asociadas:
                if str(empresa_servicio.empresa.id_empresa) not in empresa_ids:
                    empresa_servicio.delete()

            # Crear las relaciones de empresa-servicio para las empresas nuevas seleccionadas
            for empresa_id in empresa_ids:
                empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
                # Crear la relación solo si no existe ya
                if not EmpresaServicio.objects.filter(servicio=servicio, empresa=empresa).exists():
                    EmpresaServicio.objects.create(empresa=empresa, servicio=servicio)

            messages.success(request, 'Las empresas han sido asignadas o desasignadas correctamente al servicio.')
            return redirect('servicioscrud')

    else:
        form = ServicioForm()

    # Obtener la lista de empresas disponibles (solo las activas para optimizar)
    empresas = Empresa.objects.filter(verificada=True)

    # Obtener las empresas asociadas a cada servicio (solo para los servicios de la página actual)
    for servicio in servicios:
        servicio.empresas_asociadas = EmpresaServicio.objects.filter(servicio=servicio).values_list('empresa', flat=True)

    context = {
        'servicios': servicios,
        'paginator': paginator,
        'total_servicios': total_servicios,
        'filtros_activos': filtros_activos,
        'form': form, 
        'empresas': empresas,
        'total_asignaciones': total_asignaciones,
        'solicitudes_servicios': solicitudes_servicios,
        'total_solicitudes_pendientes': total_solicitudes_pendientes
    }

    return render(request, 'servicios/servicios_crud.html', context)


@admin_required
def crear_servicio(request):
    """Vista para crear un nuevo servicio"""
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio creado exitosamente.')
            return redirect('servicioscrud')
        else:
            messages.error(request, 'Error al crear el servicio. Verifica los datos ingresados.')
    else:
        form = ServicioForm()

    return render(request, 'servicios/crear_servicio.html', {'form': form})


@admin_required
def editar_servicio(request, servicio_id):
    """Vista para editar un servicio existente"""
    servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
    
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, f'Servicio "{servicio.nombre_servicio}" actualizado exitosamente.')
            return redirect('servicioscrud')
        else:
            messages.error(request, 'Error al actualizar el servicio. Verifica los datos ingresados.')
    else:
        form = ServicioForm(instance=servicio)
    
    return render(request, 'servicios/editar_servicio.html', {'form': form, 'servicio': servicio})


@admin_required
def eliminar_servicio(request, servicio_id):
    """Vista para eliminar un servicio"""
    servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
    
    if request.method == 'POST':
        nombre_servicio = servicio.nombre_servicio
        servicio.delete()
        messages.success(request, f'Servicio "{nombre_servicio}" eliminado exitosamente.')
        return redirect('servicioscrud')
    
    return render(request, 'servicios/eliminar_servicio.html', {'servicio': servicio})


@admin_required
def gestionar_asignaciones_servicios(request):
    """Vista para gestionar las asignaciones de servicios a empresas"""
    from django.core.paginator import Paginator
    from django.http import JsonResponse
    
    empresas = Empresa.objects.all().order_by('nombre_empresa')
    servicios = Servicio.objects.all().order_by('nombre_servicio')
    asignaciones = EmpresaServicio.objects.select_related('empresa', 'servicio').all()
    
    # Crear diccionario para manejo eficiente de asignaciones
    asignaciones_dict = {}
    for asignacion in asignaciones:
        if asignacion.empresa.id_empresa not in asignaciones_dict:
            asignaciones_dict[asignacion.empresa.id_empresa] = []
        asignaciones_dict[asignacion.empresa.id_empresa].append(asignacion.servicio.id_servicio)
    
    # Filtros
    empresa_filtro = request.GET.get('empresa_filtro', '')
    servicio_filtro = request.GET.get('servicio_filtro', '')
    
    if empresa_filtro:
        empresas = empresas.filter(nombre_empresa__icontains=empresa_filtro)
    
    if servicio_filtro:
        servicios = servicios.filter(nombre_servicio__icontains=servicio_filtro)
    
    # Procesar formulario de asignación masiva
    if request.method == 'POST':
        # Asignación masiva: Empresa → Servicios
        if 'asignacion_masiva_empresa' in request.POST:
            empresa_id = request.POST.get('empresa_id')
            servicios_ids = request.POST.getlist('servicios')
            
            if empresa_id and servicios_ids:
                empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
                
                # Eliminar asignaciones actuales
                EmpresaServicio.objects.filter(empresa=empresa).delete()
                
                # Crear nuevas asignaciones
                count = 0
                for servicio_id in servicios_ids:
                    servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
                    EmpresaServicio.objects.create(empresa=empresa, servicio=servicio)
                    count += 1
                
                messages.success(request, f'✓ {count} servicio(s) asignado(s) exitosamente a {empresa.nombre_empresa}.')
                return redirect('gestionar_asignaciones_servicios')
            else:
                messages.error(request, 'Debes seleccionar una empresa y al menos un servicio.')
                return redirect('gestionar_asignaciones_servicios')
        
        # Asignación masiva inversa: Servicio → Empresas
        elif 'asignacion_masiva_servicio' in request.POST:
            servicio_id = request.POST.get('servicio_id')
            empresas_ids = request.POST.getlist('empresas')
            
            if servicio_id and empresas_ids:
                servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
                
                # Eliminar asignaciones actuales de este servicio
                EmpresaServicio.objects.filter(servicio=servicio).delete()
                
                # Crear nuevas asignaciones
                count = 0
                for empresa_id in empresas_ids:
                    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
                    EmpresaServicio.objects.create(empresa=empresa, servicio=servicio)
                    count += 1
                
                messages.success(request, f'✓ El servicio "{servicio.nombre_servicio}" ha sido asignado a {count} empresa(s).')
                return redirect('gestionar_asignaciones_servicios')
            else:
                messages.error(request, 'Debes seleccionar un servicio y al menos una empresa.')
                return redirect('gestionar_asignaciones_servicios')
        
        # Compatibilidad con versión anterior
        elif 'asignacion_masiva' in request.POST:
            empresa_id = request.POST.get('empresa_id')
            servicios_ids = request.POST.getlist('servicios')
            
            if empresa_id and servicios_ids:
                empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
                
                # Eliminar asignaciones actuales
                EmpresaServicio.objects.filter(empresa=empresa).delete()
                
                # Crear nuevas asignaciones
                for servicio_id in servicios_ids:
                    servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
                    EmpresaServicio.objects.create(empresa=empresa, servicio=servicio)
                
                messages.success(request, f'Servicios asignados exitosamente a {empresa.nombre_empresa}.')
                return redirect('gestionar_asignaciones_servicios')
        
        elif 'eliminar_asignacion' in request.POST:
            asignacion_ids = request.POST.getlist('asignaciones_eliminar')
            if asignacion_ids:
                count = EmpresaServicio.objects.filter(id__in=asignacion_ids).delete()[0]
                messages.success(request, f'{count} asignaciones eliminadas exitosamente.')
                return redirect('gestionar_asignaciones_servicios')
    
    # Estadísticas
    total_empresas = empresas.count()
    total_servicios = servicios.count()
    total_asignaciones = EmpresaServicio.objects.count()
    empresas_sin_servicios = empresas.exclude(id_empresa__in=asignaciones_dict.keys()).count()
    
    # Paginación de asignaciones
    page = request.GET.get('page', 1)
    paginator = Paginator(asignaciones, 10)  # 10 asignaciones por página
    asignaciones_paginadas = paginator.get_page(page)
    
    # Si es una petición AJAX, devolver solo los datos de la tabla
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        asignaciones_html = []
        for asignacion in asignaciones_paginadas:
            asignaciones_html.append({
                'empresa_nombre': asignacion.empresa.nombre_empresa,
                'empresa_direccion': asignacion.empresa.direccion,
                'empresa_id': asignacion.empresa.id_empresa,
                'servicio_nombre': asignacion.servicio.nombre_servicio,
                'servicio_id': asignacion.servicio.id_servicio,
                'servicio_precio': float(asignacion.servicio.precio),
                'empresa_verificada': asignacion.empresa.verificada,
            })
        
        return JsonResponse({
            'asignaciones': asignaciones_html,
            'has_previous': asignaciones_paginadas.has_previous(),
            'has_next': asignaciones_paginadas.has_next(),
            'page_number': asignaciones_paginadas.number,
            'num_pages': paginator.num_pages,
            'total_asignaciones': paginator.count,
        })
    
    context = {
        'empresas': empresas,
        'servicios': servicios,
        'asignaciones': asignaciones_paginadas,
        'asignaciones_dict': asignaciones_dict,
        'total_empresas': total_empresas,
        'total_servicios': total_servicios,
        'total_asignaciones': total_asignaciones,
        'empresas_sin_servicios': empresas_sin_servicios,
        'empresa_filtro': empresa_filtro,
        'servicio_filtro': servicio_filtro,
    }
    
    return render(request, 'servicios/gestionar_asignaciones_servicios.html', context)



@admin_required
def asignar_servicio_empresa(request, empresa_id):
    """Vista para asignar servicios específicos a una empresa"""
    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
    servicios = Servicio.objects.all()
    servicios_asignados = EmpresaServicio.objects.filter(empresa=empresa).values_list('servicio_id', flat=True)
    
    if request.method == 'POST':
        servicios_ids = request.POST.getlist('servicios')
        
        # Eliminar asignaciones actuales
        EmpresaServicio.objects.filter(empresa=empresa).delete()
        
        # Crear nuevas asignaciones
        for servicio_id in servicios_ids:
            servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
            EmpresaServicio.objects.create(empresa=empresa, servicio=servicio)
        
        messages.success(request, f'Servicios actualizados para {empresa.nombre_empresa}.')
        return redirect('gestionar_asignaciones_servicios')
    
    context = {
        'empresa': empresa,
        'servicios': servicios,
        'servicios_asignados': list(servicios_asignados),
    }
    
    return render(request, 'empresas/asignar_servicio_empresa.html', context)


@admin_required
def detalle_servicio(request, servicio_id):
    """Vista para ver el detalle de un servicio y sus empresas asignadas"""
    servicio = get_object_or_404(Servicio, id_servicio=servicio_id)
    empresas_asignadas = EmpresaServicio.objects.filter(servicio=servicio).select_related('empresa')
    reservas_servicio = ReservaServicio.objects.filter(servicio=servicio).select_related('reserva').count()
    
    # Estadísticas del servicio
    total_empresas_asignadas = empresas_asignadas.count()
    total_reservas = reservas_servicio
    
    context = {
        'servicio': servicio,
        'empresas_asignadas': empresas_asignadas,
        'total_empresas_asignadas': total_empresas_asignadas,
        'total_reservas': total_reservas,
    }
    
    return render(request, 'servicios/detalle_servicio.html', context)


@admin_required
def empresas_crud(request):
    # Obtener todas las empresas para estadísticas
    todas_empresas = Empresa.objects.all()
    form = EmpresaForm()  # Si estás usando un formulario de Django para crear y actualizar empresas

    # Filtrar según los parámetros de búsqueda
    nombre_empresa = request.GET.get('nombre_empresa', '')
    verificacion = request.GET.get('verificacion', '')
    busqueda = request.GET.get('busqueda', '')
    estado = request.GET.get('estado', 'activas')  # Nuevo filtro para estado (activas/desactivadas)

    # Aplicar filtros
    empresas_filtradas = todas_empresas

    # Filtro de estado (activas o desactivadas)
    if estado == 'activas':
        empresas_filtradas = empresas_filtradas.filter(is_active=True)
    elif estado == 'desactivadas':
        empresas_filtradas = empresas_filtradas.filter(is_active=False)

    if nombre_empresa:
        empresas_filtradas = empresas_filtradas.filter(nombre_empresa__icontains=nombre_empresa)
    
    if verificacion == 'verificada':
        empresas_filtradas = empresas_filtradas.filter(verificada=True)
    elif verificacion == 'sin_verificar':
        empresas_filtradas = empresas_filtradas.filter(verificada=False)

    # Filtro de búsqueda general
    if busqueda:
        empresas_filtradas = empresas_filtradas.filter(
            Q(nombre_empresa__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(direccion__icontains=busqueda)
        )

    if request.method == "POST":
        # Manejar la eliminación de empresas
        if 'eliminar' in request.POST:
            empresa_id = request.POST.get('eliminar')
            empresa_a_eliminar = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa_a_eliminar.delete()
            messages.error(request, 'La empresa ha sido eliminada.')
            return redirect('empresascrud')

        # Manejar la activación de empresas
        if 'activar' in request.POST:
            empresa_id = request.POST.get('activar')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa.is_active = True
            empresa.save()
            messages.success(request, f'La empresa "{empresa.nombre_empresa}" ha sido activada exitosamente.')
            return redirect('empresascrud')

        # Manejar la desactivación de empresas
        if 'desactivar' in request.POST:
            empresa_id = request.POST.get('desactivar')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa.is_active = False
            empresa.save()
            messages.warning(request, f'La empresa "{empresa.nombre_empresa}" ha sido desactivada.')
            return redirect('empresascrud')

        # Manejar el cambio de estado de verificación
        if 'cambiar_verificacion' in request.POST:
            empresa_id = request.POST.get('cambiar_verificacion')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa.verificada = not empresa.verificada  # Cambiar el estado
            empresa.save()
            estado_mensaje = 'verificada' if empresa.verificada else 'sin verificar'
            messages.success(request, f'La empresa ha sido marcada como {estado_mensaje}.')
            return redirect('empresascrud')

        # Manejar la verificación de empresas
        if 'verificar' in request.POST:
            empresa_id = request.POST.get('verificar')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa.verificada = True
            empresa.save()
            messages.success(request, f'La empresa "{empresa.nombre_empresa}" ha sido verificada exitosamente.')
            return redirect('empresascrud')

        # Manejar la desverificación de empresas
        if 'desverificar' in request.POST:
            empresa_id = request.POST.get('desverificar')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            empresa.verificada = False
            empresa.save()
            messages.warning(request, f'La empresa "{empresa.nombre_empresa}" ha sido desverificada.')
            return redirect('empresascrud')

        # Manejar la creación de una nueva empresa
        if 'nombre_empresa' in request.POST and 'id_empresa' not in request.POST:
            nombre_empresa = request.POST.get('nombre_empresa')
            direccion = request.POST.get('direccion')
            telefono = request.POST.get('telefono')
            email = request.POST.get('email')
            contrasena = request.POST.get('contrasena', 'temp_password')

            # Encriptar la contraseña si se proporciona
            if contrasena and contrasena != 'temp_password':
                contrasena_encriptada = make_password(contrasena)
            else:
                contrasena_encriptada = 'temp_password'

            nueva_empresa = Empresa(
                nombre_empresa=nombre_empresa,
                direccion=direccion,
                telefono=telefono,
                email=email,
                contrasena=contrasena_encriptada
            )
            nueva_empresa.save()
            messages.success(request, 'La empresa ha sido creada exitosamente.')
            return redirect('empresascrud')

        # Manejar la actualización de la empresa
        if 'id_empresa' in request.POST:
            empresa_id = request.POST.get('id_empresa')
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)

            # Obtener los nuevos valores desde el POST
            nuevo_nombre_empresa = request.POST.get('nombre_empresa')
            nueva_direccion = request.POST.get('direccion')
            nuevo_telefono = request.POST.get('telefono')
            nuevo_email = request.POST.get('email')
            nueva_contrasena = request.POST.get('contrasena')

            # Solo actualizar si se proporciona un nuevo valor
            if nuevo_nombre_empresa:
                empresa.nombre_empresa = nuevo_nombre_empresa
            if nueva_direccion:
                empresa.direccion = nueva_direccion
            if nuevo_telefono:
                empresa.telefono = nuevo_telefono
            if nuevo_email:
                empresa.email = nuevo_email
            if nueva_contrasena:
                empresa.contrasena = make_password(nueva_contrasena)

            empresa.save()
            messages.success(request, 'La empresa ha sido actualizada exitosamente.')
            return redirect('empresascrud')

    # Verificar si hay empresas registradas en las últimas 24 horas
    from datetime import timedelta
    hace_24_horas = timezone.now() - timedelta(hours=24)
    empresas_nuevas_24h = todas_empresas.filter(fecha_registro__gte=hace_24_horas).count()
    
    # Determinar el ordenamiento según el filtro aplicado
    if verificacion == 'sin_verificar':
        # Para empresas sin verificar, mostrar las más viejas primero
        empresas_ordenadas = empresas_filtradas.order_by('fecha_registro')
    else:
        # Para el resto, ordenar por nombre de empresa
        empresas_ordenadas = empresas_filtradas.order_by('nombre_empresa')
    
    # Implementar paginación
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(empresas_ordenadas, 25)  # 25 empresas por página
    page = request.GET.get('page')
    
    try:
        empresas = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, mostrar la primera página
        empresas = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera del rango, mostrar la última página
        empresas = paginator.page(paginator.num_pages)

    # Calcular estadísticas basadas en todas las empresas
    empresas_verificadas_total = todas_empresas.filter(verificada=True).count()
    empresas_sin_verificar_total = todas_empresas.filter(verificada=False).count()
    empresas_activas_total = todas_empresas.filter(is_active=True).count()
    empresas_desactivadas_total = todas_empresas.filter(is_active=False).count()
    
    # Calcular empresas con servicios asignados
    empresas_con_servicios_total = todas_empresas.filter(
        empresaservicio__isnull=False
    ).distinct().count()
    
    return render(request, 'empresas/empresas_crud.html', {
        'empresas': empresas, 
        'paginator': paginator,
        'form': form,
        'empresas_verificadas': empresas_verificadas_total,
        'empresas_sin_verificar': empresas_sin_verificar_total,
        'empresas_pendientes': empresas_sin_verificar_total,  # Las pendientes son las sin verificar
        'empresas_con_servicios': empresas_con_servicios_total,
        'empresas_activas': empresas_activas_total,
        'empresas_desactivadas': empresas_desactivadas_total,
        'total_empresas': todas_empresas.count(),
        'empresas_nuevas_24h': empresas_nuevas_24h,  # Nuevas empresas en últimas 24 horas
        'filtros_activos': {
            'nombre_empresa': nombre_empresa,
            'verificacion': verificacion,
            'busqueda': busqueda,
            'estado': estado,
        }
    })


@admin_required
def editar_empresa(request, empresa_id):
    """Vista para editar una empresa específica"""
    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
    
    if request.method == "POST":
        # Obtener los datos básicos del formulario
        nombre_empresa = request.POST.get('nombre_empresa')
        direccion = request.POST.get('direccion')
        telefono = request.POST.get('telefono')
        email = request.POST.get('email')
        # La plantilla usa 'nueva_contrasena' y 'confirmar_contrasena'
        nueva_contrasena = request.POST.get('nueva_contrasena')
        confirmar_contrasena = request.POST.get('confirmar_contrasena')
        
        # Obtener datos fiscales
        nit_empresa = request.POST.get('nit_empresa')
        razon_social = request.POST.get('razon_social')
        regimen_tributario = request.POST.get('regimen_tributario')
        
        # Obtener datos del titular de la cuenta
        titular_cuenta = request.POST.get('titular_cuenta')
        tipo_documento_titular = request.POST.get('tipo_documento_titular')
        numero_documento_titular = request.POST.get('numero_documento_titular')
        
        # Obtener datos bancarios
        banco = request.POST.get('banco')
        tipo_cuenta = request.POST.get('tipo_cuenta')
        numero_cuenta = request.POST.get('numero_cuenta')
        swift_code = request.POST.get('swift_code')
        iban = request.POST.get('iban')
        
        # Obtener datos de contacto para facturación
        email_facturacion = request.POST.get('email_facturacion')
        telefono_facturacion = request.POST.get('telefono_facturacion')
        responsable_pagos = request.POST.get('responsable_pagos')
        
        # Obtener notas bancarias
        notas_bancarias = request.POST.get('notas_bancarias')
        
        # Obtener verificación de datos bancarios (solo admins)
        datos_bancarios_verificados = request.POST.get('datos_bancarios_verificados') == '1'
        
        # Obtener estados
        verificada = request.POST.get('verificada') == '1'
        activa = request.POST.get('activa') == '1'

        # Actualizar los campos básicos de la empresa
        if nombre_empresa:
            empresa.nombre_empresa = nombre_empresa
        if direccion:
            empresa.direccion = direccion
        if telefono:
            empresa.telefono = telefono
        if email:
            empresa.email = email
        
        # Actualizar datos fiscales
        empresa.nit_empresa = nit_empresa if nit_empresa else None
        empresa.razon_social = razon_social if razon_social else None
        empresa.regimen_tributario = regimen_tributario if regimen_tributario else None
        
        # Actualizar datos del titular
        empresa.titular_cuenta = titular_cuenta if titular_cuenta else None
        empresa.tipo_documento_titular = tipo_documento_titular if tipo_documento_titular else None
        empresa.numero_documento_titular = numero_documento_titular if numero_documento_titular else None
        
        # Actualizar datos bancarios
        empresa.banco = banco if banco else None
        empresa.tipo_cuenta = tipo_cuenta if tipo_cuenta else None
        empresa.numero_cuenta = numero_cuenta if numero_cuenta else None
        empresa.swift_code = swift_code if swift_code else None
        empresa.iban = iban if iban else None
        
        # Actualizar contacto de facturación
        empresa.email_facturacion = email_facturacion if email_facturacion else None
        empresa.telefono_facturacion = telefono_facturacion if telefono_facturacion else None
        empresa.responsable_pagos = responsable_pagos if responsable_pagos else None
        
        # Actualizar notas bancarias
        empresa.notas_bancarias = notas_bancarias if notas_bancarias else None
        
        # Actualizar verificación de datos bancarios (solo si es admin/staff)
        if request.user.is_staff or request.user.rol == 'admin':
            # Si se está marcando como verificado y antes no lo estaba
            if datos_bancarios_verificados and not empresa.datos_bancarios_verificados:
                empresa.datos_bancarios_verificados = True
                empresa.fecha_verificacion_bancaria = timezone.now()
                empresa.verificado_por = request.user
                messages.success(request, 'Los datos bancarios han sido marcados como verificados.')
            # Si se está desmarcando la verificación
            elif not datos_bancarios_verificados and empresa.datos_bancarios_verificados:
                empresa.datos_bancarios_verificados = False
                empresa.fecha_verificacion_bancaria = None
                empresa.verificado_por = None
                messages.warning(request, 'Se ha removido la verificación de datos bancarios.')
        
        # Actualizar estados
        empresa.verificada = verificada
        empresa.is_active = activa

        # Si se quiere cambiar la contraseña, validar y guardar
        if nueva_contrasena or confirmar_contrasena:
            if not nueva_contrasena or not confirmar_contrasena:
                messages.error(request, 'Para cambiar la contraseña debe completar ambos campos.')
                return redirect('empresascrud')
            if nueva_contrasena != confirmar_contrasena:
                messages.error(request, 'Las nuevas contraseñas no coinciden.')
                return redirect('empresascrud')
            if len(nueva_contrasena) < 6:
                messages.error(request, 'La nueva contraseña debe tener al menos 6 caracteres.')
                return redirect('empresascrud')

            empresa.contrasena = make_password(nueva_contrasena)

        empresa.save()
        
        # Mensajes informativos sobre datos bancarios
        if empresa.datos_bancarios_completos():
            if empresa.datos_bancarios_verificados:
                messages.success(request, f'La empresa "{empresa.nombre_empresa}" está lista para recibir pagos.')
            else:
                messages.info(request, f'La empresa "{empresa.nombre_empresa}" tiene datos bancarios completos pero aún no verificados.')
        else:
            messages.warning(request, f'La empresa "{empresa.nombre_empresa}" no tiene datos bancarios completos para recibir pagos.')
        
        messages.success(request, f'La empresa "{empresa.nombre_empresa}" ha sido actualizada exitosamente.')
        return redirect('empresascrud')
    
    # Para GET, mostrar el formulario con los datos actuales
    form = EmpresaForm(instance=empresa)
    
    return render(request, 'empresas/editar_empresa.html', {
        'empresa': empresa,
        'form': form
    })




@admin_required
def home_crud(request):
    """Vista del dashboard principal con estadísticas completas del sistema"""
    from django.db.models import Count, Sum, Q
    from datetime import date, timedelta
    
    # Obtener fechas para filtros
    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=7)
    inicio_mes = hoy.replace(day=1)
    
    # === ESTADÍSTICAS DE USUARIOS ===
    total_usuarios = Usuario.objects.count()
    # Usuarios registrados este mes
    usuarios_mes = Usuario.objects.filter(fecha_registro__gte=inicio_mes).count()
    usuarios_activos = Usuario.objects.filter(is_active=True).count()
    usuarios_admin = Usuario.objects.filter(rol='admin').count()
    usuarios_cliente = Usuario.objects.filter(rol='cliente').count()
    
    # === ESTADÍSTICAS DE EMPRESAS ===
    total_empresas = Empresa.objects.count()
    empresas_verificadas = Empresa.objects.filter(verificada=True).count()
    empresas_pendientes = Empresa.objects.filter(verificada=False).count()
    
    # === ESTADÍSTICAS DE RESERVAS/CITAS ===
    total_reservas = Reserva.objects.count()
    reservas_hoy = Reserva.objects.filter(fecha=hoy).count()
    reservas_pendientes = Reserva.objects.filter(estado='pendiente').count()
    reservas_completadas = Reserva.objects.filter(estado='completado').count()
    reservas_canceladas = Reserva.objects.filter(estado='cancelada').count()
    reservas_semana = Reserva.objects.filter(fecha__gte=inicio_semana).count()
    reservas_mes = Reserva.objects.filter(fecha__gte=inicio_mes).count()
    
    # Reservas por estado
    reservas_por_estado = Reserva.objects.values('estado').annotate(
        total=Count('id_reserva')
    ).order_by('estado')
    
    # === ESTADÍSTICAS DE SERVICIOS ===
    total_servicios = Servicio.objects.count()
    servicios_mas_reservados = Servicio.objects.annotate(
        total_reservas=Count('reserva__id_reserva')
    ).order_by('-total_reservas')[:5]
    
    # === ESTADÍSTICAS DE PLANES ===
    total_planes = Plan.objects.filter(activo=True).count()
    total_planes_empresariales = PlanEmpresarial.objects.filter(activo=True).count()
    total_suscripciones_activas = SuscripcionUsuario.objects.filter(estado='activa').count()
    total_suscripciones_empresariales = SuscripcionEmpresarial.objects.filter(estado='activa').count()
    
    # === ESTADÍSTICAS DE COMENTARIOS Y QUEJAS ===
    total_comentarios = Comentario.objects.count()
    comentarios_mes = Comentario.objects.filter(fecha__gte=inicio_mes).count()
    total_quejas = MensajeQueja.objects.count()
    quejas_pendientes = MensajeQueja.objects.filter(estado='no respondido').count()
    quejas_respondidas = MensajeQueja.objects.exclude(estado='no respondido').count()
    
    # === ACTIVIDAD RECIENTE ===
    reservas_recientes = Reserva.objects.select_related('usuario', 'empresa').order_by('-id_reserva')[:5]
    comentarios_recientes = Comentario.objects.select_related('usuario').order_by('-fecha')[:5]
    empresas_recientes = Empresa.objects.order_by('-fecha_registro')[:5]
    
    # === MÉTRICAS DE RENDIMIENTO ===
    # Tasa de finalización de reservas
    if total_reservas > 0:
        tasa_completadas = round((reservas_completadas / total_reservas) * 100, 1)
        tasa_canceladas = round((reservas_canceladas / total_reservas) * 100, 1)
        porcentaje_pendientes = round((reservas_pendientes / total_reservas) * 100, 1)
    else:
        tasa_completadas = 0
        tasa_canceladas = 0
        porcentaje_pendientes = 0
    
    # Ingresos estimados (basado en precios de servicios)
    try:
        ingresos_potenciales = Servicio.objects.aggregate(
            total=Sum('precio')
        )['total'] or 0
        ingresos_estimados_mes = reservas_mes * (ingresos_potenciales / max(total_servicios, 1))
    except:
        ingresos_estimados_mes = 0
    
    # Ingresos reales con descuentos aplicados (TODAS las reservas completadas)
    try:
        # Obtener TODAS las reservas completadas (sin filtro de fecha)
        reservas_completadas_todas = Reserva.objects.filter(estado='completado')
        
        ingresos_reales_totales = 0
        ingresos_sin_descuento_totales = 0
        total_descuentos_otorgados = 0
        reservas_con_descuento = 0
        
        for reserva in reservas_completadas_todas:
            # Obtener todos los servicios de esta reserva
            reserva_servicios = reserva.reservaservicio_set.all()
            
            for rs in reserva_servicios:
                # Prioridad: usar los precios guardados en ReservaServicio
                if rs.precio_aplicado is not None and rs.precio_original is not None:
                    # Caso 1: Precios guardados correctamente
                    precio_real = float(rs.precio_aplicado)  # Lo que realmente se cobró
                    precio_original = float(rs.precio_original)  # Precio sin descuento
                    
                    ingresos_reales_totales += precio_real
                    ingresos_sin_descuento_totales += precio_original
                    
                    # Contar si hubo descuento
                    if precio_original > precio_real:
                        descuento = precio_original - precio_real
                        total_descuentos_otorgados += descuento
                        reservas_con_descuento += 1
                
                elif rs.precio_aplicado is not None and rs.precio_original is None:
                    # Caso 2: Solo tiene precio_aplicado, usar como ambos valores
                    precio_real = float(rs.precio_aplicado)
                    ingresos_reales_totales += precio_real
                    ingresos_sin_descuento_totales += precio_real
                
                else:
                    # Caso 3: No tiene precios guardados, usar el precio actual del servicio
                    precio_servicio = float(rs.servicio.precio)
                    ingresos_reales_totales += precio_servicio
                    ingresos_sin_descuento_totales += precio_servicio
        
        # Calcular el porcentaje de descuento promedio
        porcentaje_descuento_promedio = 0
        if ingresos_sin_descuento_totales > 0:
            porcentaje_descuento_promedio = round((total_descuentos_otorgados / ingresos_sin_descuento_totales) * 100, 1)
        
        print(f"📊 DEBUG INGRESOS:")
        print(f"   Reservas completadas: {reservas_completadas_todas.count()}")
        print(f"   Ingresos reales totales: COP ${ingresos_reales_totales:,.2f}")
        print(f"   Ingresos sin descuento: COP ${ingresos_sin_descuento_totales:,.2f}")
        print(f"   Total descuentos otorgados: COP ${total_descuentos_otorgados:,.2f}")
        print(f"   Servicios con descuento: {reservas_con_descuento}")
        print(f"   Porcentaje promedio: {porcentaje_descuento_promedio}%")
        
    except Exception as e:
        print(f"❌ Error calculando ingresos reales: {e}")
        import traceback
        traceback.print_exc()
        ingresos_reales_totales = 0
        ingresos_sin_descuento_totales = 0
        total_descuentos_otorgados = 0
        porcentaje_descuento_promedio = 0
        reservas_con_descuento = 0
    
    # === DATOS PARA GRÁFICOS ===
    # Reservas por día en la última semana
    reservas_por_dia = []
    for i in range(7):
        fecha = hoy - timedelta(days=6-i)
        total_dia = Reserva.objects.filter(fecha=fecha).count()
        reservas_por_dia.append({
            'fecha': fecha.strftime('%d/%m'),
            'total': total_dia
        })
    
    # Empresas más activas
    empresas_activas = Empresa.objects.annotate(
        total_reservas=Count('reserva')
    ).filter(total_reservas__gt=0).order_by('-total_reservas')[:5]
    
    # === RANKINGS Y PREMIOS ===
    # Top 3 clientes más frecuentes (por número de reservas)
    top_clientes = Usuario.objects.filter(rol='cliente').annotate(
        total_reservas=Count('reserva')
    ).filter(total_reservas__gt=0).order_by('-total_reservas')[:3]
    
    # Top 3 empresas del mes (por reservas del mes actual)
    top_empresas_mes = Empresa.objects.annotate(
        reservas_mes=Count('reserva', filter=Q(reserva__fecha__gte=inicio_mes))
    ).filter(reservas_mes__gt=0).order_by('-reservas_mes')[:3]
    
    # Servicios mejor valorados (por número de reservas completadas)
    servicios_mejor_valorados = Servicio.objects.annotate(
        total_reservas=Count('reserva', filter=Q(reserva__estado='completado'))
    ).filter(total_reservas__gt=0).order_by('-total_reservas')[:3]
    
    context = {
        # Estadísticas principales
        'total_usuarios': total_usuarios,
        'usuarios_mes': usuarios_mes,
        'usuarios_activos': usuarios_activos,
        'usuarios_admin': usuarios_admin,
        'usuarios_cliente': usuarios_cliente,
        
        'total_empresas': total_empresas,
        'empresas_verificadas': empresas_verificadas,
        'empresas_pendientes': empresas_pendientes,
        
        'total_reservas': total_reservas,
        'reservas_hoy': reservas_hoy,
        'reservas_pendientes': reservas_pendientes,
        'reservas_completadas': reservas_completadas,
        'reservas_canceladas': reservas_canceladas,
        'reservas_semana': reservas_semana,
        'reservas_mes': reservas_mes,
        
        'total_servicios': total_servicios,
        'servicios_mas_reservados': servicios_mas_reservados,
        
        'total_planes': total_planes,
        'total_planes_empresariales': total_planes_empresariales,
        'total_suscripciones_activas': total_suscripciones_activas,
        'total_suscripciones_empresariales': total_suscripciones_empresariales,
        
        'total_comentarios': total_comentarios,
        'comentarios_mes': comentarios_mes,
        'total_quejas': total_quejas,
        'quejas_pendientes': quejas_pendientes,
        'quejas_respondidas': quejas_respondidas,
        
        # Métricas de rendimiento
        'tasa_completadas': tasa_completadas,
        'tasa_canceladas': tasa_canceladas,
        'porcentaje_pendientes': porcentaje_pendientes,
        'ingresos_estimados_mes': round(ingresos_estimados_mes, 2),
        
        # Nuevas métricas de ingresos con descuentos (TOTALES)
        'ingresos_reales_totales': round(ingresos_reales_totales, 2),
        'ingresos_sin_descuento_totales': round(ingresos_sin_descuento_totales, 2),
        'total_descuentos_otorgados': round(total_descuentos_otorgados, 2),
        'porcentaje_descuento_promedio': porcentaje_descuento_promedio,
        'reservas_completadas': reservas_completadas,
        'reservas_con_descuento': reservas_con_descuento,
        
        # Datos para actividad reciente
        'reservas_recientes': reservas_recientes,
        'comentarios_recientes': comentarios_recientes,
        'empresas_recientes': empresas_recientes,
        
        # Datos para gráficos
        'reservas_por_dia': reservas_por_dia,
        'reservas_por_estado': list(reservas_por_estado),
        'empresas_activas': empresas_activas,
        
        # Rankings y Premios
        'top_clientes': top_clientes,
        'top_empresas_mes': top_empresas_mes,
        'servicios_mejor_valorados': servicios_mejor_valorados,
        
        # Información adicional
        'fecha_hoy': hoy,
        'usuario_actual': request.user,
    }
    
    return render(request, 'home_crud.html', context)

@empresa_required
def home_empresas(request):
    """Vista del home para empresas"""
    try:
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Obtener estadísticas básicas de la empresa
        reservas_empresa = Reserva.objects.filter(empresa=empresa)
        total_reservas = reservas_empresa.count()
        reservas_completadas = reservas_empresa.filter(estado='completado').count()
        reservas_pendientes = reservas_empresa.filter(estado='pendiente').count()
        
        # Reservas de hoy y próximas
        hoy = timezone.now().date()
        reservas_hoy = reservas_empresa.filter(fecha=hoy).count()
        reservas_proximas = reservas_empresa.filter(
            fecha__gte=hoy,
            fecha__lte=hoy + timedelta(days=7),
            estado='pendiente'
        ).order_by('fecha', 'hora')[:5]
        
        # Actividad reciente (últimas 10 reservas)
        actividad_reciente = []
        reservas_recientes = reservas_empresa.order_by('-fecha', '-hora')[:10]
        
        for reserva in reservas_recientes:
            servicios = reserva.servicios.all()
            precio_total = sum(servicio.precio for servicio in servicios)
            
            actividad_reciente.append({
                'reserva': reserva,
                'servicios': servicios,
                'precio_total': precio_total,
                'tiempo_transcurrido': timezone.now().date() - reserva.fecha if reserva.fecha <= timezone.now().date() else None
            })
        
        # Obtener servicios de la empresa
        servicios_empresa = empresa.servicios.all()
        
        # Ingresos del mes
        primer_dia_mes = hoy.replace(day=1)
        reservas_mes = reservas_empresa.filter(
            fecha__gte=primer_dia_mes,
            fecha__lte=hoy,
            estado='completado'
        )
        
        print(f"🔍 Debug - Fecha actual: {hoy}")
        print(f"🔍 Debug - Primer día del mes: {primer_dia_mes}")
        print(f"🔍 Debug - Reservas del mes completadas: {reservas_mes.count()}")
        
        ingresos_mes = 0
        for reserva in reservas_mes:
            servicios = reserva.servicios.all()
            precio_reserva = sum(servicio.precio for servicio in servicios)
            print(f"🔍 Debug - Reserva {reserva.id_reserva}: servicios={servicios.count()}, precio={precio_reserva}")
            ingresos_mes += precio_reserva
            
        print(f"🔍 Debug - Total ingresos del mes: {ingresos_mes}")
        
        # También calcular ingresos totales como respaldo
        ingresos_totales = 0
        todas_reservas_completadas = reservas_empresa.filter(estado='completado')
        for reserva in todas_reservas_completadas:
            servicios = reserva.servicios.all()
            ingresos_totales += sum(servicio.precio for servicio in servicios)
        
        print(f"🔍 Debug - Ingresos totales (todas las reservas): {ingresos_totales}")
        
        # Convertir a enteros para evitar decimales en el template
        ingresos_mes = int(ingresos_mes) if ingresos_mes else 0
        ingresos_totales = int(ingresos_totales) if ingresos_totales else 0
        
        context = {
            'empresa': empresa,
            'total_reservas': total_reservas,
            'reservas_completadas': reservas_completadas,
            'reservas_pendientes': reservas_pendientes,
            'reservas_hoy': reservas_hoy,
            'reservas_proximas': reservas_proximas,
            'actividad_reciente': actividad_reciente,
            'servicios_empresa': servicios_empresa,
            'ingresos_mes': ingresos_mes,
            'ingresos_totales': ingresos_totales,
        }
        
        return render(request, 'home_empresas.html', context)
        
    except Empresa.DoesNotExist:
        messages.error(request, 'Error: Empresa no encontrada.')
        return redirect('logincrud')
    except Exception as e:
        print(f"❌ Error en home_empresas: {str(e)}")
        messages.error(request, 'Error interno. Intente nuevamente.')
        return redirect('logincrud')

def logout_empresa(request):
    """Vista para cerrar sesión de empresa"""
    empresa_nombre = request.session.get('empresa_nombre', 'Empresa')
    
    # Limpiar todas las variables de sesión relacionadas con la empresa
    request.session.pop('empresa_id', None)
    request.session.pop('empresa_nombre', None)
    request.session.pop('empresa_email', None)
    request.session.pop('es_empresa', None)
    
    messages.success(request, f'Has cerrado sesión correctamente, {empresa_nombre}.')
    return redirect('home')  # Redirigir a la página principal

@empresa_required
def citas_empresa(request):
    """Vista para mostrar todas las citas de la empresa con paginación"""
    try:
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Obtener todas las reservas de la empresa
        reservas = Reserva.objects.filter(empresa=empresa).order_by('-fecha', '-hora')
        
        # Obtener reservas con sus servicios
        reservas_con_servicios = []
        for reserva in reservas:
            servicios = reserva.servicios.all()
            # Preparar un número legible para la reserva: usar campo numero_reserva si existe, sino formatear con prefijo
            numero_reserva = getattr(reserva, 'numero_reserva', None)
            if not numero_reserva:
                # Formato ANW-00000001 (padding 8)
                numero_reserva = f"ANW-{reserva.id_reserva:08d}"

            reservas_con_servicios.append({
                'reserva': reserva,
                'numero_reserva': numero_reserva,
                'servicios': servicios,
                'total_servicios': servicios.count(),
                'precio_total': sum(servicio.precio for servicio in servicios)
            })
        
        # Configurar paginación
        items_per_page = int(request.GET.get('per_page', 10))  # Por defecto 10 items por página
        paginator = Paginator(reservas_con_servicios, items_per_page)
        
        page = request.GET.get('page')
        try:
            reservas_paginadas = paginator.page(page)
        except PageNotAnInteger:
            # Si la página no es un entero, mostrar la primera página
            reservas_paginadas = paginator.page(1)
        except EmptyPage:
            # Si la página está fuera de rango, mostrar la última página
            reservas_paginadas = paginator.page(paginator.num_pages)
        
        # Estadísticas
        total_reservas = reservas.count()
        reservas_completadas = reservas.filter(estado='completado').count()
        reservas_pendientes = reservas.filter(estado='pendiente').count()
        reservas_canceladas = reservas.filter(estado='cancelada').count()
        # Reservas por estado
        reservas_hoy = reservas.filter(fecha=timezone.now().date())
        reservas_semana = reservas.filter(
            fecha__gte=timezone.now().date(),
            fecha__lte=timezone.now().date() + timedelta(days=7)
        )
        
        # Obtener servicios de la empresa para filtros
        servicios_empresa = empresa.servicios.all()
        
        context = {
            'empresa': empresa,
            'reservas_con_servicios': reservas_paginadas,  # Ahora es paginado
            'total_reservas': total_reservas,
            'reservas_completadas': reservas_completadas,
            'reservas_pendientes': reservas_pendientes,
            'reservas_canceladas': reservas_canceladas,
            'reservas_hoy': reservas_hoy.count(),
            'reservas_semana': reservas_semana.count(),
            'servicios_empresa': servicios_empresa,
            'paginator': paginator,
            'page_obj': reservas_paginadas,
            'items_per_page': items_per_page,
        }

        return render(request, 'empresas/citas_empresa.html', context)

    except Empresa.DoesNotExist:
        messages.error(request, 'Error: Empresa no encontrada.')
        return redirect('logincrud')
    except Exception as e:
        print(f"❌ Error en citas_empresa: {str(e)}")
        messages.error(request, 'Error interno. Intente nuevamente.')
        return redirect('home_empresas')

@empresa_required
def actualizar_estado_cita(request):
    """Vista para actualizar el estado de una cita"""
    if request.method == 'POST':
        try:
            reserva_id = request.POST.get('reserva_id')
            nuevo_estado = request.POST.get('nuevo_estado')
            
            empresa_id = request.session.get('empresa_id')
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            
            # Verificar que la reserva pertenece a la empresa
            reserva = Reserva.objects.get(id_reserva=reserva_id, empresa=empresa)
            
            if nuevo_estado in ['completado', 'pendiente', 'cancelada']:
                reserva.estado = nuevo_estado
                reserva.save()
                if nuevo_estado == 'completado':
                    estado_texto = 'completada'
                elif nuevo_estado == 'pendiente':
                    estado_texto = 'pendiente'
                else:
                    estado_texto = 'cancelada'
                messages.success(request, f'Cita marcada como {estado_texto} exitosamente.')
            else:
                messages.error(request, 'Estado no válido.')
                
        except Reserva.DoesNotExist:
            messages.error(request, 'Cita no encontrada o no pertenece a tu empresa.')
        except Exception as e:
            print(f"❌ Error al actualizar estado: {str(e)}")
            messages.error(request, 'Error al actualizar el estado de la cita.')
    
    return redirect('citas_empresa')

@empresa_required
def generar_codigo_qr_reserva(request, reserva_id):
    """Vista para generar y mostrar el código QR de una reserva"""
    try:
        import qrcode
        import qrcode.image.svg
        from io import BytesIO
        import base64
        
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Verificar que la reserva pertenece a la empresa
        reserva = Reserva.objects.get(id_reserva=reserva_id, empresa=empresa)
        
        # Crear los datos para el código QR
        qr_data = {
            'numero_reserva': reserva.numero_reserva,
            'reserva_id': reserva.id_reserva,
            'empresa_id': empresa.id_empresa,
            'usuario_id': reserva.usuario.id_usuario,
            'accion': 'completar_reserva'
        }
        
        # Crear la URL que el cliente escaneará
        qr_url = f"{request.build_absolute_uri('/')[:-1]}/completar-reserva/{reserva.numero_reserva}/"
        
        # Generar el código QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        
        # Crear imagen del QR
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convertir a base64 para mostrar en el template
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        qr_image_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Obtener información de servicios
        servicios_nombres = []
        try:
            servicios_reserva = ReservaServicio.objects.filter(reserva=reserva)
            servicios_nombres = [rs.servicio.nombre_servicio for rs in servicios_reserva]
        except:
            servicios_nombres = ['No especificado']
        
        # Retornar datos JSON para el modal
        response_data = {
            'success': True,
            'qr_image': qr_image_base64,
            'qr_url': qr_url,
            'numero_reserva': reserva.numero_reserva,
            'reserva_id': reserva.id_reserva,
            'cliente_nombre': reserva.usuario.nombre_completo,
            'servicio_nombre': ', '.join(servicios_nombres),
            'fecha_reserva': f"{reserva.fecha.strftime('%d/%m/%Y')} {reserva.hora.strftime('%H:%M')}",
            'empresa_nombre': empresa.nombre_empresa
        }
        
        return JsonResponse(response_data)
        
    except Reserva.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Reserva no encontrada o no pertenece a tu empresa.'
        })
    except Exception as e:
        print(f"❌ Error al generar código QR: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Error al generar el código QR.'
        })


def completar_reserva(request, numero_reserva):
    """Vista pública que muestra información mínima sobre el proceso de completado.
    Si se accede desde el navegador, mostramos una página simple con instrucciones.
    """
    try:
        reserva = Reserva.objects.get(numero_reserva=numero_reserva)
        # Mostrar una página sencilla indicando que el usuario debe escanear desde la app
        return render(request, 'reservas/completar_reserva.html', {'numero_reserva': numero_reserva, 'reserva': reserva})
    except Reserva.DoesNotExist:
        messages.error(request, 'Reserva no encontrada.')
        return redirect('citas')


@login_required
def ajax_completar_reserva(request):
    """Endpoint AJAX que recibe JSON { numero_reserva: 'ANW-00000001' }
    Verifica que la reserva exista y pertenezca al usuario autenticado, marca como completada.
    Retorna JSON { success: True/False, message: '...' }
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

    try:
        import json
        data = json.loads(request.body.decode('utf-8'))
        numero_reserva = data.get('numero_reserva')
        if not numero_reserva:
            return JsonResponse({'success': False, 'message': 'Número de reserva faltante.'}, status=400)

        try:
            reserva = Reserva.objects.get(numero_reserva=numero_reserva)
        except Reserva.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Reserva no encontrada.'}, status=404)

        # Verificar que la reserva pertenece al usuario que realiza la petición
        if reserva.usuario != request.user:
            return JsonResponse({'success': False, 'message': 'No tienes permiso para completar esta reserva.'}, status=403)

        # Sólo permitir completar si está en estado pendiente
        if reserva.estado == 'completado':
            return JsonResponse({'success': True, 'message': 'Reserva ya estaba completada.'})

        reserva.estado = 'completado'
        reserva.save()

        return JsonResponse({'success': True, 'message': 'Reserva marcada como completada.'})

    except Exception as e:
        print(f"❌ Error en ajax_completar_reserva: {e}")
        return JsonResponse({'success': False, 'message': 'Error interno.'}, status=500)

@empresa_required
def detalle_reserva_empresa(request, reserva_id):
    """Vista para ver los detalles de una reserva desde la perspectiva de la empresa"""
    try:
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Verificar que la reserva pertenece a la empresa
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, empresa=empresa)
        
        # Obtener servicios de la reserva
        servicios = reserva.servicios.all()
        precio_total = sum(servicio.precio for servicio in servicios)
        
        context = {
            'reserva': reserva,
            'servicios': servicios,
            'precio_total': precio_total,
            'empresa': empresa,
        }
        
        # Devolver JSON para AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = {
                'id_reserva': reserva.id_reserva,
                'cliente': {
                    'nombre': reserva.usuario.nombre_completo,
                    'correo': reserva.usuario.correo,
                    'telefono': getattr(reserva.usuario, 'telefono', 'No disponible'),
                    'id': reserva.usuario.id_usuario,
                },
                'fecha': reserva.fecha.strftime('%d/%m/%Y'),
                'hora': reserva.hora.strftime('%H:%M'),
                'estado': reserva.estado,
                'servicios': [
                    {
                        'nombre': servicio.nombre_servicio,
                        'precio': float(servicio.precio),
                        'descripcion': getattr(servicio, 'descripcion', '')
                    }
                    for servicio in servicios
                ],
                'precio_total': float(precio_total),
                'fecha_creacion': reserva.fecha_registro.strftime('%d/%m/%Y %H:%M') if hasattr(reserva, 'fecha_registro') else 'No disponible',
            }
            return JsonResponse(data)
        
        return render(request, 'empresas/detalle_reserva.html', context)
        
    except Reserva.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Reserva no encontrada'}, status=404)
        messages.error(request, 'Reserva no encontrada.')
        return redirect('citas_empresa')
    except Exception as e:
        print(f"❌ Error en detalle_reserva_empresa: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Error interno'}, status=500)
        messages.error(request, 'Error al cargar los detalles de la reserva.')
        return redirect('citas_empresa')

@empresa_required  
def editar_reserva_empresa(request, reserva_id):
    """Vista para editar una reserva desde la perspectiva de la empresa"""
    try:
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Verificar que la reserva pertenece a la empresa
        reserva = get_object_or_404(Reserva, id_reserva=reserva_id, empresa=empresa)
        
        # Verificar que la reserva se puede editar (no completada ni cancelada)
        if reserva.estado in ['completado', 'cancelada']:
            messages.error(request, f'No se pueden editar reservas en estado "{reserva.estado}".')
            return redirect('citas_empresa')
        
        if request.method == 'POST':
            nueva_fecha = request.POST.get('fecha')
            nueva_hora = request.POST.get('hora')
            nuevo_estado = request.POST.get('estado')
            
            # Validaciones
            if not nueva_fecha or not nueva_hora:
                messages.error(request, 'Fecha y hora son obligatorias.')
                return redirect('editar_reserva_empresa', reserva_id=reserva_id)
            
            try:
                # Convertir fecha
                fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
                
                # Convertir hora (asumiendo formato 24h)
                hora_obj = datetime.strptime(nueva_hora, '%H:%M').time()
                
                # Verificar que no haya conflicto con otra reserva
                conflicto = Reserva.objects.filter(
                    empresa=empresa,
                    fecha=fecha_obj,
                    hora=hora_obj
                ).exclude(id_reserva=reserva_id).exists()
                
                if conflicto:
                    messages.error(request, 'Ya existe una cita en esa fecha y hora.')
                    return redirect('editar_reserva_empresa', reserva_id=reserva_id)
                
                # Actualizar reserva
                reserva.fecha = fecha_obj
                reserva.hora = hora_obj
                
                if nuevo_estado and nuevo_estado in ['pendiente', 'completado', 'cancelada']:
                    reserva.estado = nuevo_estado
                
                reserva.save()
                
                messages.success(request, 'Cita actualizada exitosamente.')
                return redirect('citas_empresa')
                
            except ValueError as e:
                messages.error(request, 'Formato de fecha u hora inválido.')
                return redirect('editar_reserva_empresa', reserva_id=reserva_id)
        
        # GET: Mostrar formulario de edición
        servicios = reserva.servicios.all()
        precio_total = sum(servicio.precio for servicio in servicios)
        
        context = {
            'reserva': reserva,
            'servicios': servicios,
            'precio_total': precio_total,
            'empresa': empresa,
        }
        
        return render(request, 'empresas/editar_reserva.html', context)
        
    except Reserva.DoesNotExist:
        messages.error(request, 'Reserva no encontrada.')
        return redirect('citas_empresa')
    except Exception as e:
        print(f"❌ Error en editar_reserva_empresa: {str(e)}")
        messages.error(request, 'Error al cargar la reserva.')
        return redirect('citas_empresa')

@empresa_required
def reportes_empresa(request):
    """Vista para mostrar reportes y estadísticas de la empresa"""
    empresa = get_object_or_404(Empresa, email=request.session.get('empresa_email'))
    
    # Obtener período seleccionado (por defecto últimos 12 meses)
    periodo = int(request.GET.get('periodo', 12))
    
    # Estadísticas generales
    total_reservas = Reserva.objects.filter(empresa=empresa).count()
    reservas_completadas = Reserva.objects.filter(empresa=empresa, estado='completado').count()
    reservas_pendientes = Reserva.objects.filter(empresa=empresa, estado='pendiente').count()
    
    # Estadísticas por mes (según período seleccionado)
    # Usar la fecha actual para asegurar que incluimos todos los meses hasta hoy
    from datetime import datetime, date, timedelta
    import calendar
    from django.db.models import Sum, Count, Q
    
    # Usar datetime.now() para obtener la fecha actual exacta
    now = datetime.now()
    hoy = now.date()
    
    reservas_por_mes = []
    ingresos_por_mes = []
    meses = []
    
    # Debug: Mostrar la fecha actual que se está usando
    print(f"🕒 Debug - Fecha/hora actual: {now}")
    print(f"🕒 Debug - Fecha actual utilizada: {hoy} (Mes: {hoy.month}, Año: {hoy.year})")
    print(f"🕒 Debug - Generando los últimos 6 meses incluyendo el actual:")
    print(f"   Esperado: Jul 2025, Jun 2025, May 2025, Abr 2025, Mar 2025, Feb 2025")
    
    # Generar lista de meses desde el más reciente hacia atrás
    meses_a_procesar = []
    
    for i in range(periodo):
        # Usar aritmética simple de meses
        mes_target = hoy.month - i
        ano_target = hoy.year
        
        # Ajustar si el mes es menor a 1
        while mes_target <= 0:
            mes_target += 12
            ano_target -= 1
        
        meses_a_procesar.append({
            'mes': mes_target,
            'ano': ano_target,
            'nombre': f"{calendar.month_name[mes_target][:3]} {ano_target}"
        })
        
        print(f"  📊 Mes {i}: {mes_target}/{ano_target} ({calendar.month_name[mes_target]})")
    
    # Procesar cada mes en orden cronológico inverso (más reciente primero)
    for mes_info in meses_a_procesar:
        mes_actual = mes_info['mes']
        ano_actual = mes_info['ano']
        
        # Obtener reservas del mes
        reservas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual
        ).count()
        
        # Calcular ingresos del mes
        ingresos_mes = 0
        reservas_completadas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual,
            estado='completado'
        )
        
        for reserva in reservas_completadas_mes:
            try:
                precio_total = sum([rs.servicio.precio for rs in ReservaServicio.objects.filter(reserva=reserva)])
                ingresos_mes += precio_total
            except:
                continue
        
        # Agregar al principio para mantener orden cronológico inverso para las gráficas
        reservas_por_mes.insert(0, reservas_mes)
        ingresos_por_mes.insert(0, round(ingresos_mes, 2))
        meses.insert(0, mes_info['nombre'])
        
        print(f"    → Reservas: {reservas_mes}, Ingresos: {round(ingresos_mes, 2)}")
    
    print(f"📊 Debug - Meses finales: {meses}")
    print(f"📊 Debug - Reservas finales: {reservas_por_mes}")
    print(f"📊 Debug - Ingresos finales: {ingresos_por_mes}")
    
    # Servicios más solicitados
    servicios_populares = []
    for servicio in Servicio.objects.all():
        count = ReservaServicio.objects.filter(
            reserva__empresa=empresa,
            servicio=servicio
        ).count()
        if count > 0:
            porcentaje = (count / total_reservas * 100) if total_reservas > 0 else 0
            servicios_populares.append({
                'servicio': servicio,
                'count': count,
                'ingresos': round(count * servicio.precio, 2),
                'porcentaje': round(porcentaje, 1)
            })
    
    servicios_populares = sorted(servicios_populares, key=lambda x: x['count'], reverse=True)[:5]
    
    # Estadísticas por día de la semana
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    reservas_por_dia = []
    
    for i in range(7):
        # Django week_day: 1=Sunday, 2=Monday, etc.
        dia_django = (i + 1) % 7 + 1  # Convertir para que Monday=2
        if dia_django == 1:  # Sunday
            dia_django = 8
        
        count = Reserva.objects.filter(
            empresa=empresa,
            fecha__week_day=dia_django
        ).count()
        reservas_por_dia.append(count)
    
    # Ingresos totales
    ingresos_totales = sum(ingresos_por_mes)
    
    # Promedio de reservas por día
    from datetime import timedelta
    dias_transcurridos = max((hoy - empresa.fecha_registro.date()).days, 1)
    promedio_reservas_dia = total_reservas / dias_transcurridos
    
    # Tasa de completado
    tasa_completado = (reservas_completadas / total_reservas * 100) if total_reservas > 0 else 0
    
    # Reservas del mes actual
    reservas_mes_actual = Reserva.objects.filter(
        empresa=empresa,
        fecha__month=hoy.month,
        fecha__year=hoy.year
    ).count()
    
    # Ingresos del mes actual
    ingresos_mes_actual = 0
    for reserva in Reserva.objects.filter(empresa=empresa, fecha__month=hoy.month, fecha__year=hoy.year, estado='completado'):
        try:
            precio_total = sum([rs.servicio.precio for rs in ReservaServicio.objects.filter(reserva=reserva)])
            ingresos_mes_actual += precio_total
        except:
            continue
    
    # Calcular crecimiento mes anterior
    mes_anterior = hoy.month - 1
    ano_anterior = hoy.year
    if mes_anterior <= 0:
        mes_anterior = 12
        ano_anterior -= 1
    
    reservas_mes_anterior = Reserva.objects.filter(
        empresa=empresa,
        fecha__month=mes_anterior,
        fecha__year=ano_anterior
    ).count()
    
    crecimiento_reservas = 0
    if reservas_mes_anterior > 0:
        crecimiento_reservas = ((reservas_mes_actual - reservas_mes_anterior) / reservas_mes_anterior) * 100
    
    # Crear datos estructurados para la tabla (mostrar los últimos 6 meses más recientes incluyendo el actual)
    datos_tabla = []
    
    # Generar los últimos 6 meses incluyendo el mes actual (Julio 2025)
    meses_tabla = []
    for i in range(6):  # 6 meses: Julio, Junio, Mayo, Abril, Marzo, Febrero
        mes_target = hoy.month - i
        ano_target = hoy.year
        
        # Ajustar si el mes es menor a 1
        while mes_target <= 0:
            mes_target += 12
            ano_target -= 1
        
        mes_nombre = f"{calendar.month_name[mes_target][:3]} {ano_target}"
        meses_tabla.append({
            'mes': mes_target,
            'ano': ano_target,
            'nombre': mes_nombre
        })
        
        print(f"  📊 Tabla - Mes {i}: {mes_target}/{ano_target} ({mes_nombre})")
    
    # Procesar cada mes para la tabla
    for i, mes_info in enumerate(meses_tabla):
        mes_actual = mes_info['mes']
        ano_actual = mes_info['ano']
        # Obtener reservas del mes
        reservas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual
        ).count()
        # Calcular reservas completadas del mes
        completadas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual,
            estado='completado'
        ).count()
        # Calcular reservas pendientes del mes
        pendientes_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual,
            estado='pendiente'
        ).count()
        # Calcular reservas canceladas del mes
        canceladas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual,
            estado='cancelada'
        ).count()
        # Calcular ingresos del mes
        ingresos_mes = 0
        reservas_completadas_mes = Reserva.objects.filter(
            empresa=empresa,
            fecha__month=mes_actual,
            fecha__year=ano_actual,
            estado='completado'
        )
        for reserva in reservas_completadas_mes:
            try:
                precio_total = sum([rs.servicio.precio for rs in ReservaServicio.objects.filter(reserva=reserva)])
                ingresos_mes += precio_total
            except:
                continue
        # Calcular porcentaje de progreso basado en el máximo de reservas de todos los meses
        max_reservas_tabla = 1  # Valor por defecto
        for mes_temp in meses_tabla:
            reservas_temp = Reserva.objects.filter(
                empresa=empresa,
                fecha__month=mes_temp['mes'],
                fecha__year=mes_temp['ano']
            ).count()
            if reservas_temp > max_reservas_tabla:
                max_reservas_tabla = reservas_temp
        porcentaje_progreso = (reservas_mes / max_reservas_tabla * 100) if max_reservas_tabla > 0 else 0
        # Calcular crecimiento comparando con el mes anterior
        # Calcular crecimiento comparando con el mes anterior
        crecimiento_mes = 0
        crecimiento_valido = False
        if i < len(meses_tabla) - 1:  # Si no es el último mes (más antiguo)
            mes_anterior_info = meses_tabla[i + 1]
            reservas_mes_anterior = Reserva.objects.filter(
                empresa=empresa,
                fecha__month=mes_anterior_info['mes'],
                fecha__year=mes_anterior_info['ano']
            ).count()
            if reservas_mes_anterior > 0:
                crecimiento_mes = ((reservas_mes - reservas_mes_anterior) / reservas_mes_anterior) * 100
                crecimiento_valido = True
            else:
                # Cuando el mes anterior tiene 0 reservas no es posible calcular un % de cambio
                crecimiento_mes = 0
                crecimiento_valido = False
        else:
            # Mes más antiguo: base de comparación
            crecimiento_mes = 0
            crecimiento_valido = False

        datos_tabla.append({
            'mes': mes_info['nombre'],
            'reservas': reservas_mes,
            'completadas': completadas_mes,
            'pendientes': pendientes_mes,
            'canceladas': canceladas_mes,
            'porcentaje_progreso': round(porcentaje_progreso, 1),
            'ingresos': round(ingresos_mes, 2),
            'crecimiento': round(crecimiento_mes, 1) if crecimiento_valido else 0,
            'crecimiento_valido': crecimiento_valido
        })
        print(f"    → Tabla {mes_info['nombre']}: Reservas: {reservas_mes}, Completadas: {completadas_mes}, Pendientes: {pendientes_mes}, Canceladas: {canceladas_mes}, Ingresos: {round(ingresos_mes, 2)}")
    
    print(f"🔍 Debug - Datos tabla final generados: {len(datos_tabla)} filas")
    for i, dato in enumerate(datos_tabla):
        print(f"  Fila {i}: {dato}")
    
    # Convertir datos a JSON para JavaScript
    import json
    
    context = {
        'empresa': empresa,
        'periodo': periodo,
        'total_reservas': total_reservas,
        'reservas_completadas': reservas_completadas,
        'reservas_pendientes': reservas_pendientes,
        'reservas_por_mes': json.dumps(reservas_por_mes),
        'ingresos_por_mes': json.dumps(ingresos_por_mes),
        'meses': json.dumps(meses),
        'servicios_populares': servicios_populares,
        'reservas_por_dia': json.dumps(reservas_por_dia),
        'dias_semana': json.dumps(dias_semana),
        'ingresos_totales': round(ingresos_totales, 2),
        'promedio_reservas_dia': round(promedio_reservas_dia, 1),
        'tasa_completado': round(tasa_completado, 1),
        'reservas_mes_actual': reservas_mes_actual,
        'ingresos_mes_actual': round(ingresos_mes_actual, 2),
        'crecimiento_reservas': round(crecimiento_reservas, 1),
        'datos_tabla': datos_tabla,
    }
    
    return render(request, 'empresas/reportes_empresa.html', context)


# =============================================================================
# VISTAS PARA MANEJO DE PLANES Y SUSCRIPCIONES
# =============================================================================

def planes_view(request):
    """Vista para mostrar todos los planes disponibles"""
    planes_list = Plan.objects.filter(activo=True).order_by('precio_mensual')
    
    # Paginación para planes - 4 por página
    planes_paginator = Paginator(planes_list, 4)  # Mostrar 4 planes por página
    planes_page = request.GET.get('page')
    
    try:
        planes = planes_paginator.page(planes_page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        planes = planes_paginator.page(1)
    except EmptyPage:
        # Si page está fuera del rango, mostrar la última página
        planes = planes_paginator.page(planes_paginator.num_pages)
    
    # Si el usuario está logueado, obtener su suscripción actual
    suscripcion_actual = None
    if request.user.is_authenticated:
        suscripcion_actual = SuscripcionUsuario.objects.filter(
            usuario=request.user,
            estado='activa'
        ).first()
    
    context = {
        'planes': planes,
        'suscripcion_actual': suscripcion_actual,
    }
    return render(request, 'planes/planes.html', context)

def planes_ajax(request):
    """Vista AJAX para cargar planes paginados sin recargar la página"""
    planes_list = Plan.objects.filter(activo=True).order_by('precio_mensual')
    
    # Paginación para planes - 4 por página
    paginator = Paginator(planes_list, 4)
    page = request.GET.get('page')
    
    try:
        planes = paginator.page(page)
    except PageNotAnInteger:
        planes = paginator.page(1)
    except EmptyPage:
        planes = paginator.page(paginator.num_pages)
    
    # Si el usuario está logueado, obtener su suscripción actual
    suscripcion_actual = None
    if request.user.is_authenticated:
        suscripcion_actual = SuscripcionUsuario.objects.filter(
            usuario=request.user,
            estado='activa'
        ).first()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Respuesta AJAX - solo renderizar la sección de planes
        return render(request, 'planes/planes_partial.html', {
            'planes': planes,
            'suscripcion_actual': suscripcion_actual
        })
    else:
        # Fallback - redireccionar a planes si no es AJAX
        return redirect('planes')

@login_required
def suscribirse_plan(request, plan_id):
    """Vista para suscribirse a un plan"""
    plan = get_object_or_404(Plan, id_plan=plan_id, activo=True)
    
    # Verificar si el usuario ya tiene una suscripción activa
    suscripcion_existente = SuscripcionUsuario.objects.filter(
        usuario=request.user,
        estado='activa'
    ).first()
    
    if suscripcion_existente:
        messages.warning(request, 'Ya tienes una suscripción activa. Cancela la actual para cambiar de plan.')
        return redirect('mi_suscripcion')
    
    if request.method == 'POST':
        # Crear nueva suscripción
        suscripcion = SuscripcionUsuario.objects.create(
            usuario=request.user,
            plan=plan,
            fecha_inicio=timezone.now(),
            fecha_fin=timezone.now() + timezone.timedelta(days=30)
        )
        
        # Crear registro de pago
        referencia = str(uuid.uuid4())
        HistorialPagosSuscripcion.objects.create(
            suscripcion=suscripcion,
            monto=plan.precio_mensual,
            referencia_pago=referencia,
            metodo_pago='PayU',  # O el método que uses
            estado='pendiente'
        )
        
        # Aquí integrarías con PayU o tu pasarela de pago
        # Por ahora, redirigimos a una página de confirmación
        messages.success(request, f'Te has suscrito al plan {plan.nombre}. Procesando pago...')
        return redirect('procesar_pago_suscripcion', referencia=referencia)
    
    context = {
        'plan': plan,
    }
    return render(request, 'planes/suscribirse.html', context)

@login_required
def mi_suscripcion(request):
    """Vista del dashboard de suscripción del usuario"""
    suscripcion = SuscripcionUsuario.objects.filter(
        usuario=request.user,
        estado='activa'
    ).first()
    
    historial_pagos = []
    servicios_restantes = 0
    
    if suscripcion:
        # Asegurar que se reinicie el contador mensual antes de mostrar la información
        suscripcion.reiniciar_contador_mensual()
        
        historial_pagos = HistorialPagosSuscripcion.objects.filter(
            suscripcion=suscripcion
        ).order_by('-fecha_pago')[:5]
        
        servicios_restantes = suscripcion.servicios_restantes()
        
        # Debug para verificar los cálculos
        print(f"🔍 Mi Suscripción - Usuario: {request.user.nombre_usuario}")
        print(f"📊 Plan: {suscripcion.plan.nombre}")
        print(f"📊 Servicios utilizados este mes: {suscripcion.servicios_utilizados_mes}")
        print(f"📊 Servicios totales permitidos: {suscripcion.plan.cantidad_servicios_mes}")
        print(f"📊 Servicios restantes: {servicios_restantes}")
        print(f"📅 Último reinicio contador: {suscripcion.ultimo_reinicio_contador}")
        print(f"⏰ Fecha fin suscripción: {suscripcion.fecha_fin}")
        print(f"✅ ¿Puede usar servicio?: {suscripcion.puede_usar_servicio()}")
    
    # Obtener reservas recientes relacionadas con la suscripción
    reservas_recientes = Reserva.objects.filter(
        usuario=request.user,
        suscripcion_utilizada=suscripcion
    ).order_by('-fecha')[:5] if suscripcion else []
    
    context = {
        'suscripcion': suscripcion,
        'historial_pagos': historial_pagos,
        'reservas_recientes': reservas_recientes,
        'servicios_restantes': servicios_restantes,
    }
    return render(request, 'planes/mi_suscripcion.html', context)

@login_required
def cancelar_suscripcion(request):
    """Vista para cancelar suscripción"""
    suscripcion = get_object_or_404(
        SuscripcionUsuario,
        usuario=request.user,
        estado='activa'
    )
    
    if request.method == 'POST':
        suscripcion.estado = 'cancelada'
        suscripcion.auto_renovar = False
        suscripcion.save()
        
        messages.success(request, 'Tu suscripción ha sido cancelada exitosamente.')
        return redirect('planes')
    
    context = {
        'suscripcion': suscripcion,
    }
    return render(request, 'planes/cancelar_suscripcion.html', context)

@login_required
def procesar_pago_suscripcion(request, referencia):
    """Vista para procesar el pago de la suscripción"""
    pago = get_object_or_404(HistorialPagosSuscripcion, referencia_pago=referencia)
    
    # Aquí integrarías con PayU
    # Por ahora simulamos un pago exitoso
    if request.method == 'POST':
        pago.estado = 'aprobado'
        pago.save()
        
        # Activar la suscripción
        pago.suscripcion.estado = 'activa'
        pago.suscripcion.save()
        
        messages.success(request, '¡Pago procesado exitosamente! Tu suscripción está activa.')
        return redirect('mi_suscripcion')
    
    context = {
        'pago': pago,
        'plan': pago.suscripcion.plan,
    }
    return render(request, 'planes/procesar_pago.html', context)

# Función auxiliar para verificar si el usuario puede hacer una reserva
def verificar_disponibilidad_suscripcion(usuario):
    """Verifica si el usuario puede hacer una reserva con su suscripción"""
    suscripcion = SuscripcionUsuario.objects.filter(
        usuario=usuario,
        estado='activa'
    ).first()
    
    if not suscripcion:
        return False, "No tienes una suscripción activa"
    
    if not suscripcion.esta_activa():
        return False, "Tu suscripción ha vencido"
    
    if not suscripcion.puede_usar_servicio():
        return False, f"Has agotado tus servicios del mes. Restantes: {suscripcion.servicios_restantes()}"
    
    return True, suscripcion

@admin_required
@ensure_csrf_cookie
def planes_crud(request):
    """Vista CRUD para gestionar planes (solo admin)"""
    planes = Plan.objects.all().order_by('precio_mensual')
    
    # Calcular estadísticas
    planes_activos = planes.filter(activo=True).count()
    precio_promedio = planes.aggregate(promedio=Avg('precio_mensual'))['promedio'] or 0
    
    context = {
        'planes': planes,
        'planes_activos': planes_activos,
        'precio_promedio': precio_promedio,
    }
    return render(request, 'planes/planes_crud.html', context)

@admin_required
def crear_plan(request):
    """Vista para crear un nuevo plan"""
    if request.method == 'POST':
        plan = Plan.objects.create(
            nombre=request.POST.get('nombre'),
            tipo=request.POST.get('tipo'),
            descripcion=request.POST.get('descripcion'),
            precio_mensual=request.POST.get('precio_mensual'),
            cantidad_servicios_mes=request.POST.get('cantidad_servicios_mes', 0),
            incluye_lavado_asientos=request.POST.get('incluye_lavado_asientos') == 'on',
            incluye_aspirado=request.POST.get('incluye_aspirado') == 'on',
            incluye_lavado_exterior=request.POST.get('incluye_lavado_exterior') == 'on',
            incluye_lavado_interior_humedo=request.POST.get('incluye_lavado_interior_humedo') == 'on',
            incluye_encerado=request.POST.get('incluye_encerado') == 'on',
            incluye_detallado_completo=request.POST.get('incluye_detallado_completo') == 'on',
        )
        
        # Agregar servicios incluidos con sus descuentos
        servicios_ids = request.POST.getlist('servicios_incluidos')
        for servicio_id in servicios_ids:
            servicio = Servicio.objects.get(id_servicio=servicio_id)
            # Obtener el descuento específico para este servicio
            descuento_key = f'descuento_{servicio_id}'
            descuento = request.POST.get(descuento_key, 0)
            
            # Crear la relación con el descuento
            PlanServicio.objects.create(
                plan=plan,
                servicio=servicio,
                porcentaje_descuento=descuento if descuento else 0
            )
        
        messages.success(request, f'Plan "{plan.nombre}" creado exitosamente.')
        return redirect('planes_crud')
    
    servicios = Servicio.objects.all()
    context = {
        'servicios': servicios,
    }
    return render(request, 'planes/planes_individuales/crear_plan.html', context)

@admin_required
def editar_plan(request, plan_id):
    """Vista para editar un plan existente"""
    plan = get_object_or_404(Plan, id_plan=plan_id)
    
    if request.method == 'POST':
        plan.nombre = request.POST.get('nombre')
        plan.tipo = request.POST.get('tipo')
        plan.descripcion = request.POST.get('descripcion')
        plan.precio_mensual = request.POST.get('precio_mensual')
        plan.cantidad_servicios_mes = request.POST.get('cantidad_servicios_mes', 0)
        plan.incluye_lavado_asientos = request.POST.get('incluye_lavado_asientos') == 'on'
        plan.incluye_aspirado = request.POST.get('incluye_aspirado') == 'on'
        plan.incluye_lavado_exterior = request.POST.get('incluye_lavado_exterior') == 'on'
        plan.incluye_lavado_interior_humedo = request.POST.get('incluye_lavado_interior_humedo') == 'on'
        plan.incluye_encerado = request.POST.get('incluye_encerado') == 'on'
        plan.incluye_detallado_completo = request.POST.get('incluye_detallado_completo') == 'on'
        plan.save()
        
        # Actualizar servicios incluidos con descuentos
        # Eliminar todas las relaciones existentes
        PlanServicio.objects.filter(plan=plan).delete()
        
        # Crear nuevas relaciones con descuentos
        servicios_ids = request.POST.getlist('servicios_incluidos')
        for servicio_id in servicios_ids:
            servicio = Servicio.objects.get(id_servicio=servicio_id)
            # Obtener el descuento específico para este servicio
            descuento_key = f'descuento_{servicio_id}'
            descuento = request.POST.get(descuento_key, 0)
            
            # Crear la relación con el descuento
            PlanServicio.objects.create(
                plan=plan,
                servicio=servicio,
                porcentaje_descuento=descuento if descuento else 0
            )
        
        messages.success(request, f'Plan "{plan.nombre}" actualizado exitosamente.')
        return redirect('planes_crud')
    
    servicios = Servicio.objects.all()
    # Obtener los servicios ya asignados al plan con sus descuentos
    servicios_plan = {
        ps.servicio.id_servicio: ps.porcentaje_descuento 
        for ps in PlanServicio.objects.filter(plan=plan)
    }
    
    context = {
        'plan': plan,
        'servicios': servicios,
        'servicios_plan': servicios_plan,
    }
    return render(request, 'planes/planes_individuales/editar_plan.html', context)

@admin_required
def eliminar_plan(request, plan_id):
    """Vista para eliminar un plan"""
    plan = get_object_or_404(Plan, id_plan=plan_id)
    
    if request.method == 'POST':
        plan.activo = False
        plan.save()
        messages.success(request, f'Plan "{plan.nombre}" desactivado exitosamente.')
        return redirect('planes_crud')
    
    # Verificar si hay suscripciones activas
    suscripciones_activas = SuscripcionUsuario.objects.filter(
        plan=plan, 
        estado='activa'
    ).count()
    
    context = {
        'plan': plan,
        'suscripciones_activas': suscripciones_activas,
    }
    return render(request, 'planes/planes_individuales/eliminar_plan.html', context)

@admin_required
@require_POST
def toggle_plan_estado(request, plan_id):
    """Vista para activar/desactivar un plan (AJAX)"""
    plan = get_object_or_404(Plan, id_plan=plan_id)
    plan.activo = not plan.activo
    plan.save()
    
    estado_texto = 'activado' if plan.activo else 'desactivado'
    messages.success(request, f'Plan "{plan.nombre}" {estado_texto} exitosamente.')
    
    return JsonResponse({
        'success': True,
        'activo': plan.activo,
        'mensaje': f'Plan {estado_texto} exitosamente'
    })

@admin_required
def suscripciones_crud(request):
    """Vista para gestionar suscripciones (solo admin)"""
    suscripciones = SuscripcionUsuario.objects.all().order_by('-fecha_inicio')
    context = {
        'suscripciones': suscripciones,
    }
    return render(request, 'planes/suscripciones_crud.html', context)


@empresa_required
def perfil_empresa(request):
    """Vista para gestionar el perfil de la empresa"""
    try:
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        if request.method == 'POST':
            form = EmpresaPerfilForm(request.POST, instance=empresa)
            if form.is_valid():
                # Guardar los datos básicos de la empresa
                empresa_actualizada = form.save(commit=False)
                
                # Procesar datos bancarios adicionales del POST
                empresa_actualizada.nit_empresa = request.POST.get('nit_empresa') or None
                empresa_actualizada.razon_social = request.POST.get('razon_social') or None
                empresa_actualizada.regimen_tributario = request.POST.get('regimen_tributario') or None
                
                empresa_actualizada.titular_cuenta = request.POST.get('titular_cuenta') or None
                empresa_actualizada.tipo_documento_titular = request.POST.get('tipo_documento_titular') or None
                empresa_actualizada.numero_documento_titular = request.POST.get('numero_documento_titular') or None
                
                empresa_actualizada.banco = request.POST.get('banco') or None
                empresa_actualizada.tipo_cuenta = request.POST.get('tipo_cuenta') or None
                empresa_actualizada.numero_cuenta = request.POST.get('numero_cuenta') or None
                empresa_actualizada.swift_code = request.POST.get('swift_code') or None
                empresa_actualizada.iban = request.POST.get('iban') or None
                
                empresa_actualizada.email_facturacion = request.POST.get('email_facturacion') or None
                empresa_actualizada.telefono_facturacion = request.POST.get('telefono_facturacion') or None
                empresa_actualizada.responsable_pagos = request.POST.get('responsable_pagos') or None
                empresa_actualizada.notas_bancarias = request.POST.get('notas_bancarias') or None
                
                # Guardar la empresa actualizada
                empresa_actualizada.save()
                
                # Manejar cambio de contraseña si se proporcionó
                contrasena_actual = form.cleaned_data.get('contrasena_actual')
                nueva_contrasena = form.cleaned_data.get('nueva_contrasena')
                
                if nueva_contrasena and contrasena_actual:
                    # Verificar contraseña actual
                    if check_password(contrasena_actual, empresa.contrasena):
                        # Actualizar con la nueva contraseña
                        empresa_actualizada.contrasena = make_password(nueva_contrasena)
                        empresa_actualizada.save()
                        messages.success(request, 'Perfil, datos bancarios y contraseña actualizados exitosamente.')
                    else:
                        messages.error(request, 'La contraseña actual es incorrecta.')
                        return render(request, 'perfil_empresa.html', {'form': form, 'empresa': empresa})
                else:
                    # Verificar el estado de los datos bancarios
                    if empresa_actualizada.datos_bancarios_completos():
                        if empresa_actualizada.datos_bancarios_verificados:
                            messages.success(request, 'Perfil actualizado. Tus datos bancarios están verificados y listos para recibir pagos.')
                        else:
                            messages.success(request, 'Perfil actualizado. Tus datos bancarios están completos y serán verificados pronto.')
                    else:
                        messages.success(request, 'Perfil actualizado.')
                        if any([empresa_actualizada.banco, empresa_actualizada.numero_cuenta, empresa_actualizada.titular_cuenta]):
                            messages.info(request, 'Completa todos los datos bancarios requeridos para poder recibir pagos.')
                
                # Actualizar datos en la sesión
                request.session['empresa_nombre'] = empresa_actualizada.nombre_empresa
                request.session['empresa_email'] = empresa_actualizada.email
                
                return redirect('perfil_empresa')
            else:
                messages.error(request, 'Por favor, corrija los errores en el formulario.')
        else:
            form = EmpresaPerfilForm(instance=empresa)
        
        # Obtener estadísticas adicionales para mostrar en el perfil
        total_reservas = Reserva.objects.filter(empresa=empresa).count()
        reservas_completadas = Reserva.objects.filter(empresa=empresa, estado='completado').count()
        servicios_ofrecidos = empresa.servicios.count()
        
        # Fecha de registro formateada
        fecha_registro = empresa.fecha_registro.strftime('%d de %B de %Y') if empresa.fecha_registro else 'No disponible'
        
        # ===== NUEVOS DATOS PARA GESTIÓN DE SERVICIOS =====
        # Servicios ya asignados a la empresa
        servicios_asignados = empresa.servicios.all()
        
        # Servicios disponibles que la empresa aún no tiene
        servicios_disponibles = Servicio.objects.exclude(id_servicio__in=servicios_asignados.values_list('id_servicio', flat=True))
        
        # Solicitudes pendientes de la empresa
        solicitudes_pendientes = SolicitudServicioEmpresa.objects.filter(
            empresa=empresa,
            estado__in=['pendiente', 'en_revision']
        ).order_by('-fecha_solicitud')
        
        context = {
            'form': form,
            'empresa': empresa,
            'total_reservas': total_reservas,
            'reservas_completadas': reservas_completadas,
            'servicios_ofrecidos': servicios_ofrecidos,
            'fecha_registro': fecha_registro,
            # Nuevos contextos para servicios
            'servicios_asignados': servicios_asignados,
            'servicios_disponibles': servicios_disponibles,
            'solicitudes_pendientes': solicitudes_pendientes,
        }

        return render(request, 'empresas/perfil_empresa.html', context)

    except Empresa.DoesNotExist:
        messages.error(request, 'Error: Empresa no encontrada.')
        return redirect('logincrud')
    except Exception as e:
        print(f"❌ Error en perfil_empresa: {str(e)}")
        messages.error(request, 'Error interno. Intente nuevamente.')
        return redirect('home_empresas')


@empresa_required
def gestion_servicios_empresa(request):
    """Vista para la página completa de gestión de servicios de empresa"""
    try:
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Servicios ya asignados a la empresa
        servicios_asignados = empresa.servicios.all().order_by('nombre_servicio')
        
        # Servicios disponibles que la empresa aún no tiene
        servicios_disponibles = Servicio.objects.exclude(
            id_servicio__in=servicios_asignados.values_list('id_servicio', flat=True)
        ).order_by('nombre_servicio')
        
        # Solicitudes pendientes de la empresa
        solicitudes_pendientes = SolicitudServicioEmpresa.objects.filter(
            empresa=empresa,
            estado__in=['pendiente', 'en_revision']
        ).select_related('servicio_solicitado').order_by('-fecha_solicitud')
        
        context = {
            'empresa': empresa,
            'servicios_asignados': servicios_asignados,
            'servicios_disponibles': servicios_disponibles,
            'solicitudes_pendientes': solicitudes_pendientes,
        }

        return render(request, 'empresas/gestion_servicios.html', context)

    except Empresa.DoesNotExist:
        messages.error(request, 'Error: Empresa no encontrada.')
        return redirect('logincrud')
    except Exception as e:
        print(f"❌ Error en gestion_servicios_empresa: {str(e)}")
        messages.error(request, 'Error interno. Intente nuevamente.')
        return redirect('home_empresas')


@empresa_required
def solicitar_servicio_empresa(request):
    """Vista para procesar solicitudes de servicios por parte de empresas"""
    print(f"🔍 solicitar_servicio_empresa llamada con método: {request.method}")
    print(f"📋 POST data: {dict(request.POST)}")
    print(f"👤 Empresa en sesión: {request.session.get('empresa_id')}")
    
    if request.method == 'POST':
        try:
            empresa_id = request.session.get('empresa_id')
            print(f"🏢 Empresa ID: {empresa_id}")
            empresa = Empresa.objects.get(id_empresa=empresa_id)
            print(f"✅ Empresa encontrada: {empresa.nombre_empresa}")
            
            servicio_id = request.POST.get('servicio_id')
            usuario_responsable = request.POST.get('usuario_responsable')
            telefono_contacto = request.POST.get('telefono_contacto')
            motivo_solicitud = request.POST.get('motivo_solicitud')
            
            print(f"📝 Datos recibidos:")
            print(f"   - Servicio ID: {servicio_id}")
            print(f"   - Usuario responsable: {usuario_responsable}")
            print(f"   - Teléfono: {telefono_contacto}")
            print(f"   - Motivo: {motivo_solicitud}")
            
            # Validaciones básicas
            if not all([servicio_id, usuario_responsable, telefono_contacto, motivo_solicitud]):
                print("❌ Faltan campos requeridos")
                messages.error(request, 'Todos los campos son requeridos.')
                return redirect('gestion_servicios_empresa')
            
            try:
                servicio = Servicio.objects.get(id_servicio=servicio_id)
                print(f"🛠️ Servicio encontrado: {servicio.nombre_servicio}")
            except Servicio.DoesNotExist:
                print("❌ Servicio no existe")
                messages.error(request, 'El servicio solicitado no existe.')
                return redirect('gestion_servicios_empresa')
            
            # Verificar si la empresa ya tiene este servicio
            if empresa.servicios.filter(id_servicio=servicio_id).exists():
                print("⚠️ Empresa ya tiene este servicio")
                messages.error(request, f'Tu empresa ya tiene acceso al servicio "{servicio.nombre_servicio}".')
                return redirect('gestion_servicios_empresa')
            
            # Verificar si ya existe una solicitud pendiente para este servicio
            solicitud_existente = SolicitudServicioEmpresa.objects.filter(
                empresa=empresa,
                servicio_solicitado=servicio,
                estado__in=['pendiente', 'en_revision']
            ).first()
            
            if solicitud_existente:
                print("⚠️ Ya existe solicitud pendiente")
                messages.warning(request, f'Ya tienes una solicitud pendiente para el servicio "{servicio.nombre_servicio}".')
                return redirect('gestion_servicios_empresa')
            
            # Crear la nueva solicitud
            nueva_solicitud = SolicitudServicioEmpresa.objects.create(
                empresa=empresa,
                servicio_solicitado=servicio,
                motivo_solicitud=motivo_solicitud,
                usuario_responsable=usuario_responsable,
                telefono_contacto=telefono_contacto,
                estado='pendiente'
            )
            
            print(f"✅ Nueva solicitud creada con ID: {nueva_solicitud.id_solicitud}")
            messages.success(request, f'Solicitud enviada exitosamente para el servicio "{servicio.nombre_servicio}". Te contactaremos pronto.')
            
            # Log para debugging
            print(f"✅ Nueva solicitud de servicio creada:")
            print(f"   - Empresa: {empresa.nombre_empresa}")
            print(f"   - Servicio: {servicio.nombre_servicio}")
            print(f"   - Responsable: {usuario_responsable}")
            print(f"   - ID Solicitud: {nueva_solicitud.id_solicitud}")
            
            return redirect('gestion_servicios_empresa')
            
        except Empresa.DoesNotExist:
            messages.error(request, 'Error: Empresa no encontrada.')
            return redirect('logincrud')
        except Exception as e:
            print(f"❌ Error en solicitar_servicio_empresa: {str(e)}")
            messages.error(request, 'Error interno al procesar la solicitud. Intente nuevamente.')
            return redirect('perfil_empresa')
    
    # Si no es POST, redirigir al perfil
    return redirect('perfil_empresa')


@empresa_required
def exportar_reporte_empresa(request):
    """Vista para exportar reportes de la empresa en formato CSV"""
    try:
        import csv
        from django.http import HttpResponse
        from datetime import datetime, timedelta
        import calendar
        
        empresa_id = request.session.get('empresa_id')
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        
        # Obtener el período seleccionado (por defecto 6 meses)
        periodo = int(request.GET.get('periodo', 6))
        formato = request.GET.get('formato', 'csv')  # csv o pdf
        
        # Calcular fechas
        hoy = timezone.now().date()
        fecha_inicio = hoy - timedelta(days=30 * periodo)
        
        print(f"🔍 Exportando reporte para empresa {empresa.nombre_empresa}")
        print(f"📅 Período: {periodo} meses (desde {fecha_inicio} hasta {hoy})")
        
        # Obtener datos de reservas
        reservas_empresa = Reserva.objects.filter(
            empresa=empresa,
            fecha__gte=fecha_inicio,
            fecha__lte=hoy
        ).order_by('-fecha')
        
        # Estadísticas generales
        total_reservas = reservas_empresa.count()
        reservas_completadas = reservas_empresa.filter(estado='completado').count()
        reservas_pendientes = reservas_empresa.filter(estado='pendiente').count()
        tasa_completado = (reservas_completadas / total_reservas * 100) if total_reservas > 0 else 0
        
        # Calcular ingresos
        ingresos_totales = 0
        for reserva in reservas_empresa.filter(estado='completado'):
            try:
                servicios = reserva.servicios.all()
                precio_reserva = sum(servicio.precio for servicio in servicios)
                ingresos_totales += precio_reserva
            except Exception as e:
                print(f"❌ Error calculando ingresos para reserva {reserva.id_reserva}: {e}")
                continue
        
        # Datos por mes
        datos_mensuales = []
        for i in range(periodo):
            fecha_mes = hoy - timedelta(days=30 * i)
            mes_actual = fecha_mes.month
            ano_actual = fecha_mes.year
            
            reservas_mes = reservas_empresa.filter(
                fecha__month=mes_actual,
                fecha__year=ano_actual
            )
            
            completadas_mes = reservas_mes.filter(estado='completado').count()
            ingresos_mes = 0
            
            for reserva in reservas_mes.filter(estado='completado'):
                try:
                    servicios = reserva.servicios.all()
                    ingresos_mes += sum(servicio.precio for servicio in servicios)
                except:
                    continue
            
            datos_mensuales.append({
                'mes': f"{calendar.month_name[mes_actual]} {ano_actual}",
                'total_reservas': reservas_mes.count(),
                'completadas': completadas_mes,
                'pendientes': reservas_mes.filter(estado='pendiente').count(),
                'ingresos': ingresos_mes,
                'tasa_completado': (completadas_mes / reservas_mes.count() * 100) if reservas_mes.count() > 0 else 0
            })
        
        # Servicios más populares
        servicios_populares = []
        servicios_stats = {}
        
        for reserva in reservas_empresa:
            for servicio in reserva.servicios.all():
                if servicio.nombre_servicio not in servicios_stats:
                    servicios_stats[servicio.nombre_servicio] = {
                        'count': 0,
                        'ingresos': 0,
                        'precio': servicio.precio
                    }
                servicios_stats[servicio.nombre_servicio]['count'] += 1
                if reserva.estado == 'completado':
                    servicios_stats[servicio.nombre_servicio]['ingresos'] += servicio.precio
        
        # Ordenar servicios por popularidad
        servicios_populares = sorted(
            [{'servicio': k, **v} for k, v in servicios_stats.items()],
            key=lambda x: x['count'],
            reverse=True
        )[:5]  # Top 5
        
        if formato == 'csv':
            # Crear respuesta CSV
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="reporte_empresa_{empresa.nombre_empresa}_{periodo}meses_{hoy.strftime("%Y%m%d")}.csv"'
            
            # Agregar BOM para UTF-8
            response.write('\ufeff')
            
            writer = csv.writer(response)
            
            # Encabezado del reporte
            writer.writerow(['REPORTE DE ANÁLISIS DE NEGOCIO - AUTONEW'])
            writer.writerow([''])
            writer.writerow(['Información de la Empresa'])
            writer.writerow(['Empresa:', empresa.nombre_empresa])
            writer.writerow(['Email:', empresa.email])
            writer.writerow(['Teléfono:', empresa.telefono])
            writer.writerow(['Dirección:', empresa.direccion])
            writer.writerow(['Fecha de Reporte:', hoy.strftime('%d/%m/%Y')])
            writer.writerow(['Período Analizado:', f'Últimos {periodo} meses'])
            writer.writerow(['Fecha Inicio:', fecha_inicio.strftime('%d/%m/%Y')])
            writer.writerow(['Fecha Fin:', hoy.strftime('%d/%m/%Y')])
            writer.writerow([''])
            
            # Métricas principales
            writer.writerow(['MÉTRICAS PRINCIPALES'])
            writer.writerow(['Métrica', 'Valor'])
            writer.writerow(['Total de Reservas', total_reservas])
            writer.writerow(['Reservas Completadas', reservas_completadas])
            writer.writerow(['Reservas Pendientes', reservas_pendientes])
            writer.writerow(['Tasa de Completado (%)', f"{tasa_completado:.1f}%"])
            writer.writerow(['Ingresos Totales', f"${ingresos_totales:,.2f}"])
            writer.writerow(['Promedio por Reserva', f"${(ingresos_totales/total_reservas):,.2f}" if total_reservas > 0 else "$0.00"])
            writer.writerow([''])
            
            # Análisis mensual
            writer.writerow(['ANÁLISIS MENSUAL DETALLADO'])
            writer.writerow(['Mes/Año', 'Total Reservas', 'Completadas', 'Pendientes', 'Tasa Completado (%)', 'Ingresos'])
            
            for dato in reversed(datos_mensuales):  # Mostrar del más antiguo al más reciente
                writer.writerow([
                    dato['mes'],
                    dato['total_reservas'],
                    dato['completadas'],
                    dato['pendientes'],
                    f"{dato['tasa_completado']:.1f}%",
                    f"${dato['ingresos']:,.2f}"
                ])
            
            writer.writerow([''])
            
            # Servicios más populares
            writer.writerow(['SERVICIOS MÁS POPULARES'])
            writer.writerow(['Servicio', 'Total Reservas', 'Ingresos Generados', 'Precio Unitario'])
            
            for servicio in servicios_populares:
                writer.writerow([
                    servicio['servicio'],
                    servicio['count'],
                    f"${servicio['ingresos']:,.2f}",
                    f"${servicio['precio']:,.2f}"
                ])
            
            writer.writerow([''])
            
            # Detalle de reservas
            writer.writerow(['DETALLE DE RESERVAS'])
            writer.writerow(['Fecha', 'Cliente', 'Servicios', 'Estado', 'Total'])
            
            for reserva in reservas_empresa.order_by('-fecha')[:50]:  # Últimas 50 reservas
                servicios_nombres = ', '.join([s.nombre_servicio for s in reserva.servicios.all()])
                total_reserva = sum(s.precio for s in reserva.servicios.all()) if reserva.estado == 'completado' else 0
                
                writer.writerow([
                    reserva.fecha.strftime('%d/%m/%Y'),
                    reserva.usuario.nombre_completo if reserva.usuario else 'N/A',
                    servicios_nombres or 'Sin servicios',
                    'Completada' if reserva.estado == 'completado' else 'Pendiente',
                    f"${total_reserva:,.2f}"
                ])
            
            writer.writerow([''])
            writer.writerow(['Reporte generado por AUTONEW - Sistema de Gestión de Lavado de Autos'])
            writer.writerow([f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")}'])
            
            print(f"✅ Reporte CSV generado exitosamente para {empresa.nombre_empresa}")
            return response
            
        else:
            # Formato no soportado por ahora
            messages.error(request, 'Formato de reporte no soportado.')
            return redirect('reportes_empresa')
            
    except Empresa.DoesNotExist:
        messages.error(request, 'Error: Empresa no encontrada.')
        return redirect('logincrud')
    except Exception as e:
        print(f"❌ Error generando reporte: {str(e)}")
        messages.error(request, 'Error generando el reporte. Intente nuevamente.')
        return redirect('reportes_empresa')


# ================================
# VISTAS CRUD PARA PLANES EMPRESARIALES
# ================================

@admin_required
def planes_empresariales_crud(request):
    """Vista CRUD para gestionar planes empresariales (solo admin)"""
    planes = PlanEmpresarial.objects.all().order_by('precio_mensual_por_vehiculo')
    
    # Obtener las solicitudes de contacto de planes empresariales más recientes
    solicitudes_contacto = SolicitudContactoPlan.objects.select_related('plan').order_by('-fecha_solicitud')[:10]
    
    context = {
        'planes': planes,
        'solicitudes_contacto': solicitudes_contacto,
    }
    return render(request, 'planes/planes_empresariales_crud.html', context)

@admin_required
def crear_plan_empresarial(request):
    """Vista para crear un nuevo plan empresarial"""
    if request.method == 'POST':
        try:
            plan = PlanEmpresarial.objects.create(
                nombre=request.POST.get('nombre'),
                tipo=request.POST.get('tipo'),
                descripcion=request.POST.get('descripcion'),
                precio_mensual_por_vehiculo=request.POST.get('precio_mensual_por_vehiculo'),
                precio_base_mensual=request.POST.get('precio_base_mensual', 0),
                vehiculos_minimos=request.POST.get('vehiculos_minimos', 5),
                vehiculos_maximos=request.POST.get('vehiculos_maximos') or None,
                servicios_por_vehiculo_mes=request.POST.get('servicios_por_vehiculo_mes', 0),
                descuento_volumen=request.POST.get('descuento_volumen', 0),
                incluye_lavado_asientos=request.POST.get('incluye_lavado_asientos') == 'on',
                incluye_aspirado=request.POST.get('incluye_aspirado') == 'on',
                incluye_lavado_exterior=request.POST.get('incluye_lavado_exterior') == 'on',
                incluye_lavado_interior_humedo=request.POST.get('incluye_lavado_interior_humedo') == 'on',
                incluye_encerado=request.POST.get('incluye_encerado') == 'on',
                incluye_detallado_completo=request.POST.get('incluye_detallado_completo') == 'on',
                incluye_servicio_domicilio=request.POST.get('incluye_servicio_domicilio') == 'on',
                incluye_mantenimiento_programado=request.POST.get('incluye_mantenimiento_programado') == 'on',
                incluye_reporte_mensual=request.POST.get('incluye_reporte_mensual') == 'on',
                incluye_soporte_24_7=request.POST.get('incluye_soporte_24_7') == 'on',
            )
            
            # Agregar servicios incluidos
            servicios_ids = request.POST.getlist('servicios_incluidos')
            for servicio_id in servicios_ids:
                servicio = Servicio.objects.get(id_servicio=servicio_id)
                plan.servicios_incluidos.add(servicio)
            
            messages.success(request, f'Plan empresarial "{plan.nombre}" creado exitosamente.')
            return redirect('planes_empresariales_crud')
            
        except Exception as e:
            messages.error(request, f'Error al crear el plan empresarial: {str(e)}')
    
    servicios = Servicio.objects.all()
    context = {
        'servicios': servicios,
    }
    return render(request, 'planes/planes_empresariales/crear_plan_empresarial.html', context)

@admin_required
def editar_plan_empresarial(request, plan_id):
    """Vista para editar un plan empresarial existente"""
    plan = get_object_or_404(PlanEmpresarial, id_plan=plan_id)
    
    if request.method == 'POST':
        try:
            plan.nombre = request.POST.get('nombre')
            plan.tipo = request.POST.get('tipo')
            plan.descripcion = request.POST.get('descripcion')
            plan.precio_mensual_por_vehiculo = request.POST.get('precio_mensual_por_vehiculo')
            plan.precio_base_mensual = request.POST.get('precio_base_mensual', 0)
            plan.vehiculos_minimos = request.POST.get('vehiculos_minimos', 5)
            plan.vehiculos_maximos = request.POST.get('vehiculos_maximos') or None
            plan.servicios_por_vehiculo_mes = request.POST.get('servicios_por_vehiculo_mes', 0)
            plan.descuento_volumen = request.POST.get('descuento_volumen', 0)
            plan.incluye_lavado_asientos = request.POST.get('incluye_lavado_asientos') == 'on'
            plan.incluye_aspirado = request.POST.get('incluye_aspirado') == 'on'
            plan.incluye_lavado_exterior = request.POST.get('incluye_lavado_exterior') == 'on'
            plan.incluye_lavado_interior_humedo = request.POST.get('incluye_lavado_interior_humedo') == 'on'
            plan.incluye_encerado = request.POST.get('incluye_encerado') == 'on'
            plan.incluye_detallado_completo = request.POST.get('incluye_detallado_completo') == 'on'
            plan.incluye_servicio_domicilio = request.POST.get('incluye_servicio_domicilio') == 'on'
            plan.incluye_mantenimiento_programado = request.POST.get('incluye_mantenimiento_programado') == 'on'
            plan.incluye_reporte_mensual = request.POST.get('incluye_reporte_mensual') == 'on'
            plan.incluye_soporte_24_7 = request.POST.get('incluye_soporte_24_7') == 'on'
            plan.save()
            
            # Actualizar servicios incluidos
            plan.servicios_incluidos.clear()
            servicios_ids = request.POST.getlist('servicios_incluidos')
            for servicio_id in servicios_ids:
                servicio = Servicio.objects.get(id_servicio=servicio_id)
                plan.servicios_incluidos.add(servicio)
            
            messages.success(request, f'Plan empresarial "{plan.nombre}" actualizado exitosamente.')
            return redirect('planes_empresariales_crud')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el plan empresarial: {str(e)}')
    
    servicios = Servicio.objects.all()
    context = {
        'plan': plan,
        'servicios': servicios,
    }
    return render(request, 'planes/planes_empresariales/editar_plan_empresarial.html', context)

@admin_required
def eliminar_plan_empresarial(request, plan_id):
    """Vista para eliminar un plan empresarial"""
    plan = get_object_or_404(PlanEmpresarial, id_plan=plan_id)
    
    if request.method == 'POST':
        try:
            # Verificar si hay suscripciones activas
            suscripciones_activas = SuscripcionEmpresarial.objects.filter(
                plan=plan, 
                estado='activa'
            ).count()
            
            if suscripciones_activas > 0:
                messages.warning(request, f'No se puede eliminar el plan "{plan.nombre}" porque tiene {suscripciones_activas} suscripciones activas. Desactivando el plan.')
                plan.activo = False
                plan.save()
            else:
                plan_nombre = plan.nombre
                plan.delete()
                messages.success(request, f'Plan empresarial "{plan_nombre}" eliminado exitosamente.')

            return redirect('planes/planes_empresariales/planes_empresariales_crud')

        except Exception as e:
            messages.error(request, f'Error al eliminar el plan empresarial: {str(e)}')
    
    # Obtener información de suscripciones para mostrar en la confirmación
    suscripciones_activas = SuscripcionEmpresarial.objects.filter(
        plan=plan, 
        estado='activa'
    ).count()
    
    context = {
        'plan': plan,
        'suscripciones_activas': suscripciones_activas,
    }
    return render(request, 'eliminar_plan_empresarial.html', context)

@admin_required
def detalle_plan_empresarial(request, plan_id):
    """Vista para ver detalles del plan empresarial"""
    plan = get_object_or_404(PlanEmpresarial, id_plan=plan_id)
    
    # Obtener suscripciones relacionadas
    suscripciones = SuscripcionEmpresarial.objects.filter(plan=plan).order_by('-fecha_inicio')[:10]
    
    # Calcular estadísticas
    total_suscripciones = SuscripcionEmpresarial.objects.filter(plan=plan).count()
    suscripciones_activas = SuscripcionEmpresarial.objects.filter(plan=plan, estado='activa').count()
    
    context = {
        'plan': plan,
        'suscripciones': suscripciones,
        'total_suscripciones': total_suscripciones,
        'suscripciones_activas': suscripciones_activas,
    }
    return render(request, 'planes/planes_empresariales/detalle_plan_empresarial.html', context)

@admin_required
def suscripciones_empresariales_crud(request):
    """Vista para gestionar suscripciones empresariales (solo admin)"""
    # Obtener todas las suscripciones para estadísticas
    todas_suscripciones = SuscripcionEmpresarial.objects.select_related('empresa', 'plan').all()
    
    # Calcular estadísticas
    suscripciones_activas = todas_suscripciones.filter(estado='activa').count()
    suscripciones_vencidas = todas_suscripciones.filter(estado='vencida').count()
    total_vehiculos = sum(suscripcion.cantidad_vehiculos for suscripcion in todas_suscripciones)
    
    # Calcular ingresos mensuales
    ingresos_mensuales = sum(
        suscripcion.precio_mensual_actual 
        for suscripcion in todas_suscripciones.filter(estado='activa')
    )
    
    # Obtener planes disponibles para el filtro
    planes_disponibles = PlanEmpresarial.objects.filter(activo=True)
    
    # Aplicar filtros de búsqueda si existen
    suscripciones_filtradas = todas_suscripciones.order_by('-fecha_inicio')
    
    # Filtro por estado
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        suscripciones_filtradas = suscripciones_filtradas.filter(estado=estado_filtro)
    
    # Filtro por plan
    plan_filtro = request.GET.get('plan')
    if plan_filtro:
        suscripciones_filtradas = suscripciones_filtradas.filter(plan__id_plan=plan_filtro)
    
    # Filtro por búsqueda de texto
    busqueda = request.GET.get('busqueda')
    if busqueda:
        suscripciones_filtradas = suscripciones_filtradas.filter(
            Q(empresa__nombre_empresa__icontains=busqueda) |
            Q(contacto_responsable__icontains=busqueda) |
            Q(plan__nombre__icontains=busqueda)
        )
    
    # Implementar paginación
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(suscripciones_filtradas, 25)  # 25 suscripciones por página
    page = request.GET.get('page')
    
    try:
        suscripciones = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, mostrar la primera página
        suscripciones = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera del rango, mostrar la última página
        suscripciones = paginator.page(paginator.num_pages)
    
    context = {
        'suscripciones': suscripciones,
        'paginator': paginator,
        'suscripciones_activas': suscripciones_activas,
        'suscripciones_vencidas': suscripciones_vencidas,
        'total_vehiculos': total_vehiculos,
        'ingresos_mensuales': ingresos_mensuales,
        'planes_disponibles': planes_disponibles,
        'total_suscripciones': todas_suscripciones.count(),
        'filtros_activos': {
            'estado': estado_filtro,
            'plan': plan_filtro,
            'busqueda': busqueda,
        }
    }
    return render(request, 'planes/suscripciones_empresariales_crud.html', context)


# ==================== VISTAS PARA SUSCRIPCIONES INDIVIDUALES ====================

@admin_required
def suscripciones_individuales_crud(request):
    """Vista para gestionar suscripciones de usuarios individuales (solo admin)"""
    # Obtener todas las suscripciones para estadísticas
    todas_suscripciones = SuscripcionUsuario.objects.select_related('usuario', 'plan').all()
    
    # Calcular estadísticas
    suscripciones_activas = todas_suscripciones.filter(estado='activa').count()
    suscripciones_vencidas = todas_suscripciones.filter(estado='vencida').count()
    
    # Calcular ingresos mensuales estimados
    ingresos_mensuales = sum(
        suscripcion.plan.precio_mensual 
        for suscripcion in todas_suscripciones.filter(estado='activa')
    )
    
    # Obtener planes disponibles para el filtro
    planes_disponibles = Plan.objects.filter(activo=True)
    
    # Aplicar filtros de búsqueda si existen
    suscripciones_filtradas = todas_suscripciones.order_by('-fecha_inicio')
    
    # Filtro por estado
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        suscripciones_filtradas = suscripciones_filtradas.filter(estado=estado_filtro)
    
    # Filtro por plan
    plan_filtro = request.GET.get('plan')
    if plan_filtro:
        suscripciones_filtradas = suscripciones_filtradas.filter(plan__id_plan=plan_filtro)
    
    # Filtro por búsqueda de texto
    busqueda = request.GET.get('busqueda')
    if busqueda:
        suscripciones_filtradas = suscripciones_filtradas.filter(
            Q(usuario__nombre_completo__icontains=busqueda) |
            Q(usuario__correo__icontains=busqueda) |
            Q(plan__nombre__icontains=busqueda)
        )
    
    # Implementar paginación
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(suscripciones_filtradas, 25)  # 25 suscripciones por página
    page = request.GET.get('page')
    
    try:
        suscripciones = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, mostrar la primera página
        suscripciones = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera del rango, mostrar la última página
        suscripciones = paginator.page(paginator.num_pages)
    
    context = {
        'suscripciones': suscripciones,
        'paginator': paginator,
        'suscripciones_activas': suscripciones_activas,
        'suscripciones_vencidas': suscripciones_vencidas,
        'ingresos_mensuales': ingresos_mensuales,
        'planes_disponibles': planes_disponibles,
        'total_suscripciones': todas_suscripciones.count(),
        'filtros_activos': {
            'estado': estado_filtro,
            'plan': plan_filtro,
            'busqueda': busqueda,
        }
    }
    return render(request, 'planes/suscripciones_individuales_crud.html', context)

@admin_required
def crear_suscripcion_individual(request):
    """Vista para crear una nueva suscripción individual"""
    if request.method == 'POST':
        usuario_id = request.POST.get('usuario')
        plan_id = request.POST.get('plan')
        
        try:
            usuario = Usuario.objects.get(id_usuario=usuario_id)
            plan = Plan.objects.get(id_plan=plan_id)
            
            # Verificar si el usuario ya tiene una suscripción activa
            suscripcion_existente = SuscripcionUsuario.objects.filter(
                usuario=usuario, 
                estado='activa'
            ).first()
            
            if suscripcion_existente:
                messages.error(request, f'El usuario {usuario.nombre_completo} ya tiene una suscripción activa.')
                return redirect('suscripciones_individuales_crud')
            
            # Crear nueva suscripción
            fecha_inicio = timezone.now()
            fecha_fin = fecha_inicio + timedelta(days=30)  # 30 días por defecto
            
            suscripcion = SuscripcionUsuario.objects.create(
                usuario=usuario,
                plan=plan,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                estado='activa',
                servicios_utilizados_mes=0,
                ultimo_reinicio_contador=fecha_inicio,
                auto_renovar=True
            )
            
            messages.success(request, f'Suscripción creada exitosamente para {usuario.nombre_completo}.')
            return redirect('suscripciones_individuales_crud')
            
        except (Usuario.DoesNotExist, Plan.DoesNotExist):
            messages.error(request, 'Usuario o plan no encontrado.')
            return redirect('suscripciones_individuales_crud')
    
    # GET request - mostrar formulario
    usuarios = Usuario.objects.filter(rol='cliente').order_by('nombre_completo')
    planes = Plan.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'usuarios': usuarios,
        'planes': planes,
    }
    return render(request, 'planes/suscripcion_individual/crear_suscripcion_individual.html', context)

@admin_required
def editar_suscripcion_individual(request, suscripcion_id):
    """Vista para editar una suscripción individual"""
    try:
        suscripcion = SuscripcionUsuario.objects.get(id_suscripcion=suscripcion_id)
    except SuscripcionUsuario.DoesNotExist:
        messages.error(request, 'Suscripción no encontrada.')
        return redirect('suscripciones_individuales_crud')
    
    if request.method == 'POST':
        estado = request.POST.get('estado')
        auto_renovar = request.POST.get('auto_renovar') == 'on'
        fecha_fin = request.POST.get('fecha_fin')
        
        # Actualizar suscripción
        suscripcion.estado = estado
        suscripcion.auto_renovar = auto_renovar
        
        if fecha_fin:
            try:
                from datetime import datetime
                suscripcion.fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Formato de fecha inválido.')
                return redirect('editar_suscripcion_individual', suscripcion_id=suscripcion_id)
        
        suscripcion.save()
        
        messages.success(request, f'Suscripción de {suscripcion.usuario.nombre_completo} actualizada exitosamente.')
        return redirect('suscripciones_individuales_crud')
    
    context = {
        'suscripcion': suscripcion,
    }
    return render(request, 'planes/suscripcion_individual/editar_suscripcion_individual.html', context)

@admin_required
def eliminar_suscripcion_individual(request, suscripcion_id):
    """Vista para eliminar una suscripción individual"""
    try:
        suscripcion = SuscripcionUsuario.objects.get(id_suscripcion=suscripcion_id)
    except SuscripcionUsuario.DoesNotExist:
        messages.error(request, 'Suscripción no encontrada.')
        return redirect('suscripciones_individuales_crud')
    
    if request.method == 'POST':
        usuario_nombre = suscripcion.usuario.nombre_completo
        suscripcion.delete()
        messages.success(request, f'Suscripción de {usuario_nombre} eliminada exitosamente.')
        return redirect('suscripciones_individuales_crud')
    
    context = {
        'suscripcion': suscripcion,
    }
    return render(request, 'planes/suscripcion_individual/eliminar_suscripcion_individual.html', context)

@admin_required
def pausar_suscripcion_individual(request, suscripcion_id):
    """Vista para pausar una suscripción individual"""
    try:
        suscripcion = SuscripcionUsuario.objects.get(id_suscripcion=suscripcion_id)
    except SuscripcionUsuario.DoesNotExist:
        messages.error(request, 'Suscripción no encontrada.')
        return redirect('suscripciones_individuales_crud')
    
    if suscripcion.estado == 'activa':
        suscripcion.estado = 'pausada'
        suscripcion.save()
        messages.success(request, f'Suscripción de {suscripcion.usuario.nombre_completo} pausada exitosamente.')
    else:
        messages.warning(request, 'Solo se pueden pausar suscripciones activas.')
    
    return redirect('suscripciones_individuales_crud')

@admin_required
def historial_pagos_suscripcion(request, suscripcion_id):
    """Vista para ver el historial de pagos de una suscripción"""
    try:
        suscripcion = SuscripcionUsuario.objects.get(id_suscripcion=suscripcion_id)
    except SuscripcionUsuario.DoesNotExist:
        messages.error(request, 'Suscripción no encontrada.')
        return redirect('suscripciones_individuales_crud')
    
    pagos = HistorialPagosSuscripcion.objects.filter(suscripcion=suscripcion).order_by('-fecha_pago')
    
    context = {
        'suscripcion': suscripcion,
        'pagos': pagos,
    }
    return render(request, 'planes/suscripcion_individual/historial_pagos_suscripcion.html', context)


@admin_required
def perfil_admin(request):
    """Vista para editar el perfil del administrador"""
    admin_user = request.user
    
    if request.method == 'POST':
        form = AdminProfileForm(request.POST, request.FILES, instance=admin_user)
        
        if form.is_valid():
            # Guardar los datos del formulario
            admin_actualizado = form.save()
            
            # Cambiar la contraseña si se proporcionó
            nueva_contrasena = form.cleaned_data.get('nueva_contrasena')
            if nueva_contrasena:
                admin_actualizado.set_password(nueva_contrasena)
                admin_actualizado.save()
                messages.success(request, 'Perfil y contraseña actualizados correctamente.')
            else:
                messages.success(request, 'Perfil actualizado correctamente.')
            
            return redirect('perfil_admin')
        else:
            # Si hay errores en el formulario, mostrarlos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = AdminProfileForm(instance=admin_user)
    
    context = {
        'form': form,
        'admin_user': admin_user,
    }
    return render(request, 'usuarios/perfil_admin.html', context)


# NUEVAS VISTAS AJAX PARA CRUD DE CITAS
@admin_required
def obtener_servicios_empresa(request):
    """Vista AJAX para obtener los servicios que presta una empresa específica"""
    print(f"🔧 obtener_servicios_empresa llamada")
    empresa_id = request.GET.get('empresa_id')
    print(f"🔧 empresa_id recibido: {empresa_id}")
    
    if not empresa_id:
        print(f"❌ No se proporcionó empresa_id")
        return JsonResponse({'error': 'ID de empresa requerido'}, status=400)
    
    try:
        empresa = Empresa.objects.get(id_empresa=empresa_id)
        print(f"🔧 Empresa encontrada: {empresa.nombre_empresa}")
        
        # Método 1: Intentar obtener servicios a través del modelo intermedio EmpresaServicio
        servicios_empresa = EmpresaServicio.objects.filter(empresa=empresa).select_related('servicio')
        print(f"🔧 Servicios encontrados via EmpresaServicio: {servicios_empresa.count()}")
        
        servicios = []
        
        # Si hay servicios asignados a través de EmpresaServicio
        if servicios_empresa.exists():
            for empresa_servicio in servicios_empresa:
                servicios.append({
                    'id_servicio': empresa_servicio.servicio.id_servicio,
                    'nombre_servicio': empresa_servicio.servicio.nombre_servicio,
                    'precio': empresa_servicio.servicio.precio,
                    'descripcion': empresa_servicio.servicio.descripcion
                })
        else:
            # Método 2: Si no hay registros en EmpresaServicio, usar la relación ManyToMany directamente
            print(f"🔧 No hay servicios en EmpresaServicio, usando relación directa many-to-many")
            servicios_directos = empresa.servicios.all()
            print(f"🔧 Servicios encontrados via many-to-many: {servicios_directos.count()}")
            
            for servicio in servicios_directos:
                servicios.append({
                    'id_servicio': servicio.id_servicio,
                    'nombre_servicio': servicio.nombre_servicio,
                    'precio': servicio.precio,
                    'descripcion': servicio.descripcion
                })
        
        # Método 3: Si aún no hay servicios, devolver todos los servicios disponibles como fallback
        if not servicios:
            print(f"🔧 No hay servicios asignados, devolviendo todos los servicios disponibles")
            todos_servicios = Servicio.objects.all()
            for servicio in todos_servicios:
                servicios.append({
                    'id_servicio': servicio.id_servicio,
                    'nombre_servicio': servicio.nombre_servicio,
                    'precio': servicio.precio,
                    'descripcion': servicio.descripcion
                })
        
        print(f"🔧 Servicios procesados: {len(servicios)}")
        resultado = {
            'servicios': servicios,
            'empresa_nombre': empresa.nombre_empresa
        }
        print(f"🔧 Resultado final: {resultado}")
        
        return JsonResponse(resultado)
    except Empresa.DoesNotExist:
        print(f"❌ Empresa no encontrada con ID: {empresa_id}")
        return JsonResponse({'error': 'Empresa no encontrada'}, status=404)
    except Exception as e:
        print(f"❌ Error en obtener_servicios_empresa: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)


@admin_required  
def obtener_horas_disponibles(request):
    """Vista AJAX para obtener las horas disponibles para una fecha y empresa específica"""
    fecha_str = request.GET.get('fecha')
    empresa_id = request.GET.get('empresa_id')
    
    if not fecha_str:
        return JsonResponse({'error': 'Fecha requerida'}, status=400)
    
    try:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hoy = datetime.now().date()
        ahora = datetime.now()
        
        # Obtener reservas existentes para esa fecha
        reservas_fecha = Reserva.objects.filter(fecha=fecha)
        
        # Si se especifica empresa, filtrar por empresa
        if empresa_id:
            reservas_fecha = reservas_fecha.filter(empresa_id=empresa_id)
        
        # Obtener horas ocupadas
        horas_ocupadas = set()
        for reserva in reservas_fecha:
            horas_ocupadas.add(reserva.hora.strftime('%H:%M'))
        
        # Generar horas disponibles
        horas_disponibles = []
        
        for h in range(8, 16):  # De 08:00 a 15:00
            hora_formateada = f"{h:02d}:00"
            
            # Si la fecha es hoy, verificar que la hora no haya pasado
            if fecha == hoy and h <= ahora.hour:
                continue
            
            # Verificar si la hora no está ocupada
            if hora_formateada not in horas_ocupadas:
                # Convertir a formato 12h para mostrar
                hora_12h = convertir_hora_12h(hora_formateada)
                horas_disponibles.append({
                    'valor': hora_12h,
                    'texto': hora_12h
                })
        
        return JsonResponse({
            'horas_disponibles': horas_disponibles,
            'fecha': fecha_str
        })
        
    except ValueError:
        return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


##### suscripciones empresariales #####

@admin_required
def detalle_suscripcion_empresarial(request, suscripcion_id):
    """Vista para ver detalles de una suscripción empresarial"""
    suscripcion = get_object_or_404(
        SuscripcionEmpresarial.objects.select_related('empresa', 'plan').prefetch_related('plan__servicios_incluidos'), 
        id_suscripcion=suscripcion_id
    )
    
    # Obtener historial de pagos
    pagos = HistorialPagosSuscripcionEmpresarial.objects.filter(suscripcion=suscripcion).order_by('-fecha_pago')[:10]
    
    # Obtener todos los servicios disponibles para referencia
    todos_servicios = Servicio.objects.all()
    servicios_plan = suscripcion.plan.servicios_incluidos.all()
    
    context = {
        'suscripcion': suscripcion,
        'pagos': pagos,
        'todos_servicios': todos_servicios,
        'servicios_plan': servicios_plan,
        'servicios_count': servicios_plan.count(),
    }
    return render(request, 'planes/suscripcion_empresarial/detalle_suscripcion_empresarial.html', context)

@admin_required
def editar_suscripcion_empresarial(request, suscripcion_id):
    """Vista para editar una suscripción empresarial"""
    suscripcion = get_object_or_404(SuscripcionEmpresarial, id_suscripcion=suscripcion_id)
    
    if request.method == 'POST':
        # Actualizar campos de la suscripción
        suscripcion.estado = request.POST.get('estado')
        suscripcion.cantidad_vehiculos = request.POST.get('cantidad_vehiculos', 1)
        
        # Calcular nuevo precio si cambió la cantidad de vehículos
        if suscripcion.plan:
            suscripcion.precio_mensual_actual = suscripcion.plan.precio_por_vehiculo * int(suscripcion.cantidad_vehiculos)
        
        suscripcion.save()
        messages.success(request, f'Suscripción actualizada exitosamente.')
        return redirect('suscripciones_empresariales_crud')
    
    context = {
        'suscripcion': suscripcion,
    }
    return render(request, 'planes/suscripcion_empresarial/editar_suscripcion_empresarial.html', context)

@admin_required
@require_http_methods(["POST"])
def aprobar_solicitud_empresarial(request):
    """Vista para aprobar o rechazar solicitudes de contacto de planes empresariales"""
    try:
        import json
        from django.http import JsonResponse
        from django.utils import timezone
        
        # Parsear el cuerpo JSON de la petición
        data = json.loads(request.body)
        solicitud_id = data.get('solicitud_id')
        accion = data.get('accion')  # 'aprobar' o 'rechazar'
        comentario = data.get('comentario', '')
        
        # Validar datos
        if not solicitud_id or not accion:
            return JsonResponse({
                'success': False, 
                'message': 'Datos incompletos'
            })
        
        if accion not in ['aprobar', 'rechazar']:
            return JsonResponse({
                'success': False, 
                'message': 'Acción no válida'
            })
        
        # Obtener la solicitud
        try:
            solicitud = SolicitudContactoPlan.objects.get(id_solicitud=solicitud_id)
        except SolicitudContactoPlan.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Solicitud no encontrada'
            })
        
        # Verificar que la solicitud esté pendiente
        if solicitud.estado != 'pendiente':
            return JsonResponse({
                'success': False, 
                'message': f'La solicitud ya ha sido {solicitud.estado}'
            })
        
        # Procesar según la acción
        if accion == 'aprobar':
            solicitud.estado = 'aprobada'
            solicitud.fecha_contacto = timezone.now()
            solicitud.notas_seguimiento = f"Solicitud aprobada por {request.user.username} el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            
            # Aquí puedes agregar lógica adicional como:
            # - Crear usuario empresarial
            # - Enviar email de confirmación
            # - Crear suscripción temporal
            
            mensaje = f'Solicitud de {solicitud.nombre_completo} aprobada exitosamente'
            
        else:  # rechazar
            solicitud.estado = 'rechazada'
            solicitud.fecha_contacto = timezone.now()
            motivo = f" - Motivo: {comentario}" if comentario.strip() else ""
            solicitud.notas_seguimiento = f"Solicitud rechazada por {request.user.username} el {timezone.now().strftime('%d/%m/%Y %H:%M')}{motivo}"
            
            mensaje = f'Solicitud de {solicitud.nombre_completo} rechazada'
        
        # Guardar cambios
        solicitud.save()
        
        # Log para auditoría
        import logging
        logger = logging.getLogger('django')
        logger.info(f"Solicitud {solicitud_id} {accion} por {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'nuevo_estado': solicitud.estado
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Error en el formato de datos'
        })
    except Exception as e:
        import logging
        logger = logging.getLogger('django')
        logger.error(f"Error procesando solicitud {solicitud_id}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor'
        })


@login_required
def analisis_reservas_empresas(request):
    """
    Vista para mostrar el análisis de reservas por empresa.
    Separa servicios con plan y sin plan, agrupa por servicio y calcula totales.
    Incluye filtrado por mes y año con paginación.
    """
    from django.db.models import Count, Sum, F, Case, When, DecimalField, Value
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from decimal import Decimal
    from datetime import datetime, date
    import calendar
    
    # Obtener filtros de mes y año
    mes_filtro = request.GET.get('mes')
    ano_filtro = request.GET.get('ano')
    
    # Si no se especifica año, usar el año actual
    if not ano_filtro:
        ano_filtro = datetime.now().year
    else:
        ano_filtro = int(ano_filtro)
    
    # Obtener todas las empresas con sus reservas
    empresas_data = []
    datos_mensuales = {}
    
    # Generar datos para todos los meses si no hay filtro específico
    if not mes_filtro:
        meses_a_procesar = range(1, 13)
    else:
        meses_a_procesar = [int(mes_filtro)]
    
    for empresa in Empresa.objects.all().order_by('nombre_empresa'):
        empresa_datos = {
            'empresa': empresa,
            'datos_por_mes': {},
            'total_ano': {
                'reservas': 0,
                'servicios_sin_plan': 0,
                'servicios_con_plan': 0,
                'valor_sin_plan': Decimal('0.00'),
                'valor_con_plan': Decimal('0.00'),
                'valor_total': Decimal('0.00')
            }
        }
        
        for mes_num in meses_a_procesar:
            # Filtrar reservas por mes y año
            reservas_mes = Reserva.objects.filter(
                empresa=empresa,
                fecha__year=ano_filtro,
                fecha__month=mes_num
            )
            
            if not reservas_mes.exists():
                continue
            
            # Obtener servicios de este mes
            reservas_servicios_mes = ReservaServicio.objects.filter(
                reserva__empresa=empresa,
                reserva__fecha__year=ano_filtro,
                reserva__fecha__month=mes_num
            )
            
            # Separar servicios con plan y sin plan
            servicios_con_plan = reservas_servicios_mes.filter(es_servicio_plan=True)
            servicios_sin_plan = reservas_servicios_mes.filter(es_servicio_plan=False)
            
            # Agrupar servicios sin plan por nombre
            servicios_sin_plan_agrupados = {}
            for rs in servicios_sin_plan:
                nombre_servicio = rs.servicio.nombre_servicio
                if nombre_servicio not in servicios_sin_plan_agrupados:
                    servicios_sin_plan_agrupados[nombre_servicio] = {
                        'nombre': nombre_servicio,
                        'cantidad': 0,
                        'total_valor': Decimal('0.00')
                    }
                servicios_sin_plan_agrupados[nombre_servicio]['cantidad'] += 1
                if rs.precio_aplicado:
                    servicios_sin_plan_agrupados[nombre_servicio]['total_valor'] += Decimal(str(rs.precio_aplicado))
            
            # Agrupar servicios con plan por nombre y plan específico
            servicios_con_plan_detallados = []
            for rs in servicios_con_plan:
                # Obtener información del plan de forma más detallada
                plan_info = "Sin Plan Específico"  # Por defecto
                
                try:
                    # Verificar primero si tiene suscripción individual
                    if rs.reserva.suscripcion_utilizada and hasattr(rs.reserva.suscripcion_utilizada, 'plan'):
                        if rs.reserva.suscripcion_utilizada.plan and hasattr(rs.reserva.suscripcion_utilizada.plan, 'nombre'):
                            plan_info = f"{rs.reserva.suscripcion_utilizada.plan.nombre}"
                        else:
                            plan_info = "Plan Individual (Sin Nombre)"
                    
                    # Verificar si tiene suscripción empresarial
                    elif rs.reserva.suscripcion_empresarial and hasattr(rs.reserva.suscripcion_empresarial, 'plan_empresarial'):
                        if rs.reserva.suscripcion_empresarial.plan_empresarial and hasattr(rs.reserva.suscripcion_empresarial.plan_empresarial, 'nombre'):
                            plan_info = f"Plan Empresarial: {rs.reserva.suscripcion_empresarial.plan_empresarial.nombre}"
                        else:
                            plan_info = "Plan Empresarial (Sin Nombre)"
                    
                    # Si es servicio de plan pero no tiene suscripción asociada
                    elif rs.es_servicio_plan:
                        plan_info = "Servicio de Plan (Sin Suscripción Asociada)"
                    
                    else:
                        plan_info = "Servicio Sin Plan Definido"
                        
                except Exception as e:
                    # En caso de error, incluir información del error para depuración
                    plan_info = f"Error al obtener plan: {str(e)}"
                
                # Calcular valores
                precio_original = Decimal(str(rs.precio_original)) if rs.precio_original else Decimal('0.00')
                precio_aplicado = Decimal(str(rs.precio_aplicado)) if rs.precio_aplicado else Decimal('0.00')
                ahorro = precio_original - precio_aplicado
                porcentaje_descuento = 0
                if precio_original > 0:
                    porcentaje_descuento = round(float((ahorro / precio_original) * 100), 1)
                
                servicios_con_plan_detallados.append({
                    'nombre_servicio': rs.servicio.nombre_servicio,
                    'plan': plan_info,
                    'precio_original': precio_original,
                    'precio_con_descuento': precio_aplicado,
                    'ahorro': ahorro,
                    'porcentaje_descuento': porcentaje_descuento,
                    'reserva_id': rs.reserva.id_reserva
                })
            
            # Agrupar por servicio y plan para totales
            servicios_con_plan_agrupados = {}
            for servicio in servicios_con_plan_detallados:
                clave = f"{servicio['nombre_servicio']} - {servicio['plan']}"
                if clave not in servicios_con_plan_agrupados:
                    servicios_con_plan_agrupados[clave] = {
                        'nombre': servicio['nombre_servicio'],
                        'plan': servicio['plan'],
                        'cantidad': 0,
                        'total_valor_original': Decimal('0.00'),
                        'total_valor_con_descuento': Decimal('0.00'),
                        'total_ahorro': Decimal('0.00'),
                        'porcentaje_descuento_promedio': 0
                    }
                
                servicios_con_plan_agrupados[clave]['cantidad'] += 1
                servicios_con_plan_agrupados[clave]['total_valor_original'] += servicio['precio_original']
                servicios_con_plan_agrupados[clave]['total_valor_con_descuento'] += servicio['precio_con_descuento']
                servicios_con_plan_agrupados[clave]['total_ahorro'] += servicio['ahorro']
            
            # Calcular porcentajes promedio de descuento
            for servicio_data in servicios_con_plan_agrupados.values():
                if servicio_data['total_valor_original'] > 0:
                    porcentaje = (servicio_data['total_ahorro'] / servicio_data['total_valor_original']) * 100
                    servicio_data['porcentaje_descuento_promedio'] = round(float(porcentaje), 1)
            
            # Calcular totales del mes
            total_reservas_mes = reservas_mes.count()
            total_servicios_sin_plan_mes = sum([data['cantidad'] for data in servicios_sin_plan_agrupados.values()])
            total_servicios_con_plan_mes = sum([data['cantidad'] for data in servicios_con_plan_agrupados.values()])
            total_valor_sin_plan_mes = sum([data['total_valor'] for data in servicios_sin_plan_agrupados.values()])
            total_valor_con_plan_mes = sum([data['total_valor_con_descuento'] for data in servicios_con_plan_agrupados.values()])
            total_valor_original_con_plan_mes = sum([data['total_valor_original'] for data in servicios_con_plan_agrupados.values()])
            total_ahorro_con_plan_mes = sum([data['total_ahorro'] for data in servicios_con_plan_agrupados.values()])
            total_general_mes = total_valor_sin_plan_mes + total_valor_con_plan_mes
            
            # Guardar datos del mes (en español)
            meses_espanol = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
            nombre_mes = meses_espanol[mes_num]
            empresa_datos['datos_por_mes'][mes_num] = {
                'nombre_mes': nombre_mes,
                'total_reservas': total_reservas_mes,
                'servicios_sin_plan': list(servicios_sin_plan_agrupados.values()),
                'servicios_con_plan': list(servicios_con_plan_agrupados.values()),
                'total_servicios_sin_plan': total_servicios_sin_plan_mes,
                'total_servicios_con_plan': total_servicios_con_plan_mes,
                'total_valor_sin_plan': total_valor_sin_plan_mes,
                'total_valor_con_plan': total_valor_con_plan_mes,
                'total_valor_original_con_plan': total_valor_original_con_plan_mes,
                'total_ahorro_con_plan': total_ahorro_con_plan_mes,
                'total_general': total_general_mes
            }
            
            # Sumar a totales anuales
            empresa_datos['total_ano']['reservas'] += total_reservas_mes
            empresa_datos['total_ano']['servicios_sin_plan'] += total_servicios_sin_plan_mes
            empresa_datos['total_ano']['servicios_con_plan'] += total_servicios_con_plan_mes
            empresa_datos['total_ano']['valor_sin_plan'] += total_valor_sin_plan_mes
            empresa_datos['total_ano']['valor_con_plan'] += total_valor_con_plan_mes
            empresa_datos['total_ano']['valor_total'] += total_general_mes
        
        # Solo agregar empresas que tienen datos
        if empresa_datos['datos_por_mes']:
            empresas_data.append(empresa_datos)
    
    # Calcular totales globales
    total_empresas = len(empresas_data)
    total_reservas_global = sum([data['total_ano']['reservas'] for data in empresas_data])
    total_servicios_sin_plan_global = sum([data['total_ano']['servicios_sin_plan'] for data in empresas_data])
    total_servicios_con_plan_global = sum([data['total_ano']['servicios_con_plan'] for data in empresas_data])
    total_valor_global = sum([data['total_ano']['valor_total'] for data in empresas_data])
    
    # Generar lista de años disponibles
    anos_disponibles = list(range(2020, datetime.now().year + 2))
    
    # Lista de meses
    meses_disponibles = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    # Implementar paginación (2 empresas por página)
    paginator = Paginator(empresas_data, 2)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'empresas_data': page_obj,
        'page_obj': page_obj,
        'totales_globales': {
            'total_empresas': total_empresas,
            'total_reservas': total_reservas_global,
            'total_servicios_sin_plan': total_servicios_sin_plan_global,
            'total_servicios_con_plan': total_servicios_con_plan_global,
            'total_valor': total_valor_global
        },
        'filtros': {
            'mes_actual': int(mes_filtro) if mes_filtro else None,
            'ano_actual': ano_filtro,
            'meses_disponibles': meses_disponibles,
            'anos_disponibles': anos_disponibles
        }
    }
    
    return render(request, 'reservas/analisis_reservas_empresas.html', context)


# ======================================================================
# VISTAS PARA GESTIÓN DE PAGOS A EMPRESAS  
# ======================================================================

@admin_required
def gestion_pagos_empresas(request):
    """
    Vista principal para la gestión de pagos a empresas
    """
    # Obtener filtros
    empresa_filtro = request.GET.get('empresa', '')
    estado_filtro = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Nueva lógica: mostrar resumen de empresas con reservas pendientes (no pagadas)
    # Consideramos reservas completadas y que no estén marcadas como pagadas a la empresa
    empresas_qs = Empresa.objects.filter(is_active=True)

    # Si hay filtro por empresa, limitar
    if empresa_filtro:
        empresas_qs = empresas_qs.filter(id_empresa=empresa_filtro)

    empresas_pendientes = []
    empresas_pagadas = []

    # Construir lista de empresas con reservas pendientes
    for empresa in empresas_qs.order_by('nombre_empresa'):
        reservas_pendientes = Reserva.objects.filter(
            empresa=empresa,
            estado='completado',
        ).filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True))

        if reservas_pendientes.exists():
            total_pendiente = 0
            for r in reservas_pendientes:
                detalle = r.obtener_detalle_servicios()
                # Usar total_original y descontar el 12% de comisión de AutoNEW
                total_original = detalle.get('total_original', 0)
                pago_empresa = total_original * 0.88  # 88% para la empresa (100% - 12% comisión)
                total_pendiente += pago_empresa

            empresas_pendientes.append({
                'empresa': empresa,
                'reservas_pendientes': reservas_pendientes,
                'cantidad_pendientes': reservas_pendientes.count(),
                'total_pendiente': total_pendiente,
            })

    # Construir lista de empresas con reservas pagadas
    # Iteramos todas las empresas activas (no solo filtradas por pendiente)
    empresas_all = Empresa.objects.filter(is_active=True).order_by('nombre_empresa')
    if empresa_filtro:
        empresas_all = empresas_all.filter(id_empresa=empresa_filtro)

    for empresa in empresas_all:
        reservas_pagadas = Reserva.objects.filter(
            empresa=empresa,
            estado='completado',
            pagado_empresa=True
        )

        if reservas_pagadas.exists():
            total_pagado = 0
            for r in reservas_pagadas:
                detalle = r.obtener_detalle_servicios()
                # Usar total_original y descontar el 12% de comisión de AutoNEW
                total_original = detalle.get('total_original', 0)
                pago_empresa = total_original * 0.88  # 88% para la empresa (100% - 12% comisión)
                total_pagado += pago_empresa

            empresas_pagadas.append({
                'empresa': empresa,
                'reservas_pagadas': reservas_pagadas,
                'cantidad_pagadas': reservas_pagadas.count(),
                'total_pagado': total_pagado,
            })

    # Paginación para pendientes y pagadas (uso de parámetros independientes)
    pending_page_no = request.GET.get('pending_page')
    paid_page_no = request.GET.get('paid_page')

    pending_paginator = Paginator(empresas_pendientes, 20)
    paid_paginator = Paginator(empresas_pagadas, 20)

    pending_page_obj = pending_paginator.get_page(pending_page_no)
    paid_page_obj = paid_paginator.get_page(paid_page_no)

    stats = {
        'total_empresas_pendientes': len(empresas_pendientes),
        'total_empresas_pagadas': len(empresas_pagadas),
        'total_reservas_pendientes': sum(e['cantidad_pendientes'] for e in empresas_pendientes),
        'total_pendiente': sum(e['total_pendiente'] for e in empresas_pendientes),
        'total_reservas_pagadas': sum(e['cantidad_pagadas'] for e in empresas_pagadas),
        'total_pagado': sum(e['total_pagado'] for e in empresas_pagadas),
    }

    context = {
        'pending_page_obj': pending_page_obj,
        'paid_page_obj': paid_page_obj,
        'stats': stats,
        'filtros': {
            'empresa': empresa_filtro,
            'estado': estado_filtro,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
    }

    return render(request, 'admin/gestion_pagos_empresas.html', context)


@admin_required
def detalle_periodo_liquidacion(request, periodo_id):
    """
    Vista para ver el detalle de un período de liquidación específico
    """
    periodo = get_object_or_404(PeriodoLiquidacion, id_periodo=periodo_id)
    
    # Si el período está activo, calcular totales actualizados
    if periodo.estado == 'activo':
        periodo.calcular_totales()
    
    # Obtener reservas del período
    reservas = periodo.reservas_incluidas.all().select_related('usuario', 'empresa')
    
    # Si no hay reservas asignadas pero el período está activo, mostrar las potenciales
    if not reservas.exists() and periodo.estado == 'activo':
        reservas = Reserva.objects.filter(
            empresa=periodo.empresa,
            estado='completado',
            fecha__range=[periodo.fecha_inicio, periodo.fecha_fin]
        ).select_related('usuario', 'empresa')
    
    # Calcular detalles de cada reserva
    detalles_reservas = []
    for reserva in reservas:
        detalle_servicios = reserva.obtener_detalle_servicios()
        
        valor_bruto = detalle_servicios.get('total', 0)
        valor_descuentos = detalle_servicios.get('ahorro_total', 0)
        valor_neto = valor_bruto - valor_descuentos
        valor_comision = (valor_neto * periodo.comision_autonew) / 100
        valor_empresa = valor_neto - valor_comision
        
        detalles_reservas.append({
            'reserva': reserva,
            'detalle_servicios': detalle_servicios,
            'valor_bruto': valor_bruto,
            'valor_descuentos': valor_descuentos,
            'valor_neto': valor_neto,
            'valor_comision': valor_comision,
            'valor_empresa': valor_empresa,
        })
    
    context = {
        'periodo': periodo,
        'detalles_reservas': detalles_reservas,
        'puede_cerrar': periodo.estado == 'activo',
        'puede_pagar': periodo.estado == 'cerrado',
    }
    
    return render(request, 'admin/detalle_periodo_liquidacion.html', context)


@admin_required
def cerrar_periodo_liquidacion(request, periodo_id):
    """
    Vista para cerrar un período de liquidación
    """
    periodo = get_object_or_404(PeriodoLiquidacion, id_periodo=periodo_id)
    
    if request.method == 'POST':
        if periodo.cerrar_periodo(request.user):
            messages.success(request, f'Período de liquidación cerrado correctamente. Total a pagar: ${periodo.total_neto:,.0f}')
            
            # Crear detalles de liquidación
            for reserva in periodo.reservas_incluidas.all():
                detalle, created = DetalleLiquidacion.objects.get_or_create(
                    periodo=periodo,
                    reserva=reserva
                )
                if created:
                    detalle.calcular_valores()
            
        else:
            messages.error(request, 'No se pudo cerrar el período. Verifique el estado actual.')
        
        return redirect('detalle_periodo_liquidacion', periodo_id=periodo.id_periodo)
    
    return redirect('gestion_pagos_empresas')


@admin_required
def marcar_como_pagado(request, periodo_id):
    """
    Vista para marcar un período como pagado
    """
    periodo = get_object_or_404(PeriodoLiquidacion, id_periodo=periodo_id)
    
    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago', '')
        referencia_pago = request.POST.get('referencia_pago', '')
        observaciones = request.POST.get('observaciones', '')
        
        if metodo_pago:
            if periodo.marcar_como_pagado(request.user, metodo_pago, referencia_pago, observaciones):
                messages.success(request, f'Período marcado como pagado correctamente. Monto: ${periodo.total_neto:,.0f}')
            else:
                messages.error(request, 'No se pudo marcar como pagado. Verifique el estado del período.')
        else:
            messages.error(request, 'Debe seleccionar un método de pago.')
        
        return redirect('detalle_periodo_liquidacion', periodo_id=periodo.id_periodo)
    
    return redirect('gestion_pagos_empresas')


@admin_required
def generar_periodos_faltantes(request):
    """
    Vista para generar períodos de liquidación faltantes
    """
    if request.method == 'POST':
        periodos_creados = 0
        empresas_activas = Empresa.objects.filter(is_active=True)
        
        for empresa in empresas_activas:
            # Verificar si ya tiene un período activo
            periodo_activo = PeriodoLiquidacion.objects.filter(
                empresa=empresa,
                estado='activo'
            ).exists()
            
            if not periodo_activo:
                # Buscar el último período para esta empresa
                ultimo_periodo = PeriodoLiquidacion.objects.filter(
                    empresa=empresa
                ).order_by('-fecha_fin').first()
                
                if ultimo_periodo:
                    # Crear período desde donde terminó el último
                    inicio = ultimo_periodo.fecha_fin
                else:
                    # Si es la primera vez, empezar desde hace 15 días
                    inicio = timezone.now() - timedelta(days=15)
                
                inicio = inicio.replace(hour=0, minute=0, second=0, microsecond=0)
                fin = inicio + timedelta(days=15)
                
                # Si el período termina en el futuro, ajustar al presente
                if fin > timezone.now():
                    fin = timezone.now()
                
                if inicio < fin:  # Solo crear si hay un rango válido
                    PeriodoLiquidacion.objects.create(
                        empresa=empresa,
                        fecha_inicio=inicio,
                        fecha_fin=fin,
                        comision_autonew=15.0  # Comisión por defecto
                    )
                    periodos_creados += 1
        
        messages.success(request, f'Se crearon {periodos_creados} períodos de liquidación.')
        
    return redirect('gestion_pagos_empresas')


@admin_required
def dashboard_pagos(request):
    """
    Dashboard con estadísticas de pagos
    """
    # Estadísticas generales
    ahora = timezone.now()
    inicio_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    stats = {
        # Períodos
        'total_periodos_activos': PeriodoLiquidacion.objects.filter(estado='activo').count(),
        'total_periodos_cerrados': PeriodoLiquidacion.objects.filter(estado='cerrado').count(),
        'periodos_vencidos': PeriodoLiquidacion.objects.filter(
            estado='activo', fecha_fin__lt=ahora
        ).count(),
        
        # Montos
        'monto_pendiente_pago': PeriodoLiquidacion.objects.filter(
            estado='cerrado'
        ).aggregate(total=Sum('total_neto'))['total'] or 0,
        
        'monto_pagado_mes_actual': PeriodoLiquidacion.objects.filter(
            estado='pagado',
            fecha_pago__gte=inicio_mes
        ).aggregate(total=Sum('total_neto'))['total'] or 0,
        
        # Empresas
        'empresas_con_pendientes': PeriodoLiquidacion.objects.filter(
            estado='cerrado'
        ).values('empresa').distinct().count(),
        
        'reservas_sin_liquidar': Reserva.objects.filter(
            estado='completado',
            periodos_liquidacion__isnull=True,
            fecha__gte=ahora - timedelta(days=30)
        ).count(),
    }
    
    # Top 5 empresas por monto pendiente
    top_empresas_pendientes = PeriodoLiquidacion.objects.filter(
        estado='cerrado'
    ).values(
        'empresa__nombre_empresa',
        'empresa__id_empresa'
    ).annotate(
        total_pendiente=Sum('total_neto'),
        periodos_pendientes=Count('id_periodo')
    ).order_by('-total_pendiente')[:5]
    
    # Períodos por estado para gráfico
    periodos_por_estado = PeriodoLiquidacion.objects.values('estado').annotate(
        cantidad=Count('id_periodo'),
        monto_total=Sum('total_neto')
    ).order_by('estado')
    
    context = {
        'stats': stats,
        'top_empresas_pendientes': top_empresas_pendientes,
        'periodos_por_estado': periodos_por_estado,
    }
    
    return render(request, 'admin/dashboard_pagos.html', context)


@admin_required
def detalle_pagos_empresa(request, empresa_id):
    """
    Muestra el detalle de reservas no pagadas para una empresa
    """
    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)

    # Permitir mostrar reservas 'pendientes' (por defecto) o 'pagadas'
    show = request.GET.get('show', 'pendientes')

    if show == 'pagadas':
        reservas_qs = Reserva.objects.filter(
            empresa=empresa,
            estado='completado',
            pagado_empresa=True
        ).select_related('usuario')
    else:
        reservas_qs = Reserva.objects.filter(
            empresa=empresa,
            estado='completado',
        ).filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True)).select_related('usuario')

    detalles = []
    total = 0
    for r in reservas_qs:
        detalle_servicios = r.obtener_detalle_servicios()
        valor = detalle_servicios.get('total', 0) - detalle_servicios.get('ahorro_total', 0)
        total += valor
        detalles.append({
            'reserva': r,
            'detalle_servicios': detalle_servicios,
            'valor': valor,
        })

    context = {
        'empresa': empresa,
        'detalles': detalles,
        'total': total,
        'show': show,
    }

    return render(request, 'admin/detalle_pagos_empresa.html', context)


@admin_required
def exportar_periodo_csv(request, periodo_id):
    """
    Exportar detalles de un período a CSV
    """
    periodo = get_object_or_404(PeriodoLiquidacion, id_periodo=periodo_id)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="periodo_{periodo.empresa.nombre_empresa}_{periodo.fecha_inicio.strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Reserva ID', 'Fecha Servicio', 'Cliente', 'Servicios',
        'Valor Bruto', 'Descuentos', 'Valor Neto', 
        'Comisión %', 'Valor Comisión', 'Total Empresa'
    ])
    
    for reserva in periodo.reservas_incluidas.all():
        detalle = reserva.obtener_detalle_servicios()
        
        servicios_nombres = []
        for categoria in ['servicios_plan', 'servicios_adicionales', 'servicios_empresariales']:
            for servicio in detalle.get(categoria, []):
                servicios_nombres.append(servicio['nombre'])
        
        valor_bruto = detalle.get('total', 0)
        valor_descuentos = detalle.get('ahorro_total', 0)
        valor_neto = valor_bruto - valor_descuentos
        valor_comision = (valor_neto * periodo.comision_autonew) / 100
        valor_empresa = valor_neto - valor_comision
        
        writer.writerow([
            reserva.id_reserva,
            reserva.fecha.strftime('%Y-%m-%d'),
            reserva.usuario.nombre_completo,
            ', '.join(servicios_nombres),
            f'{valor_bruto:.2f}',
            f'{valor_descuentos:.2f}',
            f'{valor_neto:.2f}',
            f'{periodo.comision_autonew:.1f}%',
            f'{valor_comision:.2f}',
            f'{valor_empresa:.2f}',
        ])
    
    return response


@admin_required
def exportar_empresa_csv(request, empresa_id):
    """
    Exporta a CSV las reservas pendientes (no pagadas) de una empresa
    """
    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)

    reservas = Reserva.objects.filter(
        empresa=empresa,
        estado='completado',
    ).filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True)).select_related('usuario')

    response = HttpResponse(content_type='text/csv')
    filename = f'empresa_{empresa.nombre_empresa}_{timezone.now().strftime("%Y%m%d")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(['numero_reserva', 'fecha', 'hora', 'usuario', 'valor_bruto', 'descuentos', 'valor_neto', 'pagado_empresa'])

    for r in reservas:
        detalle = r.obtener_detalle_servicios()
        valor_bruto = detalle.get('total', 0)
        descuentos = detalle.get('ahorro_total', 0)
        valor_neto = valor_bruto - descuentos
        writer.writerow([
            r.numero_reserva or '',
            r.fecha.strftime('%Y-%m-%d') if r.fecha else '',
            r.hora.strftime('%H:%M:%S') if r.hora else '',
            getattr(r.usuario, 'nombre_usuario', ''),
            f'{valor_bruto:.2f}',
            f'{descuentos:.2f}',
            f'{valor_neto:.2f}',
            str(bool(r.pagado_empresa)),
        ])

    return response


@admin_required
def api_datos_grafico_pagos(request):
    """
    API para obtener datos para gráficos de pagos
    """
    tipo_grafico = request.GET.get('tipo', 'montos_mensuales')
    
    if tipo_grafico == 'montos_mensuales':
        # Últimos 6 meses de montos pagados
        datos = []
        for i in range(6):
            fecha = timezone.now() - timedelta(days=30*i)
            inicio_mes = fecha.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if fecha.month == 12:
                fin_mes = inicio_mes.replace(year=fecha.year + 1, month=1) - timedelta(seconds=1)
            else:
                fin_mes = inicio_mes.replace(month=fecha.month + 1) - timedelta(seconds=1)
            
            monto = PeriodoLiquidacion.objects.filter(
                estado='pagado',
                fecha_pago__range=[inicio_mes, fin_mes]
            ).aggregate(total=Sum('total_neto'))['total'] or 0
            
            datos.append({
                'mes': fecha.strftime('%B %Y'),
                'monto': float(monto)
            })
        
        return JsonResponse({'datos': list(reversed(datos))})
    
    elif tipo_grafico == 'empresas_top':
        # Top 10 empresas por monto total pagado
        datos = PeriodoLiquidacion.objects.filter(
            estado='pagado'
        ).values(
            'empresa__nombre_empresa'
        ).annotate(
            total_pagado=Sum('total_neto')
        ).order_by('-total_pagado')[:10]
        
        return JsonResponse({'datos': list(datos)})
    
    return JsonResponse({'error': 'Tipo de gráfico no válido'}, status=400)


@admin_required
def marcar_reserva_pagada(request, reserva_id):
    """
    Vista para marcar una reserva individual como pagada a la empresa
    """
    if request.method == 'POST':
        try:
            reserva = get_object_or_404(Reserva, id_reserva=reserva_id)
            
            # Verificar que la reserva esté completada
            if reserva.estado != 'completado':
                return JsonResponse({
                    'success': False,
                    'error': 'Solo se pueden marcar como pagadas las reservas completadas'
                }, status=400)
            
            # Verificar que no esté ya marcada como pagada
            if reserva.pagado_empresa:
                return JsonResponse({
                    'success': False,
                    'error': 'Esta reserva ya está marcada como pagada'
                }, status=400)
            
            # Marcar como pagada
            reserva.pagado_empresa = True
            reserva.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Reserva {reserva.numero_reserva} marcada como pagada exitosamente',
                'reserva_id': reserva.id_reserva,
                'numero_reserva': reserva.numero_reserva
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@admin_required
def marcar_empresa_pagada(request, empresa_id):
    """
    Vista para marcar todas las reservas pendientes de una empresa como pagadas
    """
    if request.method == 'POST':
        try:
            empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
            
            # Obtener todas las reservas pendientes de pago
            reservas_pendientes = Reserva.objects.filter(
                empresa=empresa,
                estado='completado',
                pagado_empresa=False
            )
            
            cantidad_inicial = reservas_pendientes.count()
            
            if cantidad_inicial == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No hay reservas pendientes para esta empresa'
                }, status=400)
            
            # Marcar todas como pagadas
            reservas_actualizadas = reservas_pendientes.update(pagado_empresa=True)
            
            return JsonResponse({
                'success': True,
                'message': f'{reservas_actualizadas} reservas de {empresa.nombre_empresa} marcadas como pagadas',
                'cantidad_marcadas': reservas_actualizadas,
                'empresa': empresa.nombre_empresa
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@admin_required
def marcar_reservas_seleccionadas(request):
    """
    Vista para marcar múltiples reservas seleccionadas como pagadas
    """
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            reservas_ids = data.get('reservas_ids', [])
            
            if not reservas_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'No se enviaron reservas para marcar como pagadas'
                }, status=400)
            
            # Validar que todas las reservas existan y estén completadas
            reservas = Reserva.objects.filter(
                id_reserva__in=reservas_ids,
                estado='completado',
                pagado_empresa=False
            )
            
            cantidad_encontradas = reservas.count()
            
            if cantidad_encontradas == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No se encontraron reservas válidas para marcar como pagadas'
                }, status=400)
            
            if cantidad_encontradas != len(reservas_ids):
                return JsonResponse({
                    'success': False,
                    'error': f'Solo {cantidad_encontradas} de {len(reservas_ids)} reservas son válidas para marcar como pagadas'
                }, status=400)
            
            # Marcar todas como pagadas
            cantidad_actualizadas = reservas.update(pagado_empresa=True)
            
            return JsonResponse({
                'success': True,
                'message': f'{cantidad_actualizadas} reservas marcadas como pagadas exitosamente',
                'cantidad_marcadas': cantidad_actualizadas
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Error al decodificar los datos JSON'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


@admin_required
def exportar_pagos_excel(request):
    """
    Vista para exportar el reporte de pagos a Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen General
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws_resumen['A1'] = 'REPORTE DE PAGOS A EMPRESAS'
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen.merge_cells('A1:F1')
        
        ws_resumen['A2'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen.merge_cells('A2:F2')
        
        # Estadísticas
        empresas_qs = Empresa.objects.filter(is_active=True)
        
        total_pendiente = 0
        total_pagado = 0
        total_reservas_pendientes = 0
        total_reservas_pagadas = 0
        
        for empresa in empresas_qs:
            # Pendientes
            reservas_pend = Reserva.objects.filter(
                empresa=empresa,
                estado='completado',
                pagado_empresa=False
            )
            
            for r in reservas_pend:
                detalle = r.obtener_detalle_servicios()
                total_pendiente += detalle.get('total', 0)
            total_reservas_pendientes += reservas_pend.count()
            
            # Pagadas
            reservas_pag = Reserva.objects.filter(
                empresa=empresa,
                estado='completado',
                pagado_empresa=True
            )
            
            for r in reservas_pag:
                detalle = r.obtener_detalle_servicios()
                total_pagado += detalle.get('total', 0)
            total_reservas_pagadas += reservas_pag.count()
        
        # Escribir estadísticas
        ws_resumen['A4'] = 'ESTADÍSTICAS GENERALES'
        ws_resumen['A4'].font = Font(bold=True, size=14)
        
        stats_data = [
            ['Total Pendiente de Pago:', f'${total_pendiente:,.0f}'],
            ['Total Pagado:', f'${total_pagado:,.0f}'],
            ['Reservas Pendientes:', total_reservas_pendientes],
            ['Reservas Pagadas:', total_reservas_pagadas],
        ]
        
        row = 6
        for label, value in stats_data:
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = Font(bold=True)
            ws_resumen[f'B{row}'] = value
            row += 1
        
        # Hoja 2: Reservas Pendientes
        ws_pendientes = wb.create_sheet("Pendientes de Pago")
        
        headers_pendientes = [
            'Empresa', 'Número Reserva', 'Fecha', 'Hora', 'Cliente', 
            'Email Cliente', 'Servicios', 'Tipo', 'Total', 'Estado Pago'
        ]
        
        for col, header in enumerate(headers_pendientes, 1):
            cell = ws_pendientes.cell(1, col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 2
        for empresa in empresas_qs:
            reservas_pendientes = Reserva.objects.filter(
                empresa=empresa,
                estado='completado',
                pagado_empresa=False
            ).select_related('usuario')
            
            for reserva in reservas_pendientes:
                detalle = reserva.obtener_detalle_servicios()
                servicios = ', '.join([s.nombre_servicio for s in reserva.servicios.all()])
                
                tipo_reserva = 'Individual'
                if reserva.es_reserva_empresarial:
                    tipo_reserva = 'Empresarial'
                elif reserva.suscripcion_utilizada:
                    tipo_reserva = 'Plan'
                
                data = [
                    empresa.nombre_empresa,
                    reserva.numero_reserva,
                    reserva.fecha.strftime('%d/%m/%Y'),
                    str(reserva.hora),
                    reserva.usuario.nombre_completo,
                    reserva.usuario.correo,
                    servicios,
                    tipo_reserva,
                    detalle.get('total', 0),
                    'PENDIENTE'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws_pendientes.cell(row, col)
                    cell.value = value
                    cell.border = border
                    if col == 9:  # Columna de total
                        cell.number_format = '"$"#,##0.00'
                
                row += 1
        
        # Ajustar anchos de columna
        for col in range(1, len(headers_pendientes) + 1):
            ws_pendientes.column_dimensions[get_column_letter(col)].width = 15
        
        # Hoja 3: Reservas Pagadas
        ws_pagadas = wb.create_sheet("Pagadas")
        
        for col, header in enumerate(headers_pendientes, 1):
            cell = ws_pagadas.cell(1, col)
            cell.value = header
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 2
        for empresa in empresas_qs:
            reservas_pagadas = Reserva.objects.filter(
                empresa=empresa,
                estado='completado',
                pagado_empresa=True
            ).select_related('usuario')
            
            for reserva in reservas_pagadas:
                detalle = reserva.obtener_detalle_servicios()
                servicios = ', '.join([s.nombre_servicio for s in reserva.servicios.all()])
                
                tipo_reserva = 'Individual'
                if reserva.es_reserva_empresarial:
                    tipo_reserva = 'Empresarial'
                elif reserva.suscripcion_utilizada:
                    tipo_reserva = 'Plan'
                
                data = [
                    empresa.nombre_empresa,
                    reserva.numero_reserva,
                    reserva.fecha.strftime('%d/%m/%Y'),
                    str(reserva.hora),
                    reserva.usuario.nombre_completo,
                    reserva.usuario.correo,
                    servicios,
                    tipo_reserva,
                    detalle.get('total', 0),
                    'PAGADA'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws_pagadas.cell(row, col)
                    cell.value = value
                    cell.border = border
                    if col == 9:  # Columna de total
                        cell.number_format = '"$"#,##0.00'
                
                row += 1
        
        # Ajustar anchos de columna
        for col in range(1, len(headers_pendientes) + 1):
            ws_pagadas.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Reporte_Pagos_Empresas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar el archivo Excel: {str(e)}')
        return redirect('gestion_pagos_empresas')

@empresa_required
def empresa_mis_pagos(request):
    """
    Vista para que las empresas vean sus pagos de reservas
    """
    empresa_id = request.session.get('empresa_id')
    empresa = get_object_or_404(Empresa, id_empresa=empresa_id)
    
    # Obtener filtros
    estado_filtro = request.GET.get('estado', 'pendientes')  # pendientes o pagadas
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Query base - solo reservas completadas de esta empresa
    reservas_base = Reserva.objects.filter(
        empresa=empresa,
        estado='completado'
    ).select_related('usuario').prefetch_related('servicios')
    
    # Aplicar filtro de fechas
    if fecha_desde:
        reservas_base = reservas_base.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        reservas_base = reservas_base.filter(fecha__lte=fecha_hasta)
    
    # Separar entre pendientes y pagadas
    if estado_filtro == 'pagadas':
        reservas = reservas_base.filter(pagado_empresa=True).order_by('-fecha', '-hora')
    else:
        reservas = reservas_base.filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True)).order_by('-fecha', '-hora')
    
    # Calcular totales
    total_pendiente = 0
    total_pagado = 0
    
    for r in reservas_base.filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True)):
        detalle = r.obtener_detalle_servicios()
        total_original = detalle.get('total_original', 0)
        pago_empresa = total_original * 0.88  # 88% para la empresa
        total_pendiente += pago_empresa
    
    for r in reservas_base.filter(pagado_empresa=True):
        detalle = r.obtener_detalle_servicios()
        total_original = detalle.get('total_original', 0)
        pago_empresa = total_original * 0.88  # 88% para la empresa
        total_pagado += pago_empresa
    
    # Paginaci�n
    paginator = Paginator(reservas, 20)
    page = request.GET.get('page')
    reservas_page = paginator.get_page(page)
    
    # Estad�sticas
    stats = {
        'total_reservas_pendientes': reservas_base.filter(Q(pagado_empresa=False) | Q(pagado_empresa__isnull=True)).count(),
        'total_pendiente': total_pendiente,
        'total_reservas_pagadas': reservas_base.filter(pagado_empresa=True).count(),
        'total_pagado': total_pagado,
    }
    
    context = {
        'empresa': empresa,
        'reservas': reservas_page,
        'stats': stats,
        'estado_filtro': estado_filtro,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
    }
    
    return render(request, 'empresas/mis_pagos.html', context)


# ==================== VISTAS PARA PÁGINAS LEGALES (LEY 1581 DE 2012) ====================

def aviso_privacidad(request):
    """Vista para mostrar el Aviso de Privacidad conforme a la Ley 1581 de 2012"""
    
    contenido_html = """
    <div class="space-y-6">
        <h2>1. Responsable del Tratamiento de Datos</h2>
        <p>
            <strong>AUTONEW</strong> es el responsable del tratamiento de sus datos personales. Nos comprometemos a proteger 
            su información personal y garantizar sus derechos conforme a la <strong>Ley 1581 de 2012</strong> y sus decretos 
            reglamentarios.
        </p>
        
        <h2>2. Datos que Recopilamos</h2>
        <p>Para brindarle nuestros servicios de gestión de lavado de vehículos, recopilamos los siguientes tipos de información:</p>
        <ul>
            <li><strong>Datos de identificación:</strong> Nombre completo, número de documento de identidad</li>
            <li><strong>Datos de contacto:</strong> Correo electrónico, número de teléfono, dirección</li>
            <li><strong>Datos del vehículo:</strong> Placa, marca, modelo, tipo de vehículo</li>
            <li><strong>Datos de servicio:</strong> Historial de reservas, servicios contratados, preferencias</li>
            <li><strong>Datos de pago:</strong> Información necesaria para procesar transacciones (no almacenamos datos completos de tarjetas)</li>
        </ul>

        <h2>3. Finalidad del Tratamiento</h2>
        <p>Sus datos personales serán utilizados para:</p>
        <ul>
            <li>Gestionar y procesar sus reservas de servicios de lavado</li>
            <li>Enviar confirmaciones, recordatorios y notificaciones sobre sus citas</li>
            <li>Procesar pagos y generar facturas</li>
            <li>Mejorar nuestros servicios y experiencia de usuario</li>
            <li>Enviar información promocional (solo con su consentimiento)</li>
            <li>Cumplir con obligaciones legales y regulatorias</li>
            <li>Gestionar solicitudes, quejas y reclamos</li>
        </ul>

        <h2>4. Sus Derechos como Titular</h2>
        <p>De acuerdo con la Ley 1581 de 2012, usted tiene derecho a:</p>
        <ul>
            <li><strong>Conocer, actualizar y rectificar</strong> sus datos personales</li>
            <li><strong>Solicitar prueba</strong> de la autorización otorgada</li>
            <li><strong>Ser informado</strong> sobre el uso de sus datos</li>
            <li><strong>Revocar la autorización</strong> y solicitar la supresión de datos (excepto cuando exista obligación legal)</li>
            <li><strong>Acceder gratuitamente</strong> a sus datos personales</li>
        </ul>

        <h2>5. Seguridad de la Información</h2>
        <p>
            Implementamos medidas técnicas, humanas y administrativas para proteger sus datos personales contra pérdida, 
            uso no autorizado, acceso, divulgación, alteración o destrucción.
        </p>

        <h2>6. Contacto</h2>
        <p>Para ejercer sus derechos o realizar consultas sobre el tratamiento de sus datos, puede contactarnos:</p>
        <ul>
            <li><strong>Email:</strong> privacidad@autonew.com</li>
            <li><strong>Teléfono:</strong> +57 300 123 4567</li>
        </ul>
    </div>
    """
    
    context = {
        'titulo': 'Aviso de Privacidad',
        'contenido': contenido_html,
        'fecha_actualizacion': '12 de noviembre de 2025'
    }
    
    return render(request, 'legal/documento_legal.html', context)


def politica_tratamiento_datos(request):
    """Vista para mostrar la Política de Tratamiento de Datos Personales"""
    
    contenido_html = """
    <div class="space-y-6">
        <h2>1. Identificación del Responsable</h2>
        <p>
            <strong>AUTONEW</strong> actúa como responsable del tratamiento de datos personales, comprometiéndose a cumplir 
            con la normativa vigente en Colombia, especialmente la <strong>Ley Estatutaria 1581 de 2012</strong>, 
            el <strong>Decreto 1377 de 2013</strong> y demás normas complementarias.
        </p>

        <h2>2. Tratamiento y Finalidades</h2>
        <p>Los datos personales recopilados serán tratados con las siguientes finalidades específicas:</p>
        
        <h3>2.1. Para Usuarios</h3>
        <ul>
            <li>Crear y gestionar su cuenta de usuario en la plataforma</li>
            <li>Procesar reservas de servicios de lavado y mantenimiento vehicular</li>
            <li>Enviar confirmaciones, recordatorios y actualizaciones sobre sus reservas</li>
            <li>Gestionar programas de fidelización y suscripciones</li>
            <li>Procesar pagos y generar comprobantes</li>
            <li>Proveer soporte técnico y atención al cliente</li>
            <li>Realizar encuestas de satisfacción y mejora de servicio</li>
        </ul>

        <h3>2.2. Para Empresas Aliadas</h3>
        <ul>
            <li>Gestionar la relación comercial y el registro en nuestra plataforma</li>
            <li>Coordinar la prestación de servicios a clientes</li>
            <li>Procesar pagos y liquidaciones</li>
            <li>Comunicar actualizaciones del sistema y políticas</li>
        </ul>

        <h2>3. Derechos de los Titulares</h2>
        <p>Como titular de datos personales, usted cuenta con los siguientes derechos:</p>
        <ul>
            <li><strong>Acceso:</strong> Conocer qué datos suyos tenemos almacenados</li>
            <li><strong>Actualización:</strong> Solicitar la corrección de datos inexactos o incompletos</li>
            <li><strong>Rectificación:</strong> Modificar datos que considere incorrectos</li>
            <li><strong>Supresión:</strong> Solicitar la eliminación de sus datos (sujeto a obligaciones legales)</li>
            <li><strong>Revocación:</strong> Retirar el consentimiento otorgado para el tratamiento</li>
        </ul>

        <h2>4. Procedimiento para Ejercer sus Derechos</h2>
        <p>Para ejercer cualquiera de sus derechos, debe enviar una solicitud mediante:</p>
        <ul>
            <li><strong>Correo electrónico:</strong> privacidad@autonew.com</li>
            <li><strong>Formulario web:</strong> Desde su perfil de usuario en la sección "Privacidad"</li>
        </ul>
        <p>
            Responderemos su solicitud en un plazo máximo de <strong>15 días hábiles</strong> contados desde la fecha 
            de recepción. Si requerimos más información, se lo notificaremos dentro de los primeros 5 días.
        </p>

        <h2>5. Medidas de Seguridad</h2>
        <p>AUTONEW implementa las siguientes medidas para proteger sus datos:</p>
        <ul>
            <li>Cifrado SSL/TLS en todas las comunicaciones</li>
            <li>Autenticación de dos factores disponible</li>
            <li>Controles de acceso basados en roles</li>
            <li>Auditorías periódicas de seguridad</li>
            <li>Respaldo regular de información</li>
            <li>Capacitación continua del personal en protección de datos</li>
        </ul>

        <h2>6. Almacenamiento y Conservación</h2>
        <p>
            Sus datos personales serán conservados durante el tiempo necesario para cumplir con las finalidades 
            descritas y las obligaciones legales aplicables. Posteriormente, serán eliminados de forma segura 
            o anonimizados para fines estadísticos.
        </p>

        <h2>7. Transferencias Internacionales</h2>
        <p>
            En caso de requerir transferir datos personales a otros países, garantizamos que dichos países cuenten 
            con niveles adecuados de protección de datos o, en su defecto, aplicaremos las salvaguardas apropiadas 
            conforme a la normativa colombiana.
        </p>

        <h2>8. Modificaciones a esta Política</h2>
        <p>
            Nos reservamos el derecho de modificar esta política en cualquier momento. Los cambios significativos 
            serán notificados a través de nuestra plataforma y correo electrónico con al menos 10 días de anticipación.
        </p>

        <h2>9. Autoridad de Control</h2>
        <p>
            La Superintendencia de Industria y Comercio es la autoridad competente para conocer de las reclamaciones 
            relacionadas con el tratamiento de datos personales en Colombia.
        </p>

        <h2>10. Contacto del Área de Privacidad</h2>
        <p>
            <strong>Email:</strong> privacidad@autonew.com<br>
            <strong>Teléfono:</strong> +57 300 123 4567<br>
            <strong>Horario de atención:</strong> Lunes a Viernes, 8:00 AM - 6:00 PM
        </p>
    </div>
    """
    
    context = {
        'titulo': 'Política de Tratamiento de Datos Personales',
        'contenido': contenido_html,
        'fecha_actualizacion': '12 de noviembre de 2025'
    }
    
    return render(request, 'legal/documento_legal.html', context)


def terminos_condiciones(request):
    """Vista para mostrar los Términos y Condiciones de Uso"""
    
    contenido_html = """
    <div class="space-y-6">
        <h2>1. Aceptación de los Términos</h2>
        <p>
            Al acceder y utilizar la plataforma AUTONEW, usted acepta estar sujeto a estos Términos y Condiciones. 
            Si no está de acuerdo con alguno de estos términos, por favor absténgase de utilizar nuestros servicios.
        </p>

        <h2>2. Descripción del Servicio</h2>
        <p>
            AUTONEW es una plataforma digital que conecta usuarios con empresas prestadoras de servicios de lavado 
            y mantenimiento vehicular. Facilitamos:
        </p>
        <ul>
            <li>Reserva en línea de servicios de lavado</li>
            <li>Gestión de citas y horarios</li>
            <li>Procesamiento de pagos</li>
            <li>Programas de fidelización y suscripciones</li>
            <li>Comunicación entre usuarios y empresas aliadas</li>
        </ul>

        <h2>3. Registro y Cuenta de Usuario</h2>
        <h3>3.1. Requisitos</h3>
        <ul>
            <li>Ser mayor de 18 años o contar con autorización de un representante legal</li>
            <li>Proporcionar información veraz, precisa y actualizada</li>
            <li>Mantener la confidencialidad de sus credenciales de acceso</li>
        </ul>

        <h3>3.2. Responsabilidades</h3>
        <p>Usted es responsable de todas las actividades realizadas bajo su cuenta. Debe notificarnos inmediatamente 
        cualquier uso no autorizado.</p>

        <h2>4. Reservas y Cancelaciones</h2>
        <h3>4.1. Proceso de Reserva</h3>
        <ul>
            <li>Las reservas quedan confirmadas una vez completado el proceso en línea</li>
            <li>Recibirá una confirmación por correo electrónico</li>
            <li>Debe llegar 10 minutos antes de la hora programada</li>
        </ul>

        <h3>4.2. Política de Cancelación</h3>
        <ul>
            <li>Cancelaciones gratuitas con <strong>12 horas de anticipación</strong></li>
            <li>Cancelaciones con menos de 12 horas pueden generar un cargo del 50% del servicio</li>
            <li>No presentarse sin cancelar puede resultar en el cobro total del servicio</li>
        </ul>

        <h2>5. Pagos y Facturación</h2>
        <h3>5.1. Métodos de Pago</h3>
        <p>Aceptamos:</p>
        <ul>
            <li>Tarjetas de crédito y débito</li>
            <li>Transferencias bancarias</li>
            <li>Pagos en efectivo (directamente en el establecimiento)</li>
        </ul>

        <h3>5.2. Precios</h3>
        <ul>
            <li>Los precios mostrados están en pesos colombianos (COP)</li>
            <li>Incluyen IVA cuando aplique</li>
            <li>Pueden estar sujetos a cambios sin previo aviso</li>
        </ul>

        <h2>6. Planes de Suscripción</h2>
        <h3>6.1. Renovación Automática</h3>
        <p>
            Las suscripciones se renuevan automáticamente al final de cada período, a menos que cancele con 
            al menos 24 horas de anticipación.
        </p>

        <h3>6.2. Cancelación de Suscripción</h3>
        <ul>
            <li>Puede cancelar en cualquier momento desde su perfil</li>
            <li>La cancelación será efectiva al final del período actual</li>
            <li>No se realizan reembolsos por períodos parciales</li>
        </ul>

        <h2>7. Uso Aceptable</h2>
        <p>El usuario se compromete a NO:</p>
        <ul>
            <li>Usar la plataforma para actividades ilegales o fraudulentas</li>
            <li>Compartir su cuenta con terceros</li>
            <li>Intentar acceder sin autorización a otras cuentas</li>
            <li>Realizar reservas falsas o maliciosas</li>
            <li>Difamar, acosar o amenazar a otros usuarios o empresas</li>
            <li>Enviar spam o contenido publicitario no autorizado</li>
        </ul>

        <h2>8. Propiedad Intelectual</h2>
        <p>
            Todo el contenido de la plataforma (textos, imágenes, logotipos, software) es propiedad de AUTONEW 
            o de sus licenciantes. Está prohibida su reproducción sin autorización expresa.
        </p>

        <h2>9. Limitación de Responsabilidad</h2>
        <p>AUTONEW actúa como intermediario entre usuarios y empresas prestadoras de servicios. Por lo tanto:</p>
        <ul>
            <li>No somos responsables de la calidad del servicio prestado por terceros</li>
            <li>No garantizamos la disponibilidad continua de la plataforma (aunque nos esforzamos por ello)</li>
            <li>No nos hacemos responsables de daños indirectos o consecuentes</li>
        </ul>

        <h2>10. Quejas y Reclamos</h2>
        <p>Para presentar quejas o reclamos:</p>
        <ul>
            <li><strong>Email:</strong> soporte@autonew.com</li>
            <li><strong>Formulario web:</strong> Sección "Contacto"</li>
            <li><strong>Teléfono:</strong> +57 300 123 4567</li>
        </ul>
        <p>Responderemos en un plazo máximo de 15 días hábiles.</p>

        <h2>11. Modificaciones</h2>
        <p>
            Nos reservamos el derecho de modificar estos términos en cualquier momento. Los cambios importantes 
            serán notificados con al menos 30 días de anticipación.
        </p>

        <h2>12. Jurisdicción y Ley Aplicable</h2>
        <p>
            Estos términos se rigen por las leyes de la República de Colombia. Cualquier controversia será resuelta 
            en los tribunales competentes de Colombia.
        </p>

        <h2>13. Contacto</h2>
        <p>
            <strong>AUTONEW</strong><br>
            Email: legal@autonew.com<br>
            Teléfono: +57 300 123 4567<br>
            Ubicación: Colombia
        </p>
    </div>
    """
    
    context = {
        'titulo': 'Términos y Condiciones de Uso',
        'contenido': contenido_html,
        'fecha_actualizacion': '12 de noviembre de 2025'
    }
    
    return render(request, 'legal/documento_legal.html', context)


def politica_cookies(request):
    """Vista para mostrar la Política de Cookies"""
    
    contenido_html = """
    <div class="space-y-6">
        <h2>1. ¿Qué son las Cookies?</h2>
        <p>
            Las cookies son pequeños archivos de texto que se almacenan en su dispositivo cuando visita nuestro sitio web. 
            Nos ayudan a mejorar su experiencia, recordar sus preferencias y analizar el uso de nuestra plataforma.
        </p>

        <h2>2. Tipos de Cookies que Utilizamos</h2>
        
        <h3>2.1. Cookies Esenciales (Obligatorias)</h3>
        <p>Son necesarias para el funcionamiento básico de la plataforma:</p>
        <ul>
            <li><strong>Sesión de usuario:</strong> Mantienen su sesión activa mientras navega</li>
            <li><strong>Seguridad:</strong> Protegen contra ataques y accesos no autorizados</li>
            <li><strong>Carrito de servicios:</strong> Recuerdan los servicios que ha seleccionado</li>
        </ul>

        <h3>2.2. Cookies de Funcionalidad (Opcionales)</h3>
        <p>Mejoran la funcionalidad y personalización:</p>
        <ul>
            <li><strong>Preferencias de idioma:</strong> Recuerdan su idioma preferido</li>
            <li><strong>Ubicación:</strong> Ayudan a mostrar empresas cercanas</li>
            <li><strong>Configuración de interfaz:</strong> Guardan preferencias de visualización</li>
        </ul>

        <h3>2.3. Cookies Analíticas (Opcionales)</h3>
        <p>Nos ayudan a entender cómo usan los visitantes nuestra plataforma:</p>
        <ul>
            <li>Páginas más visitadas</li>
            <li>Tiempo de permanencia</li>
            <li>Rutas de navegación</li>
            <li>Dispositivos utilizados</li>
        </ul>

        <h3>2.4. Cookies de Marketing (Opcionales)</h3>
        <p>Permiten mostrar publicidad relevante:</p>
        <ul>
            <li>Rastrean visitas a través de sitios web</li>
            <li>Muestran anuncios personalizados</li>
            <li>Miden efectividad de campañas</li>
        </ul>

        <h2>3. Duración de las Cookies</h2>
        <table>
            <thead>
                <tr>
                    <th>Tipo</th>
                    <th>Duración</th>
                    <th>Descripción</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>Sesión</td>
                    <td>Hasta cerrar navegador</td>
                    <td>Se eliminan automáticamente al cerrar el navegador</td>
                </tr>
                <tr>
                    <td>Persistentes</td>
                    <td>30 días - 1 año</td>
                    <td>Permanecen después de cerrar el navegador</td>
                </tr>
                <tr>
                    <td>Autenticación</td>
                    <td>7-30 días</td>
                    <td>Mantienen sesión activa si selecciona "Recordarme"</td>
                </tr>
            </tbody>
        </table>

        <h2>4. Control de Cookies</h2>
        <h3>4.1. En Nuestra Plataforma</h3>
        <p>
            Puede gestionar sus preferencias de cookies desde el <strong>Centro de Preferencias</strong> disponible 
            en el pie de página o desde su perfil de usuario.
        </p>

        <h3>4.2. En su Navegador</h3>
        <p>Puede configurar su navegador para:</p>
        <ul>
            <li>Rechazar todas las cookies</li>
            <li>Aceptar solo cookies de sitios específicos</li>
            <li>Recibir notificación antes de aceptar cookies</li>
            <li>Eliminar cookies existentes</li>
        </ul>

        <h2>5. Cookies de Terceros</h2>
        <p>Utilizamos servicios de terceros que pueden establecer sus propias cookies:</p>
        <ul>
            <li><strong>Google Analytics:</strong> Para análisis de tráfico</li>
            <li><strong>Pasarelas de pago:</strong> Para procesar transacciones seguras</li>
            <li><strong>Redes sociales:</strong> Para funciones de compartir contenido</li>
        </ul>

        <h2>6. Impacto de Deshabilitar Cookies</h2>
        <p>Si deshabilita las cookies, algunas funcionalidades pueden verse afectadas:</p>
        <ul>
            <li>No podrá mantener sesión iniciada</li>
            <li>No podrá completar reservas</li>
            <li>Las preferencias no se guardarán</li>
            <li>La experiencia de usuario será limitada</li>
        </ul>

        <h2>7. Actualización de esta Política</h2>
        <p>
            Esta política puede actualizarse periódicamente. La fecha de última actualización se muestra al inicio 
            del documento. Cambios significativos serán notificados a través de nuestra plataforma.
        </p>

        <h2>8. Más Información</h2>
        <p>Para consultas sobre nuestra política de cookies:</p>
        <ul>
            <li><strong>Email:</strong> privacidad@autonew.com</li>
            <li><strong>Teléfono:</strong> +57 300 123 4567</li>
        </ul>
    </div>
    """
    
    context = {
        'titulo': 'Política de Cookies',
        'contenido': contenido_html,
        'fecha_actualizacion': '12 de noviembre de 2025'
    }
    
    return render(request, 'legal/documento_legal.html', context)
